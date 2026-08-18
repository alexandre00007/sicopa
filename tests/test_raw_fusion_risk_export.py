from openpyxl import load_workbook

from controle_paie.database import Database
from controle_paie.raw_fusion_complete_export import CompleteExportRawFusionService


def _add_payroll(con, line_id, execution_id, regime, matricule, nom, amount, line_no):
    normalized_mat = matricule if matricule not in (None, '', 'NU') else ('NU' if matricule == 'NU' else '')
    normalized_name = nom.upper()
    con.execute("""INSERT INTO paie_standardisee
        (ligne_paie_id,execution_id,institution_id,regime,trimestre,annee,table_source,
         matricule_source,matricule_normalise,nom,prenom,nom_normalise,section,categorie,grade,
         unite_affectation,province,remuneration_base,transport,prime,logement,pension_rente,
         autres_remunerations,retenues,montant_net,remuneration_brute_calculee,ligne_source)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        [line_id, execution_id, 'i1', regime, 'T1', 2026, 'A' if execution_id == 'e1' else 'B',
         matricule, normalized_mat, nom, '', normalized_name, 'SEC', 'CAT', 'GR', 'UNIT', 'PROV',
         amount, 0, 0, 0, 0, 0, 0, amount, amount, line_no])


def _seed(db):
    with db.connect() as con:
        con.execute("CREATE TABLE raw_a (execution_id VARCHAR, matricule VARCHAR, nom VARCHAR, montant DOUBLE)")
        con.execute("CREATE TABLE raw_b (execution_id VARCHAR, matricule VARCHAR, nom VARCHAR, montant DOUBLE)")
        con.execute("INSERT INTO raw_a VALUES ('e1','001','Agent Risque',100),('e1','900','Agent Sain',80),('e1','NU','Sans Matricule',70)")
        con.execute("INSERT INTO raw_b VALUES ('e2','001','Agent Risque',120)")
        con.execute("""INSERT INTO journal_executions
            (execution_id,type_operation,fichier_source,table_source,table_destination,institution_id,regime,
             trimestre,annee,mode_chargement,lignes_lues,lignes_chargees,statut,date_fin)
            VALUES
            ('e1','IMPORT_ACCESS','a.accdb','A','raw_a','i1','CNSS','T1',2026,'append',3,3,'TERMINE',CURRENT_TIMESTAMP),
            ('e2','IMPORT_ACCESS','b.accdb','B','raw_b','i1','CARC','T1',2026,'append',1,1,'TERMINE',CURRENT_TIMESTAMP)
        """)
        _add_payroll(con, 'p1', 'e1', 'CNSS', '001', 'Agent Risque', 100, 1)
        _add_payroll(con, 'p2', 'e2', 'CARC', '001', 'Agent Risque', 120, 1)
        _add_payroll(con, 'p3', 'e1', 'CNSS', '900', 'Agent Sain', 80, 2)
        _add_payroll(con, 'p4', 'e1', 'CNSS', 'NU', 'Sans Matricule', 70, 3)


def _all_rows(workbook, prefix):
    rows = []
    for ws in workbook.worksheets:
        if ws.title.startswith(prefix):
            rows.extend(list(ws.iter_rows(min_row=2, values_only=True)))
    return rows


def test_annexe_12_excludes_healthy_single_regime_and_keeps_risks(tmp_path):
    db = Database(tmp_path / 'risk_export.duckdb')
    db.migrate()
    _seed(db)
    service = CompleteExportRawFusionService(db)
    info = service.create_fusion(['raw_a', 'raw_b'], 'T1', 2026, 'risk')

    folder = tmp_path / 'export'
    folder.mkdir()
    path = service.risk_exporter.export(info['id'], folder)
    assert path.name == '12_synthese_occurrences_agents_a_risque.xlsx'

    wb = load_workbook(path, read_only=True, data_only=True)
    assert 'Synthese generale' in wb.sheetnames
    assert '04_Matricule_NU' in wb.sheetnames
    assert '08_Multi_regimes' in wb.sheetnames
    assert 'Controle' in wb.sheetnames

    summary = list(wb['Synthese generale'].iter_rows(min_row=2, values_only=True))
    names = {row[4] for row in summary}
    assert 'Agent Risque' in names
    assert 'Sans Matricule' in names
    assert 'Agent Sain' not in names

    nu_rows = _all_rows(wb, '04_Matricule_NU')
    assert len(nu_rows) == 1
    assert nu_rows[0][9] == 'NU'
    assert nu_rows[0][11] == 'Sans Matricule'

    multi_rows = _all_rows(wb, '08_Multi_regimes')
    assert len(multi_rows) == 2
    assert {row[5] for row in multi_rows} == {'CNSS', 'CARC'}
    assert all(row[10] == '001' for row in multi_rows)

    # Aucun détail spécialisé ne doit contenir l'agent parfaitement sain.
    all_detail_names = set()
    for ws in wb.worksheets:
        if ws.title[:2].isdigit():
            all_detail_names.update(row[11] for row in ws.iter_rows(min_row=2, values_only=True))
    assert 'Agent Sain' not in all_detail_names

    control = {row[0]: row[1] for row in wb['Controle'].iter_rows(min_row=2, values_only=True)}
    assert control['Agents analyses au total'] == 3
    assert control['Agents a risque'] == 2
    assert control['Agents sains mono-regime exclus'] == 1
    assert control['Lignes physiques a risque attendues'] == 3
    assert control['Lignes detail exportees'] == 3
    assert control['Controle exclusion'] == 'OK'
    assert control['Controle detail'] == 'OK'
