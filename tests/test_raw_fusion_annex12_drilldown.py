from pathlib import Path

from openpyxl import load_workbook

from controle_paie.database import Database
from controle_paie.raw_fusion_annex12_drilldown import (
    normalize_matricule_python,
    normalize_name_python,
)
from controle_paie.raw_fusion_complete_export import CompleteExportRawFusionService


def test_matching_normalization_rules():
    assert normalize_name_python("MULUMBA ALEXANDRE") == "ALEXANDREMULUMBA"
    assert normalize_name_python("Alexandre,  Mulúmba") == "ALEXANDREMULUMBA"
    assert normalize_matricule_python("000-123") == "123"
    assert normalize_matricule_python(" 001 23 ") == "123"


def _seed(db):
    with db.connect() as con:
        con.execute("CREATE TABLE raw_a (execution_id VARCHAR, matricule VARCHAR, nom VARCHAR)")
        con.execute("CREATE TABLE raw_b (execution_id VARCHAR, matricule VARCHAR, nom VARCHAR)")
        con.execute("INSERT INTO raw_a VALUES ('e1','000123','MULUMBA ALEXANDRE')")
        con.execute("INSERT INTO raw_b VALUES ('e2','123','ALEXANDRE MULUMBA')")
        con.execute("""INSERT INTO journal_executions
            (execution_id,type_operation,fichier_source,table_source,table_destination,institution_id,regime,
             trimestre,annee,mode_chargement,lignes_lues,lignes_chargees,statut,date_fin)
            VALUES
            ('e1','IMPORT_ACCESS','a.accdb','A','raw_a','i1','REGIME_A','T1',2026,'append',1,1,'TERMINE',CURRENT_TIMESTAMP),
            ('e2','IMPORT_ACCESS','b.accdb','B','raw_b','i2','REGIME_B','T1',2026,'append',1,1,'TERMINE',CURRENT_TIMESTAMP)
        """)
        rows = [
            ('p1','e1','i1','REGIME_A','raw_a','000123','MULUMBA','ALEXANDRE',1,100.0,90.0),
            ('p2','e2','i2','REGIME_B','raw_b','123','ALEXANDRE','MULUMBA',1,120.0,100.0),
        ]
        for line_id, execution_id, inst, regime, table_source, mat, nom, prenom, line_no, gross, net in rows:
            con.execute("""INSERT INTO paie_standardisee
                (ligne_paie_id,execution_id,institution_id,regime,trimestre,annee,table_source,
                 matricule_source,matricule_normalise,nom,prenom,nom_normalise,section,categorie,grade,
                 unite_affectation,province,remuneration_base,transport,prime,logement,pension_rente,
                 autres_remunerations,retenues,montant_net,remuneration_brute_calculee,ligne_source)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                [line_id,execution_id,inst,regime,'T1',2026,table_source,mat,mat,nom,prenom,
                 f'{nom} {prenom}','SEC','CAT','GR','UNIT','PROV',gross,0,0,0,0,0,0,net,gross,line_no])


def test_annex12_has_exactly_seven_linked_sheets_and_financial_metrics(tmp_path):
    db = Database(tmp_path / 'annex12.duckdb')
    db.migrate()
    _seed(db)
    service = CompleteExportRawFusionService(db)
    info = service.create_fusion(['raw_a','raw_b'],'T1',2026,'drilldown')

    out = tmp_path / 'exports'
    out.mkdir()
    folder = Path(service.export_all(info['id'], out))
    path = folder / '12_matching_multi_regimes_drilldown.xlsx'
    assert path.exists()

    wb = load_workbook(path, read_only=False, data_only=False)
    assert wb.sheetnames == [
        'Synthese Globale','Synthese_Nom','Synthese_Matricule','Synthese_Matricule_Nom',
        'Detail_Nom','Detail_Matricule','Detail_Matricule_Nom',
    ]

    global_ws = wb['Synthese Globale']
    assert global_ws.max_row == 4
    assert global_ws['B2'].value == 1
    assert global_ws['C2'].value == 2
    assert global_ws['E2'].value == 220
    assert global_ws['F2'].value == 190
    assert global_ws['G2'].value == 100
    assert global_ws['H2'].value == 90
    assert global_ws['I2'].value == 1
    assert global_ws['J2'].hyperlink is not None

    nom_ws = wb['Synthese_Nom']
    assert nom_ws['A3'].value == 'ALEXANDREMULUMBA'
    assert nom_ws['B3'].value == 2
    assert nom_ws['C3'].value == 2
    assert nom_ws['C3'].hyperlink is not None
    assert "Detail_Nom" in nom_ws['C3'].hyperlink.target
    assert nom_ws['E3'].value == 220
    assert nom_ws['F3'].value == 190
    assert nom_ws['G3'].value == 110
    assert nom_ws['I3'].value == 100
    assert nom_ws['J3'].value == 120
    assert nom_ws['K3'].value == 20
    assert nom_ws['O3'].value == 100
    assert nom_ws['P3'].value == 90
    assert 'REGIME_A=1' in nom_ws['Q3'].value
    assert 'REGIME_B=1' in nom_ws['Q3'].value
    assert nom_ws['S3'].value == 'RISQUE_DOUBLE_PAIEMENT'

    mat_ws = wb['Synthese_Matricule']
    assert mat_ws['A3'].value == '123'
    assert mat_ws['C3'].value == 2
    assert mat_ws['O3'].value == 100

    combo_ws = wb['Synthese_Matricule_Nom']
    assert combo_ws['A3'].value == '123|ALEXANDREMULUMBA'
    assert combo_ws['C3'].value == 2

    detail_nom = wb['Detail_Nom']
    assert detail_nom.max_row == 4
    assert detail_nom['A3'].value == 'p1'
    assert detail_nom['A4'].value == 'p2'
    assert detail_nom['V3'].value == 220
    assert detail_nom['AF3'].value == 100
    assert detail_nom['AI3'].value == 'RISQUE_DOUBLE_PAIEMENT'
