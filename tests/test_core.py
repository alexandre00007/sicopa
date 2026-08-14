import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd

from openpyxl import load_workbook

from controle_paie.database import Database
from controle_paie.config import AppConfig, RegimeConfig
from controle_paie.explorer import DataExplorerService
from controle_paie.loaders import IngestionService
from controle_paie.matching import MatchingService
from controle_paie.multiregime import MultiRegimeAnalysisService
from controle_paie.listing_analysis import ListingGroupAnalysisService
from controle_paie.letters import generate_interpretation_letter
from controle_paie.standardization import normalize_identifier, standardize_declaration, standardize_payroll
from controle_paie.ui import fitted_window_geometry, validate_scope_values


class CoreTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.db = Database(Path(self.temp.name) / "test.duckdb")
        self.db.migrate()
        self.institution = self.db.add_institution("TEST", "Institution Test")

    def tearDown(self):
        self.temp.cleanup()

    def metadata(self):
        return dict(execution_id="e1", institution_id=self.institution, regime="PNC",
                    trimestre="T1", annee=2024, table_source="Tab_PNC_T1_2024",
                    fichier_source="declaratif.xlsx", feuille_source="Agents")

    def test_normalization_and_aliases(self):
        self.assertEqual(normalize_identifier(" AB-12 / ç "), "AB12C")
        self.assertEqual(normalize_identifier("straße"), "STRASSE")
        frame = pd.DataFrame({"NumMatricule": ["A-1","straße"], "Noms": ["Jean Test","Agent Test"],
                              "TraitementBase": [100,50], "Primes": [20,0]})
        result = standardize_payroll(frame, self.metadata())
        self.assertEqual(result.loc[0, "matricule_normalise"], "A1")
        self.assertEqual(result.loc[1, "matricule_normalise"], "STRASSE")
        self.assertEqual(result.loc[0, "remuneration_brute_calculee"], 120)

    def test_child_window_geometry_is_always_visible(self):
        width,height,x,y=fitted_window_geometry(1024,600,1400,900,512,300)
        self.assertLessEqual(x+width,1000)
        self.assertLessEqual(y+height,528)
        self.assertGreaterEqual(x,24)
        self.assertGreaterEqual(y,48)
        self.assertEqual((width,height),(976,480))

    def test_scope_validation_lists_missing_fields(self):
        with self.assertRaisesRegex(ValueError, "institution, trimestre"):
            validate_scope_values("", "PNC", "", "2024")
        self.assertEqual(validate_scope_values("inst-1", "PNC", "T1", "2024"),
                         ("inst-1", "PNC", "T1", 2024))

    def test_dynamic_regime_configuration(self):
        self.db.upsert_regime("SANTE", "Régime de la santé", r"^Tab_Sante_T[1-4]_\d{4}$", "raw_sante")
        rows = self.db.list_regimes()
        self.assertEqual(rows[0][0], "SANTE")
        config = AppConfig(regimes={row[0]: RegimeConfig(row[0], row[2], row[3]) for row in rows})
        self.assertEqual(config.detect_regime("Tab_Sante_T2_2025"), "SANTE")
        self.db.set_regime_active("SANTE", False)
        self.assertEqual(self.db.list_regimes(), [])
        self.db.delete_regime("SANTE")
        self.assertEqual(self.db.list_regimes(active_only=False), [])

    def test_data_scope_deletion_removes_standardized_records(self):
        pay = standardize_payroll(pd.DataFrame({
            "Matricule": ["A1", "B2"], "NomPostnom": ["Alpha", "Beta"], "Base": [100, 200]
        }), dict(execution_id="pay-del", institution_id=self.institution, regime="PNC",
                 trimestre="T1", annee=2024, table_source="raw_pnc"))
        declaration = standardize_declaration(pd.DataFrame({
            "Matricule": ["A1", "B2"], "Noms": ["Alpha", "Beta"]
        }), dict(execution_id="decl-del", institution_id=self.institution, regime="PNC",
                 trimestre="T1", annee=2024, fichier_source="decl.xlsx", feuille_source="Agents"))
        with self.db.connect() as connection:
            connection.register("pay", pay)
            connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM pay")
            connection.register("declaration", declaration)
            connection.execute("INSERT INTO declaratif_standardise BY NAME SELECT * FROM declaration")
            connection.execute("INSERT INTO resultats_rapprochement (rapprochement_id, execution_id, institution_id, regime, trimestre, annee, ligne_paie_id, ligne_declaratif_id, methode_correspondance, score_correspondance, statut_rapprochement, masse_financiere_controlee, impact_potentiel) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                ["r1","pay-del",self.institution,"PNC","T1",2024,"p1","d1","MANUEL",1.0,"OK",100,0])
        deleted = self.db.delete_data_scope(self.institution, "PNC", "T1", 2024)
        self.assertEqual(deleted["paie_rows"], 2)
        self.assertEqual(deleted["declaratif_rows"], 2)
        self.assertEqual(deleted["matching_rows"], 1)
        with self.db.connect() as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM paie_standardisee WHERE institution_id=? AND regime=? AND trimestre=? AND annee=?",
                [self.institution, "PNC", "T1", 2024]).fetchone()[0], 0)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM declaratif_standardise WHERE institution_id=? AND regime=? AND trimestre=? AND annee=?",
                [self.institution, "PNC", "T1", 2024]).fetchone()[0], 0)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM resultats_rapprochement WHERE institution_id=? AND regime=? AND trimestre=? AND annee=?",
                [self.institution, "PNC", "T1", 2024]).fetchone()[0], 0)

    def test_table_search_can_delete_filtered_rows_without_raw_prefix(self):
        service = DataExplorerService(self.db)
        with self.db.connect() as connection:
            connection.execute("CREATE TABLE custom_data AS SELECT 'A1' AS matricule, 'Alpha' AS nom, 'T1' AS trimestre, 2024 AS annee, 'PNC' AS regime UNION ALL SELECT 'B2','Beta','T1',2024,'PNC' UNION ALL SELECT 'C3','Charlie','T2',2024,'PNC'")
        deleted = service.delete_rows("custom_data", "regime", "égal à", "PNC", "trimestre", "égal à", "T1")
        self.assertEqual(deleted, 2)
        with self.db.connect() as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM custom_data").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT matricule FROM custom_data").fetchone()[0], "C3")

    def test_app_config_can_add_regime(self):
        config = AppConfig()
        config.add_regime("SANTE", "Régime de la santé", r"^Tab_Sante_T[1-4]_\d{4}$", "raw_sante")
        self.assertIn("SANTE", config.regimes)
        self.assertEqual(config.detect_regime("Tab_Sante_T2_2025"), "SANTE")

    def test_access_import_reports_every_progress_stage(self):
        frame=pd.DataFrame({"Matricule":["A1","A2"],"NomPostnom":["Alpha","Beta"],"Base":[100,200]})
        events=[];config=AppConfig(database_path=Path(self.temp.name)/"test.duckdb")
        service=IngestionService(self.db,config)
        with patch("controle_paie.loaders.read_access_table",return_value=frame):
            service.load_access("source.accdb","Tab_PNC_T1_2024",self.institution,"PNC","T1",2024,progress=lambda value,text:events.append((value,text)))
        values=[value for value,_text in events]
        self.assertEqual(values,[-1,30,-1,55,75,90,100])
        self.assertIn("2 lignes",events[1][1])
        self.assertIn("Table Access chargée",events[-1][1])

    def test_payroll_excel_import_uses_payroll_schema_and_progress(self):
        source=Path(self.temp.name)/"listing.xlsx"
        pd.DataFrame({"ID Agent":["A1","A2"],"Nom agent":["Alpha","Beta"],"Traitement":[125,275]}).to_excel(source,index=False,sheet_name="Paie")
        self.db.upsert_column_mapping("PNC","PAIE_EXCEL","ID Agent","matricule_source",True)
        self.db.upsert_column_mapping("PNC","PAIE_EXCEL","Nom agent","nom",True)
        self.db.upsert_column_mapping("PNC","PAIE_EXCEL","Traitement","remuneration_base",True)
        events=[];config=AppConfig(database_path=Path(self.temp.name)/"test.duckdb")
        execution=IngestionService(self.db,config).load_payroll_excel(
            str(source),"Paie",1,self.institution,"PNC","T1",2024,
            progress=lambda value,text:events.append((value,text)))
        with self.db.connect() as connection:
            rows=connection.execute("SELECT matricule_normalise,nom,remuneration_base FROM paie_standardisee WHERE execution_id=? ORDER BY matricule_normalise",[execution]).fetchall()
            operation=connection.execute("SELECT type_operation FROM journal_executions WHERE execution_id=?",[execution]).fetchone()[0]
        self.assertEqual(rows,[("A1","Alpha",125),("A2","Beta",275)])
        self.assertEqual(operation,"IMPORT_PAIE_EXCEL")
        self.assertEqual([value for value,_text in events],[-1,30,-1,60,90,100])

    def test_multi_regime_analysis_filters_sources_and_uses_payment_regime(self):
        other=self.db.add_institution("OTHER","Institution Autre")
        target_pay=standardize_payroll(pd.DataFrame({
            "Matricule":["A1"],"NomPostnom":["Alpha"],"Base":[100],"Section":["OK"]}),
            dict(execution_id="pay-target",institution_id=self.institution,regime="PNC",
                 trimestre="T1",annee=2024,table_source="raw_pnc"))
        other_pay=standardize_payroll(pd.DataFrame({
            "Matricule":["A1","A2"],"NomPostnom":["Alpha","Beta"],"Base":[200,300],
            "Section":["KEEP","DROP"]}),
            dict(execution_id="pay-other",institution_id=other,regime="FARDC",
                 trimestre="T1",annee=2024,table_source="raw_fardc"))
        declaration=standardize_declaration(pd.DataFrame({
            "Matricule":["A1","A2"],"Noms":["Alpha","Beta"]}),
            dict(execution_id="decl",institution_id=self.institution,regime="PNC",
                 trimestre="T1",annee=2024,fichier_source="decl.xlsx",feuille_source="Agents"))
        with self.db.connect() as connection:
            connection.register("target_pay",target_pay);connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM target_pay")
            connection.register("other_pay",other_pay);connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM other_pay")
            connection.register("declaration",declaration);connection.execute("INSERT INTO declaratif_standardise BY NAME SELECT * FROM declaration")
        self.db.add_treatment_filter(other,"FARDC","section","égal à","KEEP")
        events=[];result=MultiRegimeAnalysisService(self.db).run(
            self.institution,"PNC","T1",2024,["pay-target","pay-other"],
            progress=lambda value,text:events.append((value,text)))
        with self.db.connect() as connection:
            retained=dict(connection.execute(
                "SELECT execution_id,lignes_retenues FROM sources_analyse_multi WHERE campagne_id=?",
                [result["campaign_id"]]).fetchall())
            rows=connection.execute(
                "SELECT matricule_normalise,regime_paiement,statut_analyse,impact_potentiel,formule_impact_id FROM resultats_analyse_multi WHERE campagne_id=? ORDER BY matricule_normalise,regime_paiement",
                [result["campaign_id"]]).fetchall()
        self.assertEqual(retained,{"pay-target":1,"pay-other":1})
        self.assertEqual(rows[0][2],"PAYE_PLUSIEURS_REGIMES")
        self.assertEqual(rows[1][1],"PNC")
        self.assertEqual(rows[1][3],0)
        self.assertEqual(rows[2][0],"A2")
        self.assertEqual(rows[2][2],"DECLARE_NON_PAYE")
        self.assertEqual(rows[0][4],"FORMULE_DEFAUT")
        self.assertEqual(events[-1][0],100)
        export_dir=Path(self.temp.name)/"exports"
        package=MultiRegimeAnalysisService(self.db).export(result["campaign_id"],str(export_dir))
        self.assertTrue((package/"rapport_multi_regimes.xlsx").is_file())
        self.assertTrue((package/"annexe_resultats_multi_regimes.xlsx").is_file())
        self.assertTrue((package/"effectifs_uniques_multi_regimes.xlsx").is_file())
        self.assertTrue((package/"lettre_interpretation_multi_regimes.docx").is_file())
        category_files=list((package/"annexes_par_categorie").glob("*.xlsx"))
        self.assertGreaterEqual(len(category_files),2)
        campaign=MultiRegimeAnalysisService(self.db).list_campaigns()[0]
        self.assertEqual(Path(campaign[11]),package)

    def test_multi_regime_requires_an_exact_declaration_version(self):
        pay=standardize_payroll(pd.DataFrame({"Matricule":["A1","B2"],"Noms":["Alpha","Beta"],"Base":[100,200]}),
            dict(execution_id="pay-one",institution_id=self.institution,regime="PNC",
                 trimestre="T1",annee=2024,table_source="raw_pnc"))
        decl_one=standardize_declaration(pd.DataFrame({"Matricule":["A1"],"Noms":["Alpha"]}),
            dict(execution_id="decl-one",institution_id=self.institution,regime="PNC",
                 trimestre="T1",annee=2024,fichier_source="one.xlsx",feuille_source="Agents"))
        decl_two=standardize_declaration(pd.DataFrame({"Matricule":["B2"],"Noms":["Beta"]}),
            dict(execution_id="decl-two",institution_id=self.institution,regime="PNC",
                 trimestre="T1",annee=2024,fichier_source="two.xlsx",feuille_source="Agents"))
        with self.db.connect() as connection:
            connection.register("pay",pay);connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM pay")
            connection.register("d1",decl_one);connection.execute("INSERT INTO declaratif_standardise BY NAME SELECT * FROM d1")
            connection.register("d2",decl_two);connection.execute("INSERT INTO declaratif_standardise BY NAME SELECT * FROM d2")
        service=MultiRegimeAnalysisService(self.db)
        with self.assertRaisesRegex(ValueError,"Plusieurs versions"):
            service.run(self.institution,"PNC","T1",2024,["pay-one"])
        diagnosis=service.diagnose(self.institution,"PNC","T1",2024,"decl-one",["pay-one"])
        self.assertEqual(diagnosis["declaration_rows"],1)
        self.assertTrue(diagnosis["ready"])
        self.assertIn("STANDARD",diagnosis["sources"][0]["mapping"])
        columns,sample=service.sample_source("pay-one",1)
        self.assertEqual(columns[0],"matricule_source")
        self.assertEqual(len(sample),1)
        result=service.run(self.institution,"PNC","T1",2024,["pay-one"],
                           declaration_execution_id="decl-one")
        with self.db.connect() as connection:
            declarative=connection.execute(
                "SELECT declaratif_execution_id FROM campagnes_analyse_multi WHERE campagne_id=?",
                [result["campaign_id"]]).fetchone()[0]
            people=connection.execute(
                "SELECT DISTINCT matricule_normalise FROM resultats_analyse_multi WHERE campagne_id=?",
                [result["campaign_id"]]).fetchall()
        self.assertEqual(declarative,"decl-one")
        self.assertEqual(people,[("A1",)])
        self.assertEqual(service.list_campaigns()[0][0],result["campaign_id"])
        service.archive_campaign(result["campaign_id"])
        self.assertEqual(service.list_campaigns(),[])
        self.assertEqual(service.list_campaigns(True)[0][0],result["campaign_id"])

    def test_listing_group_analysis_builds_filters_classifies_and_exports(self):
        other=self.db.add_institution("OTHER","Institution Autre")
        first=standardize_payroll(pd.DataFrame({
            "Matricule":["A1","B2","B2","NU"],"NomPostnom":["Alpha","Beta","Beta","Sans matricule"],
            "Base":[100,200,200,50],"Section":["KEEP","KEEP","KEEP","KEEP"]}),
            dict(execution_id="listing-pnc",institution_id=self.institution,regime="PNC",
                 trimestre="T1",annee=2024,table_source="raw_pnc"))
        second=standardize_payroll(pd.DataFrame({
            "Matricule":["A1","C3","D4"],"NomPostnom":["Alpha","Charlie","Delta"],
            "Base":[300,400,500],"Section":["KEEP","KEEP","DROP"]}),
            dict(execution_id="listing-fardc",institution_id=other,regime="FARDC",
                 trimestre="T1",annee=2024,table_source="raw_fardc"))
        with self.db.connect() as connection:
            connection.register("first_listing",first)
            connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM first_listing")
            connection.register("second_listing",second)
            connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM second_listing")
        self.db.add_treatment_filter(other,"FARDC","section","égal à","KEEP")
        service=ListingGroupAnalysisService(self.db)
        diagnosis=service.preview("T1",2024,["listing-pnc","listing-fardc"])
        self.assertEqual([row["retained"] for row in diagnosis],[4,2])
        events=[];result=service.run("Groupe trimestriel","T1",2024,
            ["listing-pnc","listing-fardc"],progress=lambda value,text:events.append((value,text)))
        self.assertEqual(result["base_rows"],6);self.assertEqual(events[-1][0],100)
        with self.db.connect() as connection:
            statuses=dict(connection.execute("""SELECT statut_analyse,COUNT(*)
                FROM resultats_analyse_listing WHERE groupe_id=? GROUP BY 1""",
                [result["group_id"]]).fetchall())
            retained=dict(connection.execute("""SELECT execution_id,lignes_retenues
                FROM sources_analyse_listing WHERE groupe_id=?""",[result["group_id"]]).fetchall())
        self.assertEqual(retained,{"listing-pnc":4,"listing-fardc":2})
        self.assertEqual(statuses["PAYE_PLUSIEURS_REGIMES"],2)
        self.assertEqual(statuses["DOUBLON_MATRICULE"],2)
        self.assertEqual(statuses["MATRICULE_NON_EXPLOITABLE"],1)
        self.assertEqual(statuses["UNIQUE_DANS_GROUPE"],1)
        package=service.export(result["group_id"],str(Path(self.temp.name)/"exports"))
        self.assertTrue((package/"rapport_analyse_listings.xlsx").is_file())
        workbook = load_workbook(package/"annexe_globale_listings.xlsx")
        sheet = workbook["Résultats"]
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("Nom complet", headers)
        self.assertTrue(any(row[5] == "Alpha" for row in sheet.iter_rows(min_row=2, values_only=True)))

        report = load_workbook(package/"rapport_analyse_listings.xlsx")
        report_links = [cell.hyperlink.target for row in report["Synthèse"].iter_rows() for cell in row if getattr(cell, "hyperlink", None)]
        self.assertTrue(any(target.endswith("annexe_globale_listings.xlsx") for target in report_links))
        self.assertTrue((package/"annexe_globale_listings.xlsx").is_file())
        self.assertTrue((package/"effectifs_uniques_listings.xlsx").is_file())
        self.assertTrue((package/"lettre_interpretation_listings.docx").is_file())
        self.assertGreaterEqual(len(list((package/"annexes_par_categorie").glob("*.xlsx"))),4)
        self.assertEqual(Path(service.list_groups()[0][8]),package)
        service.archive_group(result["group_id"])
        self.assertEqual(service.list_groups(),[])

    def test_listing_inter_regime_summary_is_available(self):
        other=self.db.add_institution("OTHER","Institution Autre")
        first=standardize_payroll(pd.DataFrame({
            "Matricule":["A1","B2"],"NomPostnom":["Alpha","Beta"],"Base":[100,200],"Section":["KEEP","KEEP"]
        }), dict(execution_id="listing-pnc-compare",institution_id=self.institution,regime="PNC",
                 trimestre="T1",annee=2024,table_source="raw_pnc"))
        second=standardize_payroll(pd.DataFrame({
            "Matricule":["A1","C3"],"NomPostnom":["Alpha","Charlie"],"Base":[300,400],"Section":["KEEP","KEEP"]
        }), dict(execution_id="listing-fardc-compare",institution_id=other,regime="FARDC",
                 trimestre="T1",annee=2024,table_source="raw_fardc"))
        with self.db.connect() as connection:
            connection.register("first_compare",first);connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM first_compare")
            connection.register("second_compare",second);connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM second_compare")
        service=ListingGroupAnalysisService(self.db)
        result=service.run("Groupe comparatif","T1",2024,["listing-pnc-compare","listing-fardc-compare"])
        summary=service.inter_regime_summary(result["group_id"])
        regimes={row[0] for row in summary}
        self.assertIn("PNC",regimes)
        self.assertIn("FARDC",regimes)
        self.assertTrue(any(row[1]=="PAYE_PLUSIEURS_REGIMES" for row in summary))

    def test_one_classification_per_payroll_row(self):
        pay = standardize_payroll(pd.DataFrame({
            "Matricule": ["A", "B", "B", "C"], "NomPostnom": ["Alpha", "Beta", "Beta", "Gamma"],
            "Base": [100, 200, 200, 300]}), self.metadata())
        declaration = standardize_declaration(
            pd.DataFrame({"matricule": ["A", "B"], "noms": ["Alpha", "Beta"]}), self.metadata())
        with self.db.connect() as connection:
            connection.register("pay", pay)
            connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM pay")
            connection.register("declaration", declaration)
            connection.execute("INSERT INTO declaratif_standardise BY NAME SELECT * FROM declaration")
        execution = MatchingService(self.db).run(self.institution, "PNC", "T1", 2024)
        with self.db.connect() as connection:
            rows = dict(connection.execute(
                "SELECT statut_rapprochement, COUNT(*) FROM resultats_rapprochement WHERE execution_id=? GROUP BY 1",
                [execution]).fetchall())
            total = connection.execute(
                "SELECT COUNT(*) FROM resultats_rapprochement WHERE execution_id=? AND ligne_paie_id IS NOT NULL",
                [execution]).fetchone()[0]
        self.assertEqual(total, 4)
        self.assertEqual(rows["DOUBLON_MATRICULE"], 1)
        self.assertEqual(rows["PAYE_NON_DECLARE"], 1)

    def test_nu_variants_are_never_matricule_duplicates(self):
        pay = standardize_payroll(pd.DataFrame({
            "Matricule": ["NU", "N.U", "nu", "nU"],
            "NomPostnom": ["Alpha", "Beta", "Gamma", "Delta"],
            "Base": [100, 100, 100, 100]}), self.metadata())
        with self.db.connect() as connection:
            connection.register("pay_nu", pay)
            connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM pay_nu")
        execution = MatchingService(self.db).run(self.institution, "PNC", "T1", 2024)
        with self.db.connect() as connection:
            statuses = dict(connection.execute(
                "SELECT statut_rapprochement,COUNT(*) FROM resultats_rapprochement WHERE execution_id=? AND ligne_paie_id IS NOT NULL GROUP BY 1",
                [execution]).fetchall())
        self.assertNotIn("DOUBLON_MATRICULE", statuses)
        self.assertEqual(statuses["MATRICULE_MANQUANT"], 4)

    def test_listing_filters_are_applied_before_matching(self):
        pay = standardize_payroll(pd.DataFrame({
            "Matricule": ["A", "B"], "NomPostnom": ["Alpha", "Beta"],
            "Section": ["CABINET", "DIRECTION"], "Base": [100, 200]}), self.metadata())
        with self.db.connect() as connection:
            connection.register("pay_filter", pay)
            connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM pay_filter")
        self.db.add_treatment_filter(self.institution,"PNC","section","égal à","CABINET")
        execution=MatchingService(self.db).run(self.institution,"PNC","T1",2024)
        with self.db.connect() as connection:
            rows=connection.execute("SELECT COUNT(*) FROM resultats_rapprochement WHERE execution_id=? AND ligne_paie_id IS NOT NULL",[execution]).fetchone()[0]
        self.assertEqual(rows,1)

    def test_interpretation_letter_is_created(self):
        from docx import Document
        target=Path(self.temp.name)/"lettre_interpretation.docx"
        summaries=[
            {"label":"Données du listing de paie filtré","count":10,"concerned":9,"impact":1000,"group":"annexes_listing"},
            {"label":"Données déclaratives","count":8,"concerned":8,"impact":0,"group":"annexes_declaratif"},
            {"label":"Doublons par matricule hors NU","count":2,"concerned":1,"impact":200,"group":"annexes_listing"},
        ]
        generate_interpretation_letter(target,"Institution Test","PNC","T1",2024,summaries,[])
        self.assertTrue(target.exists())
        text="\n".join(paragraph.text for paragraph in Document(target).paragraphs)
        self.assertIn("Institution Test",text)
        self.assertIn("Interprétation générale",text)

    def test_versioned_impact_formula_and_extra_component(self):
        self.db.add_financial_component("PRIME_TECHNIQUE","Prime technique")
        formula_id=self.db.save_impact_formula("Net technique",self.institution,"PNC","PAYE_NON_DECLARE","T1",2024,"TOUTES_LIGNES",[
            {"code":"REMUNERATION_BASE","coefficient":1},{"code":"PRIME_TECHNIQUE","coefficient":1},{"code":"RETENUES","coefficient":-1}])
        pay=standardize_payroll(pd.DataFrame({"Matricule":["X"],"NomPostnom":["Agent X"],"Base":[100],"Retenue":[10],"PrimeTech":[25]}),self.metadata(),{"PrimeTech":"composante_PRIME_TECHNIQUE"})
        with self.db.connect() as connection:
            connection.register("pay_formula",pay);connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM pay_formula")
        execution=MatchingService(self.db).run(self.institution,"PNC","T1",2024)
        with self.db.connect() as connection:
            row=connection.execute("SELECT impact_potentiel,formule_impact_id FROM resultats_rapprochement WHERE execution_id=? AND ligne_paie_id IS NOT NULL",[execution]).fetchone()
        self.assertEqual(float(row[0]),115.0);self.assertEqual(row[1],formula_id)
    def test_matching_can_force_a_global_formula_and_detect_existing_fields(self):
        self.db.upsert_column_mapping("PNC","ACCESS","TransportSource","transport",False)
        specific=self.db.save_impact_formula(
            "Impact non déclaré",self.institution,"PNC","PAYE_NON_DECLARE","T1",2024,
            "TOUTES_LIGNES",[{"code":"REMUNERATION_BASE","coefficient":1}])
        forced=self.db.save_impact_formula(
            "Profil doublé",self.institution,"PNC","*","T1",2024,
            "TOUTES_LIGNES",[{"code":"REMUNERATION_BASE","coefficient":2}])
        pay=standardize_payroll(pd.DataFrame({
            "Matricule":["FORCE-1"],"NomPostnom":["Agent Force"],"Base":[100]}),self.metadata())
        with self.db.connect() as connection:
            connection.register("pay_forced",pay)
            connection.execute("INSERT INTO paie_standardisee BY NAME SELECT * FROM pay_forced")

        available={row[0]:row for row in self.db.available_financial_components(
            self.institution,"PNC","T1",2024)}
        self.assertIn("REMUNERATION_BASE",available)
        self.assertIn("REMUNERATION_BRUTE_CALCULEE",available)
        self.assertTrue(available["TRANSPORT"][3])
        selectable={formula["id"] for formula in self.db.selectable_impact_formulas(
            self.institution,"PNC","T1",2024)}
        self.assertIn(forced,selectable);self.assertNotIn(specific,selectable)

        execution=MatchingService(self.db).run(
            self.institution,"PNC","T1",2024,impact_formula_id=forced)
        with self.db.connect() as connection:
            row=connection.execute(
                "SELECT impact_potentiel,formule_impact_id FROM resultats_rapprochement "
                "WHERE execution_id=? AND ligne_paie_id IS NOT NULL",[execution]).fetchone()
        self.assertEqual(float(row[0]),200.0);self.assertEqual(row[1],forced)
        from controle_paie.reports import ReportService
        prepared=ReportService(self.db)._prepare_impact_specs([{
            "label":"Non déclarés mais présents dans le listing filtré par matricule",
            "query":"SELECT * FROM paie_standardisee","params":[]}],
            self.institution,"PNC","T1",2024,forced)
        self.assertEqual(prepared[0]["formula"]["id"],forced)

        with self.assertRaisesRegex(ValueError,"ne correspond pas"):
            self.db.selected_impact_formula(forced,self.institution,"FARDC","T1",2024)



    def test_formula_occurrences_supplementaires(self):
        self.db.save_impact_formula("Doublons supplémentaires",self.institution,"PNC","DOUBLON_MATRICULE","T1",2024,"OCCURRENCES_SUPPLEMENTAIRES",[{"code":"REMUNERATION_BASE","coefficient":1}])
        expression,formula=self.db.impact_sql(self.institution,"PNC","T1",2024,"DOUBLON_MATRICULE","p","rn")
        self.assertIn("CASE WHEN rn>1",expression);self.assertEqual(formula["aggregation"],"OCCURRENCES_SUPPLEMENTAIRES")

    def test_default_formula_is_visible_and_used_as_fallback(self):
        formula=self.db.default_impact_formula()
        self.assertEqual(formula["id"],"FORMULE_DEFAUT")
        self.assertEqual(formula["aggregation"],"TOUTES_LIGNES")
        self.assertEqual([term["code"] for term in formula["terms"]],["REMUNERATION_BASE","TRANSPORT","PRIME","LOGEMENT","PENSION_RENTE","AUTRES_REMUNERATIONS"])
        resolved=self.db.resolve_impact_formula(self.institution,"PNC","T1",2024,"PAYE_NON_DECLARE")
        self.assertEqual(resolved["id"],"FORMULE_DEFAUT")

    def test_formula_details_and_non_destructive_modification(self):
        first=self.db.save_impact_formula("Impact initial",self.institution,"PNC","PAYE_NON_DECLARE","T1",2024,"TOUTES_LIGNES",[{"code":"REMUNERATION_BASE","coefficient":1}])
        detail=self.db.get_impact_formula(first)
        self.assertEqual(detail["version"],1);self.assertEqual(detail["terms"][0]["code"],"REMUNERATION_BASE")
        second=self.db.save_impact_formula("Impact modifié",self.institution,"PNC","PAYE_NON_DECLARE","T1",2024,"TOUTES_LIGNES",[{"code":"REMUNERATION_BASE","coefficient":1},{"code":"RETENUES","coefficient":-1}])
        self.assertEqual(self.db.get_impact_formula(second)["version"],2)
        self.assertEqual(self.db.get_impact_formula(first)["name"],"Impact initial")
        self.assertEqual(self.db.resolve_impact_formula(self.institution,"PNC","T1",2024,"PAYE_NON_DECLARE")["id"],second)


if __name__ == "__main__":
    unittest.main()
