from pathlib import Path

from openpyxl import load_workbook

from controle_paie.database import Database
from controle_paie.raw_period_bilateral_export import BilateralRawPeriodComparisonExporter
from controle_paie.raw_period_occurrences import OccurrenceAwareRawPeriodComparisonService


def _seed_occurrences(db):
    service = OccurrenceAwareRawPeriodComparisonService(db)
    service.ensure_schema()
    rows = [
        # commun exact 1/1
        ('A','a1','001','DUPONT','JEAN','RA'),
        ('B','b1','1','JEAN','DUPONT','RB'),
        # commun avec occurrences differentes 2/1
        ('A','a2','002','MULUMBA','ALEXANDRE','RA'),
        ('A','a3','0002','ALEXANDRE','MULUMBA','RA'),
        ('B','b2','2','MULUMBA','ALEXANDRE','RB'),
        # discordance meme matricule, noms differents
        ('A','a4','003','KABILA','JEAN','RA'),
        ('B','b3','3','MBUYI','PAUL','RB'),
        # exclusifs
        ('A','a5','004','SOLO','ALPHA','RA'),
        ('B','b4','005','SOLO','BETA','RB'),
    ]
    with db.connect() as con:
        for side, line_id, mat, nom, prenom, regime in rows:
            con.execute("""INSERT INTO occurrences_comparaison_raw
                (comparaison_id,cote,table_source,execution_id,ligne_paie_id,ligne_source,
                 matricule_normalise,nom_normalise,nom,prenom,institution_id,regime,
                 section,categorie,grade,unite_affectation,province,brut,net)
                VALUES ('cmp1',?,'raw_'||lower(?),'e_'||?, ?,1,?,?,?,?,'INST',?,
                        'SEC','CAT','GR','UNIT','PROV',100,90)""",
                [side, side, line_id, line_id, mat, f'{nom} {prenom}', nom, prenom, regime])


def test_bilateral_export_matrix_and_drilldown(tmp_path):
    db = Database(tmp_path / 'bilateral.duckdb')
    db.migrate()
    _seed_occurrences(db)
    exporter = BilateralRawPeriodComparisonExporter(db)
    out = tmp_path / 'out'
    out.mkdir()
    path = exporter.export('cmp1', out)
    assert path.exists()

    wb = load_workbook(path, read_only=False, data_only=False)
    assert wb.sheetnames == [
        'Synthese Globale Comparee','Synthese_Nom_Comparee','Synthese_Matricule_Comparee',
        'Synthese_MatNom_Comparee','Detail_Nom_Compare','Detail_Matricule_Compare','Detail_MatNom_Compare'
    ]

    global_ws = wb['Synthese Globale Comparee']
    # Ligne Matricule: 3 communs, 1 exclusif A, 1 exclusif B, 1 discordance => overlap 60%
    assert global_ws['B3'].value == 3
    assert global_ws['C3'].value == 1
    assert global_ws['D3'].value == 1
    assert global_ws['E3'].value == 1
    assert global_ws['F3'].value == 60
    assert global_ws['G3'].hyperlink is not None

    mat_ws = wb['Synthese_Matricule_Comparee']
    values = {row[0].value: row for row in mat_ws.iter_rows(min_row=3)}
    assert values['2'][1].value == 2
    assert values['2'][2].value == 1
    assert values['2'][3].value == 3
    assert values['2'][4].value == 1
    assert values['2'][5].value == 'COMMUN_AVEC_OCCURRENCES_DIFFERENTES'
    assert values['2'][3].hyperlink is not None

    assert values['3'][5].value == 'COMMUN_DISCORDANCE'
    assert values['3'][6].value == 'ALERTE'
    assert values['4'][5].value == 'EXCLUSIF_REGIME_A'
    assert values['5'][5].value == 'EXCLUSIF_REGIME_B'

    detail = wb['Detail_Matricule_Compare']
    # Pour le matricule 2, 2 occurrences A et 1 B donnent 2 lignes cote-a-cote, pas 2x1.
    target = values['2'][3].hyperlink.target
    start_row = int(target.split('!A')[-1])
    assert detail.cell(start_row, 1).value == '2'
    assert detail.cell(start_row + 1, 1).value == '2'
    assert detail.cell(start_row + 1, 7).value is None
