from pathlib import Path

from openpyxl import load_workbook

from controle_paie.database import Database
from controle_paie.raw_period_a_to_b_export import AToBRawPeriodExporter, AToBExportRawPeriodComparisonService


def _insert_occ(con, cmp_id, side, line_id, line_no, mat, nom, prenom, regime, brut, net):
    con.execute("""INSERT INTO occurrences_comparaison_raw
        (comparaison_id,cote,table_source,execution_id,ligne_paie_id,ligne_source,
         matricule_normalise,nom_normalise,nom,prenom,institution_id,regime,section,categorie,
         grade,unite_affectation,province,brut,net)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        [cmp_id,side,f'raw_{side.lower()}',f'exec_{side}',line_id,line_no,mat,
         f'{nom}{prenom}',nom,prenom,f'inst_{side}',regime,'SEC','CAT','GR','UNIT','PROV',brut,net])


def test_a_to_b_export_matches_annex12_logic_without_exporting_missing_details(tmp_path):
    db=Database(tmp_path/'a2b.duckdb')
    db.migrate()
    service=AToBExportRawPeriodComparisonService(db)
    service.ensure_schema()
    cmp_id='cmp-a2b'
    with db.connect() as con:
        # Exact Nom+Matricule, 2 occurrences A contre 1 B.
        _insert_occ(con,cmp_id,'A','a1',1,'0001','MULUMBA','ALEXANDRE','REG_A',100,90)
        _insert_occ(con,cmp_id,'A','a2',2,'0001','ALEXANDRE','MULUMBA','REG_A',120,100)
        _insert_occ(con,cmp_id,'B','b1',1,'1','ALEXANDRE','MULUMBA','REG_B',110,95)
        # Meme matricule mais nom different -> discordance identite.
        _insert_occ(con,cmp_id,'A','a3',3,'0002','BETA','TEST','REG_A',200,180)
        _insert_occ(con,cmp_id,'B','b2',2,'2','GAMMA','TEST','REG_B',210,185)
        # Meme nom mais matricule different -> trouve par nom.
        _insert_occ(con,cmp_id,'B','b3',3,'999','BETA','TEST','REG_B',205,182)
        # Exclusif A : doit rester dans la synthese mais pas gonfler le detail.
        _insert_occ(con,cmp_id,'A','a4',4,'0003','DELTA','TEST','REG_A',300,270)

    out=tmp_path/'exports'; out.mkdir()
    path=AToBRawPeriodExporter(db).export(cmp_id,out)
    assert path.name == '20_analyse_RAW_A_vers_B.xlsx'
    wb=load_workbook(path,read_only=False,data_only=False)
    assert wb.sheetnames == [
        'Synthese Globale A vers B','Synthese_Nom_A_vers_B','Synthese_Matricule_A_vers_B',
        'Synthese_MatNom_A_vers_B','Detail_Nom_A_vers_B','Detail_Matricule_A_vers_B','Detail_MatNom_A_vers_B'
    ]

    global_ws=wb['Synthese Globale A vers B']
    # Nom: ALPHA/MULUMBA, BETA et DELTA dans A; deux trouves dans B, DELTA absent.
    assert global_ws['B2'].value == 3
    assert global_ws['C2'].value == 2
    assert global_ws['D2'].value == 1
    assert global_ws['N2'].hyperlink is not None

    nom=wb['Synthese_Nom_A_vers_B']
    statuses={nom.cell(r,1).value: nom.cell(r,8).value for r in range(3,nom.max_row+1)}
    assert statuses['BETATEST'] == 'TROUVE_DANS_B_PAR_NOM'
    assert statuses['DELTATEST'] == 'NON_TROUVE_DANS_B'

    mat=wb['Synthese_Matricule_A_vers_B']
    mat_status={mat.cell(r,1).value: mat.cell(r,8).value for r in range(3,mat.max_row+1)}
    assert mat_status['2'] == 'TROUVE_DANS_B_DISCORDANCE_IDENTITE'

    combo=wb['Synthese_MatNom_A_vers_B']
    combo_rows={combo.cell(r,1).value:r for r in range(3,combo.max_row+1)}
    exact_key='1|ALEXANDREMULUMBA'
    assert exact_key in combo_rows
    r=combo_rows[exact_key]
    assert combo.cell(r,8).value == 'TROUVE_DANS_B_NOM_MATRICULE'
    assert combo.cell(r,2).value == 2
    assert combo.cell(r,3).value == 1
    assert combo.cell(r,19).hyperlink is not None

    # DELTA est absent de B : aucune ligne DELTA ne doit etre recopiee dans le detail Nom.
    detail_nom=wb['Detail_Nom_A_vers_B']
    values=[cell.value for row in detail_nom.iter_rows(min_row=3,values_only=False) for cell in row]
    assert 'DELTA' not in values
