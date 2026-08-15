from openpyxl import load_workbook

from controle_paie.database import Database
from controle_paie.raw_fusion_occurrence_export import OccurrenceExportRawFusionService


def _seed(db):
    with db.connect() as con:
        con.execute("CREATE TABLE raw_a (execution_id VARCHAR, matricule VARCHAR, montant DOUBLE)")
        con.execute("CREATE TABLE raw_b (execution_id VARCHAR, matricule VARCHAR, montant DOUBLE)")
        con.execute("INSERT INTO raw_a VALUES ('e1','001',100),('e1','001',120)")
        con.execute("INSERT INTO raw_b VALUES ('e2','001',200)")
        con.execute("""INSERT INTO journal_executions
            (execution_id,type_operation,fichier_source,table_source,table_destination,institution_id,regime,
             trimestre,annee,mode_chargement,lignes_lues,lignes_chargees,statut,date_fin)
            VALUES
            ('e1','IMPORT_ACCESS','a.accdb','A','raw_a','i1','CNSS','T1',2026,'append',2,2,'TERMINE',CURRENT_TIMESTAMP),
            ('e2','IMPORT_ACCESS','b.accdb','B','raw_b','i1','CARC','T1',2026,'append',1,1,'TERMINE',CURRENT_TIMESTAMP)
        """)
        rows = [
            ('p1','e1','CNSS','T1',2026,100,1,'AGENT TEST'),
            ('p2','e1','CNSS','T1',2026,120,2,'AGENT TEST'),
            ('p3','e2','CARC','T1',2026,200,1,'AGENT TEST'),
            # Ligne volontairement parasite : meme execution_id mais mauvaise periode.
            # Elle ne doit jamais entrer dans l'annexe de T1 2026.
            ('p_old','e1','CNSS','T4',2025,999,99,'AGENT TEST'),
        ]
        for line_id, execution_id, regime, quarter, year, amount, line_no, normalized_name in rows:
            con.execute("""INSERT INTO paie_standardisee
                (ligne_paie_id,execution_id,institution_id,regime,trimestre,annee,table_source,
                 matricule_source,matricule_normalise,nom,prenom,nom_normalise,section,categorie,grade,
                 unite_affectation,province,remuneration_base,transport,prime,logement,pension_rente,
                 autres_remunerations,retenues,montant_net,remuneration_brute_calculee,ligne_source)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                [line_id,execution_id,'i1',regime,quarter,year,'A' if execution_id=='e1' else 'B',
                 '001','001','Agent','Test',normalized_name,'SEC','CAT','GR','UNIT','PROV',
                 amount,0,0,0,0,0,0,amount,amount,line_no])


def _sheet_data_rows(ws):
    return sum(1 for _ in ws.iter_rows(values_only=True)) - 1


def test_occurrence_annex_keeps_every_physical_source_line(tmp_path):
    db = Database(tmp_path / 'fusion_occurrences.duckdb')
    db.migrate()
    _seed(db)
    service = OccurrenceExportRawFusionService(db)
    info = service.create_fusion(['raw_a','raw_b'],'T1',2026,'occurrences')

    check = service.occurrence_consistency(info['id'])
    assert check['physical_rows'] == 3
    assert check['aggregated_occurrences'] == 3
    assert check['agents'] == 1
    assert check['difference'] == 0
    assert check['physical_gross'] == 420
    assert check['aggregated_gross'] == 420
    assert check['physical_net'] == 420
    assert check['aggregated_net'] == 420
    assert check['ok'] is True

    folder = tmp_path / 'export'
    folder.mkdir()
    path = service.export_occurrences(info['id'], folder)
    assert path.name == '11_toutes_occurrences_confondues.xlsx'

    wb = load_workbook(path, read_only=True, data_only=True)
    assert 'Synthese agents' in wb.sheetnames
    assert 'Toutes les lignes' in wb.sheetnames
    assert 'Controle coherence' in wb.sheetnames

    detail_sheets = [ws for ws in wb.worksheets if ws.title.startswith('Toutes les lignes')]
    assert sum(_sheet_data_rows(ws) for ws in detail_sheets) == 3

    detail_rows = []
    for ws in detail_sheets:
        detail_rows.extend(list(ws.iter_rows(min_row=2, values_only=True)))
    assert {row[1] for row in detail_rows} == {'raw_a', 'raw_b'}
    assert {row[2] for row in detail_rows} == {'e1', 'e2'}
    assert {row[3] for row in detail_rows} == {'p1', 'p2', 'p3'}
    assert sorted(row[4] for row in detail_rows if row[2] == 'e1') == [1, 2]
    assert all(row[7] == 'T1' and row[8] == 2026 for row in detail_rows)
    assert all(row[23] == 3 for row in detail_rows)
    assert all(row[24] == 2 for row in detail_rows)
    assert all(row[25] == 2 for row in detail_rows)
    assert all(row[26] == 2 for row in detail_rows)
    assert all(row[33] is True for row in detail_rows)

    control = wb['Controle coherence']
    values = {row[0]: row[1] for row in control.iter_rows(min_row=2, values_only=True)}
    assert values['Periode'] == 'T1 2026'
    assert values['Lignes physiques sources'] == 3
    assert values['Occurrences agregees'] == 3
    assert values['Lignes exportees'] == 3
    assert values['Difference brut'] == 0
    assert values['Difference net'] == 0
    assert values['Controle'] == 'OK'


def test_occurrences_remain_coherent_after_reanalysis(tmp_path):
    db = Database(tmp_path / 'fusion_reanalysis.duckdb')
    db.migrate()
    _seed(db)
    service = OccurrenceExportRawFusionService(db)
    info = service.create_fusion(['raw_a','raw_b'],'T1',2026,'reanalyze')

    before = service.occurrence_consistency(info['id'])
    service.reanalyze(info['id'])
    after = service.occurrence_consistency(info['id'])

    assert before['physical_rows'] == after['physical_rows'] == 3
    assert after['aggregated_occurrences'] == 3
    assert after['physical_gross'] == after['aggregated_gross'] == 420
    assert after['physical_net'] == after['aggregated_net'] == 420
    assert after['ok'] is True


def test_strict_identity_export_filter_uses_final_status(tmp_path):
    db = Database(tmp_path / 'fusion_strict_export.duckdb')
    db.migrate()
    _seed(db)
    service = OccurrenceExportRawFusionService(db)
    info = service.create_fusion(['raw_a','raw_b'],'T1',2026,'strict')
    with db.connect() as con:
        con.execute("""UPDATE resultats_fusion_multi
            SET statut='MATRICULE_PARTAGE_IDENTITES_DIFFERENTES',identite_incoherente=TRUE
            WHERE fusion_id=?""", [info['id']])
        query, params = service._result_query('IDENTITE_INCOHERENTE')
        rows = con.execute(query, [info['id']] + params).fetchall()
    assert len(rows) == 1
    assert rows[0][0] == 'MATRICULE_PARTAGE_IDENTITES_DIFFERENTES'
    assert rows[0][17] is True
