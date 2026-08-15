import duckdb
from openpyxl import load_workbook

from controle_paie.agent_identity import normalize_matricule, normalize_name, person_key
from controle_paie.export_streaming import write_query_xlsx


def test_agent_identity_is_canonical_and_stable():
    assert normalize_matricule(" 00-123 /A ") == "00123A"
    assert normalize_name("Kábila ", " Jean-Pierre") == "KABIL AJEANPIERRE".replace(" ", "")
    assert person_key("00-123", "KABILAJEAN") == "M:00123"
    assert person_key("NU", "KABILA JEAN") == "N:KABIL AJEAN".replace(" ", "")


def test_streaming_excel_exports_all_rows(tmp_path):
    con = duckdb.connect()
    con.execute("CREATE TABLE t AS SELECT i AS id, 'Agent ' || i::VARCHAR AS nom FROM range(0, 12050) x(i)")
    path = tmp_path / "export.xlsx"
    count = write_query_xlsx(con, path, "SELECT * FROM t ORDER BY id")
    assert count == 12050
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 12051
    assert rows[1][0] == 0
    assert rows[-1][0] == 12049
