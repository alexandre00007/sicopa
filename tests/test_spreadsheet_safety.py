import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from controle_paie.database import Database
from controle_paie.reports import ReportService
from controle_paie.spreadsheet_utils import (
    EXCEL_MAX_TEXT_LENGTH,
    sanitize_excel_dataframe,
    sanitize_excel_value,
    sanitize_xml_text,
)


class SpreadsheetSafetyTests(unittest.TestCase):
    def test_illegal_xml_controls_are_replaced_without_losing_words(self):
        source = "KASONGO KASONGO L\x1eYLY"
        self.assertEqual(sanitize_xml_text(source), "KASONGO KASONGO L YLY")
        self.assertEqual(sanitize_excel_value(source), "KASONGO KASONGO L YLY")
        self.assertEqual(sanitize_excel_value("A" * (EXCEL_MAX_TEXT_LENGTH + 5)),
                         "A" * EXCEL_MAX_TEXT_LENGTH)
        self.assertEqual(sanitize_excel_value(1250), 1250)

        frame = sanitize_excel_dataframe(pd.DataFrame({"Nom\x1eagent": [source], "Montant": [1250]}))
        self.assertEqual(list(frame.columns), ["Nom agent", "Montant"])
        self.assertEqual(frame.iloc[0, 0], "KASONGO KASONGO L YLY")
        self.assertEqual(frame.iloc[0, 1], 1250)

    def test_streamed_report_accepts_existing_dirty_duckdb_values(self):
        with tempfile.TemporaryDirectory() as folder:
            database = Database(Path(folder) / "test.duckdb")
            database.migrate()
            target = Path(folder) / "annexe.xlsx"
            ReportService(database)._stream_query(
                target,
                "SELECT ?::VARCHAR AS nom, 1250::INTEGER AS montant",
                ["KASONGO KASONGO L\x1eYLY"],
                "Institution Test", "PNC", "T1", 2024, "Test de sécurité", 1,
            )

            workbook = load_workbook(target, read_only=True, data_only=True)
            try:
                row = next(workbook["Données"].iter_rows(min_row=6, max_row=6,
                                                         values_only=True))
            finally:
                workbook.close()
            self.assertEqual(row, ("KASONGO KASONGO L YLY", 1250))


if __name__ == "__main__":
    unittest.main()
