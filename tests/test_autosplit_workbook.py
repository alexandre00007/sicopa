from openpyxl import load_workbook

import controle_paie.autosplit_workbook as autosplit


def test_autosplit_write_only_workbook_repeats_header(tmp_path, monkeypatch):
    monkeypatch.setattr(autosplit, "EXCEL_MAX_DATA_ROWS", 2)
    book = autosplit.AutoSplitWriteOnlyWorkbook()
    sheet = book.create_sheet("Details")
    sheet.append(["id", "nom"])
    for index in range(5):
        sheet.append([index, f"Agent {index}"])

    target = tmp_path / "split.xlsx"
    book.save(target)

    loaded = load_workbook(target, read_only=True, data_only=True)
    assert loaded.sheetnames == ["Details", "Details_2", "Details_3"]
    assert list(loaded["Details"].values) == [("id", "nom"), (0, "Agent 0"), (1, "Agent 1")]
    assert list(loaded["Details_2"].values) == [("id", "nom"), (2, "Agent 2"), (3, "Agent 3")]
    assert list(loaded["Details_3"].values) == [("id", "nom"), (4, "Agent 4")]
