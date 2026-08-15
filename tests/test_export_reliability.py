from pathlib import Path

from openpyxl import load_workbook

from controle_paie.database import Database
from controle_paie.export_streaming import write_query_csv, write_query_xlsx
from controle_paie.explorer import DataExplorerService
from controle_paie.sql_console import SqlConsoleService


def _seed_numbers(db: Database, count: int = 12050):
    with db.connect() as con:
        con.execute("CREATE TABLE sample_export AS SELECT i id, 'ROW-'||CAST(i AS VARCHAR) label FROM range(?) t(i)", [count])


def test_streaming_xlsx_and_csv_export_all_rows(tmp_path: Path):
    db = Database(tmp_path / "exports.duckdb")
    _seed_numbers(db)
    xlsx = tmp_path / "all.xlsx"
    csv = tmp_path / "all.csv"
    with db.connect() as con:
        assert write_query_xlsx(con, xlsx, "SELECT * FROM sample_export ORDER BY id") == 12050
    with db.connect() as con:
        assert write_query_csv(con, csv, "SELECT * FROM sample_export ORDER BY id") == 12050
    wb = load_workbook(xlsx, read_only=True)
    assert sum(max(0, ws.max_row - 1) for ws in wb.worksheets) == 12050
    assert sum(1 for _ in csv.open(encoding="utf-8-sig")) - 1 == 12050
    assert not list(tmp_path.glob("*.part"))


def test_explorer_export_is_not_limited_by_display_pagination(tmp_path: Path):
    db = Database(tmp_path / "explorer.duckdb")
    _seed_numbers(db)
    service = DataExplorerService(db)
    page = service.read("sample_export", limit=500)
    assert len(page) == 500
    target = service.export(str(tmp_path / "explorer.xlsx"), table="sample_export", limit=10)
    wb = load_workbook(target, read_only=True)
    assert sum(max(0, ws.max_row - 1) for ws in wb.worksheets) == 12050


def test_sql_console_exports_full_query_without_fetchall_limit(tmp_path: Path):
    db = Database(tmp_path / "sql.duckdb")
    _seed_numbers(db)
    service = SqlConsoleService(db)
    query = "SELECT * FROM sample_export ORDER BY id"
    result = service.execute(query, display_limit=100)
    assert result["displayed"] == 100
    assert result["truncated"] is True
    target = service.export_excel(query, tmp_path / "sql.xlsx")
    wb = load_workbook(target, read_only=True)
    result_sheets = [ws for ws in wb.worksheets if ws.title.startswith("Résultats")]
    assert sum(max(0, ws.max_row - 1) for ws in result_sheets) == 12050
