from pathlib import Path

from openpyxl import load_workbook

from controle_paie.database import Database
from controle_paie.raw_fusion_complete_export import CompleteExportRawFusionService


def _seed(db):
    with db.connect() as con:
        con.execute("CREATE TABLE raw_a (execution_id VARCHAR, matricule VARCHAR, montant DOUBLE)")
        con.execute("CREATE TABLE raw_b (execution_id VARCHAR, matricule VARCHAR, montant DOUBLE)")
        con.execute("INSERT INTO raw_a VALUES ('e1','001',100),('e1','001',120)")
        con.execute("INSERT INTO raw_b VALUES ('e2','001',200)")
        con.execute("""INSERT INTO journal_executions
            (execution_id,type_operation,fichier_source,table_source,table_destination,institution_id,regime,
             trimestre,annee,mode_chargement,lignes_lues,lignes_chargees,statut,date_fin)
            VALUES
            ('e1','IMPORT_ACCESS','a.accdb','A','raw_a','i1','CNSS','T1',2026,'append',2,2,'TERMINE',CURRENT_TIMESTAMP),
            ('e2','IMPORT_ACCESS','b.accdb','B','raw_b','i1','CARC','T1',2026,'append',1,1,'TERMINE',CURRENT_TIMESTAMP)
        """)
        for row in [
            ('p1','e1','CNSS',100,1),('p2','e1','CNSS',120,2),('p3','e2','CARC',200,1)
        ]:
            line_id, execution_id, regime, amount, line_no = row
            con.execute("""INSERT INTO paie_standardisee
                (ligne_paie_id,execution_id,institution_id,regime,trimestre,annee,table_source,
                 matricule_source,matricule_normalise,nom,prenom,nom_normalise,section,categorie,grade,
                 unite_affectation,province,remuneration_base,transport,prime,logement,pension_rente,
                 autres_remunerations,retenues,montant_net,remuneration_brute_calculee,ligne_source)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                [line_id,execution_id,'i1',regime,'T1',2026,'A' if execution_id=='e1' else 'B',
                 '001','001','Agent','Test','AGENT TEST','SEC','CAT','GR','UNIT','PROV',
                 amount,0,0,0,0,0,0,amount,amount,line_no])


def _count_rows(wb, prefix):
    total = 0
    for ws in wb.worksheets:
        if ws.title.startswith(prefix):
            total += max(0, sum(1 for _ in ws.iter_rows(values_only=True)) - 1)
    return total


def test_complete_export_uses_partitioned_annexes(tmp_path):
    db = Database(tmp_path / 'fusion_partitioned.duckdb')
    db.migrate()
    _seed(db)
    service = CompleteExportRawFusionService(db)
    info = service.create_fusion(['raw_a','raw_b'],'T1',2026,'partitioned')

    parent = tmp_path / 'exports'
    parent.mkdir()
    folder = Path(service.export_all(info['id'], parent))

    annex11 = folder / '11_toutes_occurrences_confondues.xlsx'
    annex12 = folder / '12_synthese_occurrences_agents_a_risque.xlsx'
    assert annex11.exists() and annex12.exists()

    wb11 = load_workbook(annex11, read_only=True, data_only=True)
    control11 = {r[0]: r[1] for r in wb11['Controle coherence'].iter_rows(min_row=2, values_only=True)}
    assert control11['Mode export'] == 'PARTITIONNE_PAR_EXECUTION'
    assert control11['Executions traitees'] == 2
    assert control11['Lignes physiques sources'] == 3
    assert control11['Lignes exportees'] == 3
    assert control11['Controle'] == 'OK'
    assert _count_rows(wb11, 'Toutes les lignes') == 3

    wb12 = load_workbook(annex12, read_only=True, data_only=True)
    control12 = {r[0]: r[1] for r in wb12['Controle'].iter_rows(min_row=2, values_only=True)}
    assert control12['Mode export'] == 'PARTITIONNE_PAR_EXECUTION'
    assert control12['Organisation'] == 'SYNTHESE_CONCISE_PUIS_OCCURRENCES'
    assert control12['Nature categories'] == 'NON_EXCLUSIVES'
    assert control12['Executions traitees'] == 2
    assert control12['Synthese matricules'] == 1
    assert control12['Synthese noms'] == 1
    assert control12['Synthese multi-regimes'] == 1
    assert control12['Controle'] == 'OK'

    names = wb12.sheetnames
    assert names[:4] == [
        '00_Synthese_anomalies',
        '01_Synthese_matricule',
        '02_Synthese_nom',
        '03_Synthese_multi_regimes',
    ]
    assert 'D01_Occ_matricule' in names
    assert 'D02_Occ_nom' in names
    assert 'D05_Occ_multi_regimes' in names

    summary_rows = list(wb12['00_Synthese_anomalies'].iter_rows(min_row=2, values_only=True))
    labels = {row[0] for row in summary_rows}
    assert 'ANOMALIE_PAR_MATRICULE' in labels
    assert 'ANOMALIE_PAR_NOM' in labels
    assert 'MULTI_REGIME_REPETE' in labels

    assert _count_rows(wb12, 'D01_Occ_matricule') == 3
    assert _count_rows(wb12, 'D02_Occ_nom') == 3
    assert _count_rows(wb12, 'D05_Occ_multi_regimes') == 3
