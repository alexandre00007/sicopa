from __future__ import annotations

import re
import os
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from .database import Database

Progress = Optional[Callable[[int, str], None]]

class ReportService:
    """Generate bounded-memory Excel deliverables from DuckDB cursors."""
    def __init__(self, database: Database): self.db = database

    def export(self, output: str, institution_id: str, regime: str, quarter: str, year: int, progress: Progress = None) -> Path:
        target=Path(output);target.parent.mkdir(parents=True,exist_ok=True);progress and progress(5,"Préparation de la synthèse")
        params=[institution_id,regime,quarter,year]
        with self.db.connect() as con:
            found=con.execute("SELECT nom_officiel FROM institutions WHERE institution_id=?",[institution_id]).fetchone();institution=found[0] if found else institution_id
            rows=con.execute("""SELECT statut_rapprochement,COUNT(*),COALESCE(SUM(masse_financiere_controlee),0),COALESCE(SUM(impact_potentiel),0),COALESCE(SUM(impact_confirme),0) FROM resultats_rapprochement WHERE institution_id=? AND regime=? AND trimestre=? AND annee=? GROUP BY statut_rapprochement ORDER BY statut_rapprochement""",params).fetchall()
        wb=Workbook();ws=wb.active;ws.title="Synthèse";self._title(ws,institution,regime,quarter,year)
        ws.append(["Catégorie","Enregistrements","Masse contrôlée","Impact potentiel","Impact confirmé"]);self._style_header(ws[4])
        for row in rows:ws.append(list(row))
        ws.freeze_panes="A6";ws.auto_filter.ref=f"A5:E{ws.max_row}"
        for letter,width in zip("ABCDE",[36,18,22,22,22]):ws.column_dimensions[letter].width=width
        note=wb.create_sheet("Méthodologie");note.append(["MÉTHODOLOGIE DU CONTRÔLE"]);note["A1"].font=Font(bold=True,size=15,color="12355B")
        for text in ["Une ligne de paie reçoit une seule classification finale.","Le matricule est prioritaire, puis le nom normalisé.","L’impact potentiel doit être confirmé par un contrôle métier.","Les détails sont fournis dans le dossier annexes."]:note.append([text])
        note.column_dimensions["A"].width=100
        temporary=target.with_name(f".{target.stem}.part.xlsx")
        try:
            wb.save(temporary)
            os.replace(temporary,target)
        finally:
            if temporary.exists():temporary.unlink()
        progress and progress(20,f"Fichier généré : {target.name}");return target

    def generate_package(self, root: str, institution_id: str, regime: str, quarter: str, year: int, progress: Progress = None) -> Path:
        """Generate the historical listing, declarative and comparison annex families."""
        with self.db.connect() as con:
            found=con.execute("SELECT nom_officiel FROM institutions WHERE institution_id=?",[institution_id]).fetchone()
            institution=found[0] if found else institution_id
        slug=re.sub(r"[^A-Za-z0-9]+","_",institution).strip("_").lower();stamp=datetime.now().strftime("%Y%m%d_%H%M%S")
        package=Path(root)/f"{year}_{quarter}_{regime}_{slug}_{stamp}";package.mkdir(parents=True,exist_ok=True)
        progress and progress(1,"Initialisation du dossier de résultats")
        specs=self._historical_annex_specs(institution_id,regime,quarter,year)
        summaries=[];total=max(1,len(specs))
        for index,spec in enumerate(specs,1):
            folder=package/spec["folder"];folder.mkdir(parents=True,exist_ok=True);target=folder/spec["filename"]
            effectif_query=self._effectif_query(spec);effectif_folder_name=spec["folder"].replace("annexes_","effectifs_")
            effectif_filename=f"effectif_{Path(spec['filename']).stem}.xlsx";effectif_folder=package/effectif_folder_name;effectif_folder.mkdir(parents=True,exist_ok=True);effectif_target=effectif_folder/effectif_filename
            with self.db.connect() as con:
                count,impact=con.execute(f'SELECT COUNT(*),COALESCE(SUM(impact_calcule),0) FROM ({spec["query"]}) donnees',spec["params"]).fetchone()
                concerned=con.execute(f"SELECT COUNT(*) FROM ({effectif_query}) effectifs",spec["params"]).fetchone()[0]
            base=22+73*(index-1)/total;span=73/total;detail_span=span*0.62
            progress and progress(int(base),f'Annexe {index}/{total} : {spec["label"]} ({count} lignes, {concerned} concernés)')
            self._stream_query(target,spec["query"],spec["params"],institution,regime,quarter,year,spec["label"],count,progress,base,detail_span)
            progress and progress(int(base+detail_span),f"Fichier généré : {spec['folder']}/{spec['filename']}")
            self._stream_query(effectif_target,effectif_query,spec["params"],institution,regime,quarter,year,f"Effectif unique — {spec['label']}",concerned,progress,base+detail_span,span-detail_span)
            progress and progress(int(base+span),f"Fichier généré : {effectif_folder_name}/{effectif_filename}")
            summaries.append({**spec,"count":count,"concerned":concerned,"impact":impact,"relative_path":str(Path(spec["folder"])/spec["filename"]),"effectif_path":str(Path(effectif_folder_name)/effectif_filename)})
        progress and progress(96,"Finalisation du rapport et des liens")
        report=package/"rapport_final.xlsx";filters=self.db.list_treatment_filters(institution_id,regime);self._write_historical_report(report,institution,regime,quarter,year,summaries,filters)
        progress and progress(99,"Fichier généré : rapport_final.xlsx")
        progress and progress(100,f"{len(specs)} annexes, {len(specs)} fichiers d’effectifs et rapport final terminés")
        return package

    def _historical_annex_specs(self,institution_id: str,regime: str,quarter: str,year: int) -> list[dict]:
        scope=[institution_id,regime,quarter,year]
        def base(alias, table="paie_standardisee"):
            return f"{alias}.institution_id=? AND {alias}.regime=? AND {alias}.trimestre=? AND {alias}.annee=?"
        def selected(alias):
            clause,values=self.db.payroll_filter_clause(institution_id,regime,alias)
            return base(alias)+clause,scope+values
        def valid(key,alias=""):
            prefix=f"{alias}." if alias else ""
            return f"{prefix}{key} NOT IN ('','NU')" if key=="matricule_normalise" else f"{prefix}{key}<>''"
        def spec(group,label,filename,query,params):
            return {"group":group,"folder":group,"label":label,"filename":filename,"query":query,"params":params}

        items=[];ps,pp=selected("p")
        items += [
          spec("annexes_listing","Données du listing de paie filtré","donnees_fichier_paie.xlsx",f"SELECT p.*,p.remuneration_brute_calculee impact_calcule FROM paie_standardisee p WHERE {ps}",pp),
          spec("annexes_listing","Matricules manquants ou non exploitables (NU)","matricules_manquants.xlsx",f"SELECT p.*,p.remuneration_brute_calculee impact_calcule FROM paie_standardisee p WHERE {ps} AND COALESCE(p.matricule_normalise,'') IN ('','NU')",pp),
          spec("annexes_listing","Rapport de validation du listing","rapport_validation.xlsx",f"SELECT p.*,p.remuneration_brute_calculee impact_calcule FROM paie_standardisee p WHERE {ps} AND (COALESCE(p.matricule_normalise,'') IN ('','NU') OR COALESCE(p.nom_normalise,'')='')",pp),
          spec("annexes_listing","Projection des identifiants du listing","projection_identifiants.xlsx",f"SELECT p.institution_id,p.matricule_source,p.matricule_normalise,p.nom,p.prenom,p.nom_normalise,p.section,p.remuneration_brute_calculee,p.ligne_source,p.remuneration_brute_calculee impact_calcule FROM paie_standardisee p WHERE {ps}",pp),
          spec("annexes_listing","Enregistrements uniques par matricule","enregistrements_uniques_par_matricule.xlsx",f"SELECT p.*,p.remuneration_brute_calculee impact_calcule FROM paie_standardisee p WHERE {ps} AND {valid('matricule_normalise','p')} QUALIFY ROW_NUMBER() OVER(PARTITION BY p.matricule_normalise ORDER BY p.ligne_source)=1",pp),
          spec("annexes_listing","Doublons par matricule hors NU","doublons_par_matricule.xlsx",f"SELECT p.*,p.remuneration_brute_calculee impact_calcule FROM paie_standardisee p WHERE {ps} AND {valid('matricule_normalise','p')} QUALIFY COUNT(*) OVER(PARTITION BY p.matricule_normalise)>1",pp),
          spec("annexes_listing","Matricules partagés uniques hors NU","matricules_partages_uniques.xlsx",f"SELECT p.*,0::DECIMAL(38,2) impact_calcule FROM paie_standardisee p WHERE {ps} AND {valid('matricule_normalise','p')} QUALIFY COUNT(*) OVER(PARTITION BY p.matricule_normalise)>1 AND ROW_NUMBER() OVER(PARTITION BY p.matricule_normalise ORDER BY p.ligne_source)=1",pp),
          spec("annexes_listing","Enregistrements uniques par nom","enregistrements_uniques_par_nom.xlsx",f"SELECT p.*,p.remuneration_brute_calculee impact_calcule FROM paie_standardisee p WHERE {ps} AND {valid('nom_normalise','p')} QUALIFY ROW_NUMBER() OVER(PARTITION BY p.nom_normalise ORDER BY p.ligne_source)=1",pp),
          spec("annexes_listing","Doublons par nom","doublons_par_nom.xlsx",f"SELECT p.*,p.remuneration_brute_calculee impact_calcule FROM paie_standardisee p WHERE {ps} AND {valid('nom_normalise','p')} QUALIFY COUNT(*) OVER(PARTITION BY p.nom_normalise)>1",pp),
          spec("annexes_listing","Noms partagés uniques","noms_partages_uniques.xlsx",f"SELECT p.*,0::DECIMAL(38,2) impact_calcule FROM paie_standardisee p WHERE {ps} AND {valid('nom_normalise','p')} QUALIFY COUNT(*) OVER(PARTITION BY p.nom_normalise)>1 AND ROW_NUMBER() OVER(PARTITION BY p.nom_normalise ORDER BY p.ligne_source)=1",pp),
        ]
        # Cohorte contrôlée : agents déclarés ET présents dans le listing filtré.
        # Recherche : tout le listing du trimestre, à l'exception exacte du listing filtré.
        ds=base("d","declaratif_standardise");cs,cp=selected("c")
        outside_filter,outside_values=self.db.payroll_filter_clause(institution_id,regime,"o")
        outside_clause=f"NOT (o.institution_id=?{outside_filter})"
        elsewhere_queries={}
        for key,label in [("matricule_normalise","matricule"),("nom_normalise","nom")]:
            cohort=f"""EXISTS(SELECT 1 FROM declaratif_standardise d WHERE {ds} AND {valid(key,'d')} AND d.{key}=o.{key}
                AND EXISTS(SELECT 1 FROM paie_standardisee c WHERE {cs} AND {valid(key,'c')} AND c.{key}=d.{key}))"""
            query=f"""SELECT o.*,o.remuneration_brute_calculee impact_calcule FROM paie_standardisee o
                WHERE o.regime=? AND o.trimestre=? AND o.annee=? AND {valid(key,'o')}
                AND {cohort} AND {outside_clause}"""
            params=[regime,quarter,year]+scope+cp+[institution_id]+outside_values
            elsewhere_queries[key]=(query,params)
            items.append(spec("annexes_listing",f"Agents déclarés du listing filtré retrouvés hors périmètre par {label}",f"agents_autres_sections_par_{label}.xlsx",query,params))
            items.append(spec("annexes_listing",f"Agents déclarés du listing filtré payés hors périmètre avec rémunération positive par {label}",f"agents_autres_sections_par_{label}_filtres.xlsx",query+" AND o.remuneration_brute_calculee>0",params))

        items += [
          spec("annexes_declaratif","Données déclaratives","donnees_declaratives.xlsx",f"SELECT d.*,0::DECIMAL(38,2) impact_calcule FROM declaratif_standardise d WHERE {ds}",scope),
          spec("annexes_declaratif","Matricules déclaratifs manquants ou non exploitables (NU)","matricules_manquants_declaratif.xlsx",f"SELECT d.*,0::DECIMAL(38,2) impact_calcule FROM declaratif_standardise d WHERE {ds} AND COALESCE(d.matricule_normalise,'') IN ('','NU')",scope),
          spec("annexes_declaratif","Enregistrements déclaratifs uniques par matricule","declaratifs_uniques_par_matricule.xlsx",f"SELECT d.*,0::DECIMAL(38,2) impact_calcule FROM declaratif_standardise d WHERE {ds} AND {valid('matricule_normalise','d')} QUALIFY ROW_NUMBER() OVER(PARTITION BY d.matricule_normalise ORDER BY d.ligne_source)=1",scope),
          spec("annexes_declaratif","Doublons déclaratifs par matricule hors NU","doublons_declaratifs_par_matricule.xlsx",f"SELECT d.*,0::DECIMAL(38,2) impact_calcule FROM declaratif_standardise d WHERE {ds} AND {valid('matricule_normalise','d')} QUALIFY COUNT(*) OVER(PARTITION BY d.matricule_normalise)>1",scope),
          spec("annexes_declaratif","Enregistrements déclaratifs uniques par nom","declaratifs_uniques_par_nom.xlsx",f"SELECT d.*,0::DECIMAL(38,2) impact_calcule FROM declaratif_standardise d WHERE {ds} AND {valid('nom_normalise','d')} QUALIFY ROW_NUMBER() OVER(PARTITION BY d.nom_normalise ORDER BY d.ligne_source)=1",scope),
          spec("annexes_declaratif","Doublons déclaratifs par nom","doublons_declaratifs_par_nom.xlsx",f"SELECT d.*,0::DECIMAL(38,2) impact_calcule FROM declaratif_standardise d WHERE {ds} AND {valid('nom_normalise','d')} QUALIFY COUNT(*) OVER(PARTITION BY d.nom_normalise)>1",scope),
        ]
        p0s,p0p=selected("p0")
        for key,label in [("matricule_normalise","matricule"),("nom_normalise","nom")]:
            other,other_params=elsewhere_queries[key];other=other+" AND o.remuneration_brute_calculee>0"
            same=f"""SELECT p0.*,p0.remuneration_brute_calculee impact_calcule FROM paie_standardisee p0 WHERE {p0s} AND p0.remuneration_brute_calculee>0 AND {valid(key,'p0')} AND EXISTS(SELECT 1 FROM declaratif_standardise d WHERE {ds} AND {valid(key,'d')} AND d.{key}=p0.{key})"""
            missing=f"""SELECT p0.*,p0.remuneration_brute_calculee impact_calcule FROM paie_standardisee p0 WHERE {p0s} AND p0.remuneration_brute_calculee>0 AND NOT EXISTS(SELECT 1 FROM declaratif_standardise d WHERE {ds} AND {valid(key,'d')} AND d.{key}=p0.{key})"""
            items.append(spec("annexes_comparaisons",f"Agents déclarés du listing filtré payés hors périmètre par {label}",f"declares_payes_ailleurs_par_{label}.xlsx",other,other_params))
            items.append(spec("annexes_comparaisons",f"Déclarés présents dans le listing filtré par {label}",f"declares_presents_listing_par_{label}.xlsx",same,p0p+scope))
            items.append(spec("annexes_comparaisons",f"Non déclarés mais présents dans le listing filtré par {label}",f"non_declares_payes_par_{label}.xlsx",missing,p0p+scope))
        return items

    def _effectif_query(self,spec: dict) -> str:
        """Return one representative row per person concerned by an annex rubric."""
        label=spec["label"].lower()
        if "par matricule" in label or "matricules partagés" in label:
            key="matricule_normalise"
        elif "par nom" in label or "noms partagés" in label:
            key="nom_normalise"
        else:
            key="CASE WHEN COALESCE(matricule_normalise,'') NOT IN ('','NU') THEN 'M:'||matricule_normalise WHEN COALESCE(nom_normalise,'')<>'' THEN 'N:'||nom_normalise ELSE 'L:'||COALESCE(institution_id,'')||':'||CAST(ligne_source AS VARCHAR) END"
        return f"SELECT * FROM ({spec['query']}) effectif_source QUALIFY ROW_NUMBER() OVER(PARTITION BY {key} ORDER BY ligne_source)=1"

    def _write_historical_report(self,path: Path,institution: str,regime: str,quarter: str,year: int,summaries: list[dict],filters: list[tuple] = None) -> None:
        wb=Workbook();wb.remove(wb.active)
        overview=wb.create_sheet("Vue d’ensemble");self._title(overview,institution,regime,quarter,year)
        overview.append(["Source du contrôle","Annexes","Enregistrements","Concernés","Masse / impact"]);self._style_header(overview[4])
        group_labels=[("annexes_listing","Listing de paie"),("annexes_declaratif","Liste déclarative"),("annexes_comparaisons","Comparaisons")]
        for code,label in group_labels:
            rows=[x for x in summaries if x["group"]==code];overview.append([label,len(rows),sum(x["count"] for x in rows),sum(x["concerned"] for x in rows),sum(x["impact"] for x in rows)])
        overview.column_dimensions["A"].width=34;overview.column_dimensions["B"].width=15;overview.column_dimensions["C"].width=20;overview.column_dimensions["D"].width=18;overview.column_dimensions["E"].width=24
        groups=[("annexes_listing","Fichier listing de paie"),("annexes_declaratif","Liste déclarative"),("annexes_comparaisons","Comparaisons")]
        for group,title in groups:
            ws=wb.create_sheet(title);self._title(ws,institution,regime,quarter,year)
            ws.append(["Catégories","Enregistrements","Masse / impact contrôlé","Annexe détaillée","Nombre de concernés","Lien de l’effectif unique"]);self._style_header(ws[4])
            for item in [x for x in summaries if x["group"]==group]:
                ws.append([item["label"],item["count"],item["impact"],item["relative_path"],item["concerned"],item["effectif_path"]])
                detail_cell=ws.cell(ws.max_row,4);detail_cell.hyperlink=item["relative_path"];detail_cell.style="Hyperlink"
                effectif_cell=ws.cell(ws.max_row,6);effectif_cell.hyperlink=item["effectif_path"];effectif_cell.style="Hyperlink"
            ws.freeze_panes="A5";ws.auto_filter.ref=f"A4:F{ws.max_row}";ws.column_dimensions["A"].width=58;ws.column_dimensions["B"].width=18;ws.column_dimensions["C"].width=24;ws.column_dimensions["D"].width=55;ws.column_dimensions["E"].width=22;ws.column_dimensions["F"].width=58
        filter_ws=wb.create_sheet("Filtres du listing");filter_ws.append(["Colonne","Opérateur","Contenu"]);self._style_header(filter_ws[1])
        for _,column,operator,value in (filters or []):filter_ws.append([column,operator,value])
        if not filters:filter_ws.append(["Aucun filtre","—","Tout le listing de l’institution est traité"])
        filter_ws.column_dimensions["A"].width=34;filter_ws.column_dimensions["B"].width=20;filter_ws.column_dimensions["C"].width=50
        note=wb.create_sheet("Note méthodologique");note.append(["NOTE MÉTHODOLOGIQUE"]);note["A1"].font=Font(bold=True,size=15,color="12355B")
        for line in [f"Institution : {institution}",f"Régime : {regime} — {quarter} {year}","La cohorte de contrôle contient uniquement les agents déclarés présents dans le listing filtré.","Cette cohorte est recherchée dans tout le listing trimestriel hors du périmètre filtré, y compris dans la même institution.","Le nombre de concernés correspond à l’effectif unique de chaque rubrique; son fichier est lié dans le rapport.","NU, N.U et leurs variantes sont exclus des doublons de matricule.","Les annexes par matricule et par nom sont des méthodes de contrôle parallèles et ne doivent pas être additionnées.","Les impacts sont potentiels et doivent être confirmés par un contrôle métier."]:note.append([line])
        note.column_dimensions["A"].width=120
        temporary=path.with_name(f".{path.stem}.part.xlsx")
        try:wb.save(temporary);os.replace(temporary,path)
        finally:
            if temporary.exists():temporary.unlink()

    def _stream_query(self,path: Path,query: str,params: list,institution: str,regime: str,quarter: str,year: int,status: str,total_rows: int = 0,progress: Progress = None,progress_base: float = 0,progress_span: float = 0) -> None:
        wb=Workbook(write_only=True);ws=wb.create_sheet("Données");ws.freeze_panes="A6"
        title=WriteOnlyCell(ws,"ANNEXE DE CONTRÔLE DE LA PAIE");title.font=Font(bold=True,size=15,color="FFFFFF");title.fill=PatternFill("solid",fgColor="12355B")
        ws.append([title]);ws.append([f"Institution : {institution}"]);ws.append([f"Régime : {regime} — Période : {quarter} {year}"]);ws.append([f"Catégorie : {status}"])
        with self.db.connect() as con:
            cursor=con.execute(query,params);headers=[item[0] for item in cursor.description];styled=[]
            for header in headers:
                cell=WriteOnlyCell(ws,header);cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="1677FF");cell.alignment=Alignment(horizontal="center");styled.append(cell)
            ws.append(styled)
            written=0
            while True:
                batch=cursor.fetchmany(1000)
                if not batch:break
                for row in batch:ws.append(list(row))
                written += len(batch)
                if progress and total_rows:
                    value=int(progress_base+progress_span*min(1,written/total_rows))
                    progress(value,f"{status} : {written:,}/{total_rows:,} lignes".replace(","," "))
        temporary=path.with_name(f".{path.stem}.part.xlsx")
        try:
            wb.save(temporary)
            os.replace(temporary,path)
        finally:
            if temporary.exists():temporary.unlink()

    @staticmethod
    def _title(ws,institution: str,regime: str,quarter: str,year: int) -> None:
        ws.merge_cells("A1:F1");cell=ws["A1"];cell.value="RAPPORT FINAL DE CONTRÔLE ET RAPPROCHEMENT DE LA PAIE";cell.fill=PatternFill("solid",fgColor="12355B");cell.font=Font(color="FFFFFF",bold=True,size=15);cell.alignment=Alignment(horizontal="center");ws.row_dimensions[1].height=30;ws["A2"]=f"Institution : {institution}";ws["A3"]=f"Régime : {regime} — Période : {quarter} {year}"

    @staticmethod
    def _style_header(cells) -> None:
        for cell in cells:cell.fill=PatternFill("solid",fgColor="1677FF");cell.font=Font(color="FFFFFF",bold=True);cell.alignment=Alignment(horizontal="center")
