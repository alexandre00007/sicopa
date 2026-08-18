from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from .export_streaming import EXCEL_MAX_DATA_ROWS, atomic_save_workbook
from .spreadsheet_utils import sanitize_excel_row
from .raw_fusion_occurrence_export_low_memory import LowMemoryOccurrenceExportRawFusionService
from .raw_fusion_risk_export import RISK_DETAIL_HEADERS, RISK_SUMMARY_HEADERS
from .raw_fusion_risk_export_low_memory import LowMemoryRawFusionRiskExporter


class _WorkbookPartitionWriter:
    """Ecrit plusieurs curseurs successifs dans une meme famille de feuilles Excel."""

    def __init__(self, book: Workbook, headers, sheet_name: str):
        self.book = book
        self.headers = list(headers)
        self.base_name = sheet_name
        self.sheet_index = 0
        self.rows_in_sheet = 0
        self.total = 0
        self.sheet = None
        self._new_sheet()

    def _new_sheet(self):
        self.sheet_index += 1
        suffix = "" if self.sheet_index == 1 else f"_{self.sheet_index}"
        name = (self.base_name[: 31 - len(suffix)] + suffix)[:31]
        self.sheet = self.book.create_sheet(name)
        self.sheet.append(list(sanitize_excel_row(self.headers)))
        self.rows_in_sheet = 0

    def append_cursor(self, cursor, chunk_size: int = 1000) -> int:
        added = 0
        while True:
            chunk = cursor.fetchmany(max(100, int(chunk_size)))
            if not chunk:
                break
            for row in chunk:
                if self.rows_in_sheet >= EXCEL_MAX_DATA_ROWS:
                    self._new_sheet()
                self.sheet.append(list(sanitize_excel_row(row)))
                self.rows_in_sheet += 1
                self.total += 1
                added += 1
        return added


def _person_key_sql(alias: str = "p") -> str:
    return f"""CASE
        WHEN COALESCE({alias}.matricule_normalise,'') NOT IN ('','NU')
            THEN 'M:'||{alias}.matricule_normalise
        WHEN COALESCE({alias}.nom_normalise,'')<>''
            THEN 'N:'||{alias}.nom_normalise
        ELSE 'L:'||{alias}.ligne_paie_id
    END"""


def _configure_partitioned(con):
    con.execute("SET preserve_insertion_order=false")
    con.execute("SET threads=1")


def _fusion_scope(con, fusion_id: str):
    row = con.execute("SELECT trimestre,annee FROM fusions_raw WHERE fusion_id=?", [fusion_id]).fetchone()
    if not row:
        raise ValueError(f"Fusion introuvable : {fusion_id}")
    return row[0], int(row[1])


def _fusion_executions(con, fusion_id: str):
    return con.execute("""SELECT s.execution_id,MIN(s.table_source) AS table_source
        FROM sources_fusion_raw s
        WHERE s.fusion_id=? AND s.execution_id IS NOT NULL
        GROUP BY s.execution_id ORDER BY s.execution_id""", [fusion_id]).fetchall()


def _prepare_compact_identity_stats(con, fusion_id: str, quarter: str, year: int, executions) -> None:
    """Construit seulement une ligne partielle par identite/execution, jamais toutes les lignes physiques."""
    con.execute("DROP TABLE IF EXISTS tmp_sicorpa_identity_parts")
    con.execute("DROP TABLE IF EXISTS tmp_sicorpa_identity_stats")
    con.execute("""CREATE TEMP TABLE tmp_sicorpa_identity_parts (
        person_key VARCHAR, execution_id VARCHAR, table_source VARCHAR,
        nom_normalise VARCHAR, matricule_normalise VARCHAR,
        has_matricule_vide BOOLEAN, has_matricule_nu BOOLEAN,
        has_matricule_non_exploitable BOOLEAN
    )""")
    key = _person_key_sql("p")
    for execution_id, table_source in executions:
        con.execute(f"""INSERT INTO tmp_sicorpa_identity_parts
            SELECT {key}, ?, ?,
                   MIN(NULLIF(p.nom_normalise,'')),MIN(NULLIF(p.matricule_normalise,'')),
                   BOOL_OR(COALESCE(NULLIF(TRIM(p.matricule_source),''),'')=''),
                   BOOL_OR(UPPER(TRIM(COALESCE(p.matricule_source,'')))='NU'),
                   BOOL_OR(UPPER(TRIM(COALESCE(p.matricule_source,''))) IN ('NULL','N/A','NA','NEANT','NÉANT','INCONNU','NONE'))
            FROM paie_standardisee p
            WHERE p.execution_id=? AND p.trimestre=? AND p.annee=?
            GROUP BY {key}""", [execution_id, table_source or "", execution_id, quarter, year])

    con.execute("""CREATE TEMP TABLE tmp_sicorpa_identity_stats AS
        SELECT person_key,
               COUNT(DISTINCT execution_id) AS nb_executions,
               COUNT(DISTINCT COALESCE(table_source,'')) AS nb_tables,
               STRING_AGG(DISTINCT NULLIF(nom_normalise,''),' | ') AS noms_distincts,
               STRING_AGG(DISTINCT NULLIF(matricule_normalise,''),' | ') AS matricules_distincts,
               BOOL_OR(has_matricule_vide) AS has_matricule_vide,
               BOOL_OR(has_matricule_nu) AS has_matricule_nu,
               BOOL_OR(has_matricule_non_exploitable) AS has_matricule_non_exploitable
        FROM tmp_sicorpa_identity_parts GROUP BY person_key""")
    try:
        con.execute("CREATE INDEX idx_tmp_identity_stats_key ON tmp_sicorpa_identity_stats(person_key)")
    except Exception:
        pass


def _cleanup_compact_stats(con):
    con.execute("DROP TABLE IF EXISTS tmp_sicorpa_identity_stats")
    con.execute("DROP TABLE IF EXISTS tmp_sicorpa_identity_parts")


class PartitionedOccurrenceExportRawFusionService(LowMemoryOccurrenceExportRawFusionService):
    """Annexe 11 : lecture et ecriture execution par execution, memoire bornee."""

    def _execution_detail_query(self) -> str:
        key = _person_key_sql("p")
        return f"""SELECT
            r.fusion_id,?,p.execution_id,p.ligne_paie_id,p.ligne_source,p.regime,p.institution_id,
            p.trimestre,p.annee,p.matricule_source,p.matricule_normalise,p.nom,p.prenom,
            p.nom_normalise,p.section,p.categorie,p.grade,p.unite_affectation,p.province,
            p.remuneration_brute_calculee,p.montant_net,r.statut,r.nb_regimes,r.occurrences,
            GREATEST(r.occurrences-1,0),s.nb_executions,s.nb_tables,r.regimes,r.institutions,
            s.noms_distincts,s.matricules_distincts,
            COALESCE(d.doublon_matricule,FALSE),COALESCE(d.doublon_nom,FALSE),
            r.paiement_multi_regime,r.paiement_multiple_meme_regime,r.identite_incoherente,
            CASE
                WHEN r.identite_incoherente THEN 'IDENTITE_INCOHERENTE'
                WHEN COALESCE(d.doublon_matricule,FALSE) AND COALESCE(d.doublon_nom,FALSE) THEN 'MATRICULE_ET_NOM_REPETES'
                WHEN r.nb_regimes>1 THEN 'MEME_IDENTITE_MULTI_REGIME'
                WHEN r.paiement_multiple_meme_regime THEN 'PAIEMENT_MULTIPLE_MEME_REGIME'
                WHEN COALESCE(d.doublon_matricule,FALSE) THEN 'MATRICULE_REPETE'
                WHEN COALESCE(d.doublon_nom,FALSE) THEN 'NOM_REPETE'
                WHEN r.occurrences>1 THEN 'MEME_IDENTITE_REPETEE'
                ELSE 'OCCURRENCE_UNIQUE'
            END,r.diagnostic
        FROM paie_standardisee p
        JOIN resultats_fusion_multi r ON r.fusion_id=? AND r.person_key=({key})
        JOIN tmp_sicorpa_identity_stats s ON s.person_key=r.person_key
        LEFT JOIN resultats_fusion_doublons d ON d.fusion_id=r.fusion_id AND d.person_key=r.person_key
        WHERE p.execution_id=? AND p.trimestre=? AND p.annee=?"""

    def export_occurrences(self, fusion_id: str, folder: str | Path, progress=None) -> Path:
        folder = Path(folder)
        target = folder / "11_toutes_occurrences_confondues.xlsx"
        book = Workbook(write_only=True)
        with self.db.connect() as con:
            _configure_partitioned(con)
            quarter, year = _fusion_scope(con, fusion_id)
            executions = _fusion_executions(con, fusion_id)
            _prepare_compact_identity_stats(con, fusion_id, quarter, year, executions)
            try:
                aggregated, aggregated_gross, aggregated_net = con.execute("""SELECT
                    COALESCE(SUM(occurrences),0),COALESCE(SUM(masse_brute),0),COALESCE(SUM(masse_net),0)
                    FROM resultats_fusion_multi WHERE fusion_id=?""", [fusion_id]).fetchone()
                agents = int(con.execute("SELECT COUNT(*) FROM resultats_fusion_multi WHERE fusion_id=?", [fusion_id]).fetchone()[0] or 0)
                physical_rows = 0; physical_gross = 0.0; physical_net = 0.0
                for execution_id, _table_source in executions:
                    row = con.execute("""SELECT COUNT(*),COALESCE(SUM(remuneration_brute_calculee),0),COALESCE(SUM(montant_net),0)
                        FROM paie_standardisee WHERE execution_id=? AND trimestre=? AND annee=?""",
                        [execution_id, quarter, year]).fetchone()
                    physical_rows += int(row[0] or 0); physical_gross += float(row[1] or 0); physical_net += float(row[2] or 0)
                aggregated = int(aggregated or 0); aggregated_gross=float(aggregated_gross or 0); aggregated_net=float(aggregated_net or 0)
                gross_diff=physical_gross-aggregated_gross; net_diff=physical_net-aggregated_net
                if physical_rows != aggregated or abs(gross_diff)>0.01 or abs(net_diff)>0.01:
                    raise ValueError(f"Incoherence de la fusion : lignes {physical_rows}/{aggregated}, ecart brut {gross_diff:.2f}, ecart net {net_diff:.2f}.")

                progress and progress(91, "Annexe 11 : synthese par agent")
                summary_writer = _WorkbookPartitionWriter(book, self.SUMMARY_HEADERS, "Synthese agents")
                summary_writer.append_cursor(con.execute(self._summary_query(), [fusion_id]), 1000)

                detail_writer = _WorkbookPartitionWriter(book, self.OCCURRENCE_HEADERS, "Toutes les lignes")
                total_exec = max(1, len(executions))
                for index, (execution_id, table_source) in enumerate(executions, 1):
                    progress and progress(92 + int(4 * index / total_exec), f"Annexe 11 : source {index}/{total_exec}")
                    cursor = con.execute(self._execution_detail_query(), [table_source or "", fusion_id, execution_id, quarter, year])
                    detail_writer.append_cursor(cursor, 1000)
                if detail_writer.total != physical_rows:
                    raise ValueError(f"Export incomplet des occurrences : {detail_writer.total} lignes sur {physical_rows} attendues.")

                control=book.create_sheet("Controle coherence")
                for row in [
                    ["Indicateur","Valeur"],["Mode export","PARTITIONNE_PAR_EXECUTION"],["Periode",f"{quarter} {year}"],
                    ["Executions traitees",len(executions)],["Agents analyses",agents],["Lignes physiques sources",physical_rows],
                    ["Occurrences agregees",aggregated],["Lignes exportees",detail_writer.total],["Difference lignes",physical_rows-aggregated],
                    ["Brut lignes physiques",physical_gross],["Brut agrege",aggregated_gross],["Difference brut",gross_diff],
                    ["Net lignes physiques",physical_net],["Net agrege",aggregated_net],["Difference net",net_diff],["Controle","OK"]
                ]: control.append(row)
            finally:
                _cleanup_compact_stats(con)
        atomic_save_workbook(book,target)
        progress and progress(97,f"Fichier genere : {target.name}")
        return target


class PartitionedRawFusionRiskExporter(LowMemoryRawFusionRiskExporter):
    """Annexe 12 : syntheses d'abord, details ligne par ligne ensuite."""

    def _risk_summary_partitioned(self, category_filter: str | None = None):
        category=self._category_expr_low_memory(); risk=self._risk_predicate_low_memory()
        extra=""; params=[]
        if category_filter:
            extra=f" AND ({category})=?"; params=[category_filter]
        sql=f"""SELECT {category},r.statut,r.matricule_normalise,r.nom_normalise,r.nom,r.prenom,r.regimes,r.institutions,
            r.nb_regimes,r.nb_institutions,r.occurrences,GREATEST(r.occurrences-1,0),s.nb_executions,s.nb_tables,
            r.masse_brute,r.masse_net,COALESCE(d.doublon_matricule,FALSE),COALESCE(d.doublon_nom,FALSE),
            r.paiement_multi_regime,r.paiement_multiple_meme_regime,r.identite_incoherente,r.diagnostic
        FROM resultats_fusion_multi r JOIN tmp_sicorpa_identity_stats s ON s.person_key=r.person_key
        LEFT JOIN resultats_fusion_doublons d ON d.fusion_id=r.fusion_id AND d.person_key=r.person_key
        WHERE r.fusion_id=? AND {risk}{extra}"""
        return sql,params

    def _risk_execution_detail(self, category_filter: str | None = None):
        key=_person_key_sql("p"); category=self._category_expr_low_memory(); risk=self._risk_predicate_low_memory()
        extra=""; params=[]
        if category_filter:
            extra=f" AND ({category})=?"; params=[category_filter]
        sql=f"""SELECT {category},?,p.execution_id,p.ligne_paie_id,p.ligne_source,p.regime,p.institution_id,p.trimestre,p.annee,
            p.matricule_source,p.matricule_normalise,p.nom,p.prenom,p.nom_normalise,p.section,p.categorie,p.grade,
            p.unite_affectation,p.province,p.remuneration_brute_calculee,p.montant_net,r.statut,r.nb_regimes,r.occurrences,
            GREATEST(r.occurrences-1,0),s.nb_executions,s.nb_tables,r.regimes,r.institutions,s.noms_distincts,s.matricules_distincts,
            COALESCE(d.doublon_matricule,FALSE),COALESCE(d.doublon_nom,FALSE),r.paiement_multi_regime,
            r.paiement_multiple_meme_regime,r.identite_incoherente,r.diagnostic
        FROM paie_standardisee p
        JOIN resultats_fusion_multi r ON r.fusion_id=? AND r.person_key=({key})
        JOIN tmp_sicorpa_identity_stats s ON s.person_key=r.person_key
        LEFT JOIN resultats_fusion_doublons d ON d.fusion_id=r.fusion_id AND d.person_key=r.person_key
        WHERE p.execution_id=? AND p.trimestre=? AND p.annee=? AND {risk}{extra}"""
        return sql,params

    def export(self, fusion_id: str, folder: str | Path, progress=None) -> Path:
        folder=Path(folder); target=folder/"12_synthese_occurrences_agents_a_risque.xlsx"; book=Workbook(write_only=True)
        categories=[
            ("Matricule","PAR_MATRICULE"),
            ("Nom","PAR_NOM"),
            ("Matricule_Nom","PAR_MATRICULE_ET_NOM"),
            ("Matricule_NU","MATRICULE_NU"),
            ("Null_vide","MATRICULE_NULL_OU_VIDE"),
            ("Non_exploitable","MATRICULE_NON_EXPLOITABLE"),
            ("Identites_incoh","MATRICULE_PARTAGE_NOMS_DIFFERENTS"),
            ("Multi_regimes","MULTI_REGIME"),
            ("Paiements_multiples","PAIEMENT_MULTIPLE_MEME_REGIME"),
            ("Plusieurs_instit","PLUSIEURS_INSTITUTIONS"),
            ("Autres_anomalies","AUTRE_ANOMALIE"),
        ]
        with self.db.connect() as con:
            _configure_partitioned(con); quarter,year=_fusion_scope(con,fusion_id); executions=_fusion_executions(con,fusion_id)
            _prepare_compact_identity_stats(con,fusion_id,quarter,year,executions)
            try:
                # 1) Toutes les syntheses en premier.
                summary_sql,summary_extra=self._risk_summary_partitioned()
                global_summary=_WorkbookPartitionWriter(book,RISK_SUMMARY_HEADERS,"00_Synthese_generale")
                global_summary.append_cursor(con.execute(summary_sql,[fusion_id]+summary_extra),1000)

                category_summary_counts={}
                for index,(label,category_name) in enumerate(categories,1):
                    sql,extra=self._risk_summary_partitioned(category_name)
                    writer=_WorkbookPartitionWriter(book,RISK_SUMMARY_HEADERS,f"S{index:02d}_{label}")
                    writer.append_cursor(con.execute(sql,[fusion_id]+extra),1000)
                    category_summary_counts[category_name]=writer.total

                # 2) Ensuite seulement les annexes de detail ligne par ligne.
                detail_total=0
                detail_counts={}
                total_exec=max(1,len(executions))
                for cat_index,(label,category_name) in enumerate(categories,1):
                    writer=_WorkbookPartitionWriter(book,RISK_DETAIL_HEADERS,f"D{cat_index:02d}_{label}")
                    for exec_index,(execution_id,table_source) in enumerate(executions,1):
                        progress and progress(
                            97+int(exec_index/total_exec),
                            f"Annexe 12 : detail {label} - source {exec_index}/{total_exec}"
                        )
                        sql,extra=self._risk_execution_detail(category_name)
                        cursor=con.execute(sql,[table_source or "",fusion_id,execution_id,quarter,year]+extra)
                        detail_total += writer.append_cursor(cursor,1000)
                    detail_counts[category_name]=writer.total

                all_agents=int(con.execute("SELECT COUNT(*) FROM resultats_fusion_multi WHERE fusion_id=?",[fusion_id]).fetchone()[0] or 0)
                control=book.create_sheet("Controle")
                control.append(["Indicateur","Valeur"])
                control.append(["Mode export","PARTITIONNE_PAR_EXECUTION"])
                control.append(["Organisation","SYNTHESES_PUIS_DETAILS"])
                control.append(["Agents analyses au total",all_agents])
                control.append(["Agents a risque",global_summary.total])
                control.append(["Executions traitees",len(executions)])
                control.append(["Lignes detail exportees",detail_total])
                for label,category_name in categories:
                    control.append([f"Synthese {label}",category_summary_counts.get(category_name,0)])
                    control.append([f"Detail {label}",detail_counts.get(category_name,0)])
                control.append(["Controle","OK"])
            finally:
                _cleanup_compact_stats(con)
        atomic_save_workbook(book,target); progress and progress(99,f"Fichier genere : {target.name}"); return target
