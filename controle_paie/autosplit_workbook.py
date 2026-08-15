from __future__ import annotations

from contextlib import contextmanager

from openpyxl import Workbook as OpenpyxlWorkbook


EXCEL_MAX_DATA_ROWS = 1_048_575


class _AutoSplitSheet:
    def __init__(self, owner, title: str):
        self._owner = owner
        self._base_title = (title or "Resultats")[:31]
        self._sheet_index = 1
        self._sheet = owner._book.create_sheet(self._base_title)
        self._header = None
        self._data_rows = 0

    def _next_sheet(self):
        self._sheet_index += 1
        suffix = f"_{self._sheet_index}"
        title = (self._base_title[:31-len(suffix)] + suffix)[:31]
        self._sheet = self._owner._book.create_sheet(title)
        self._data_rows = 0
        if self._header is not None:
            self._sheet.append(self._header)

    def append(self, row):
        values = list(row)
        if self._header is None:
            self._header = values
            self._sheet.append(values)
            return
        if self._data_rows >= EXCEL_MAX_DATA_ROWS:
            self._next_sheet()
        self._sheet.append(values)
        self._data_rows += 1

    def __getattr__(self, name):
        return getattr(self._sheet, name)


class AutoSplitWriteOnlyWorkbook:
    """Proxy openpyxl qui fractionne automatiquement chaque feuille write-only.

    La premiere ligne ajoutee a chaque feuille est traitee comme en-tete et est
    repetee dans les feuilles suivantes. Les classeurs classiques ne passent pas
    par ce proxy et conservent donc toutes les fonctions de style openpyxl.
    """

    def __init__(self):
        self._book = OpenpyxlWorkbook(write_only=True)

    def create_sheet(self, title=None, index=None):
        # Le mode write-only ne gere pas utilement l'insertion par index dans les
        # exports SICORPA; on conserve la signature pour compatibilite.
        return _AutoSplitSheet(self, title or "Resultats")

    def save(self, filename):
        return self._book.save(filename)

    def __getattr__(self, name):
        return getattr(self._book, name)


def reliable_workbook(*args, **kwargs):
    """Factory compatible avec openpyxl.Workbook.

    Seuls les classeurs write-only sont proxifies. Les rapports de synthese
    classiques restent de vrais Workbook openpyxl afin de conserver styles,
    hyperlinks, dimensions et acces par cellule.
    """
    if kwargs.get("write_only", False):
        return AutoSplitWriteOnlyWorkbook()
    return OpenpyxlWorkbook(*args, **kwargs)


@contextmanager
def patch_workbook_factory(*modules):
    """Remplace temporairement Workbook dans des modules d'export legacy."""
    originals = []
    try:
        for module in modules:
            originals.append((module, getattr(module, "Workbook", None)))
            module.Workbook = reliable_workbook
        yield
    finally:
        for module, original in originals:
            if original is None:
                try:
                    delattr(module, "Workbook")
                except AttributeError:
                    pass
            else:
                module.Workbook = original
