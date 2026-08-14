from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Font

from .spreadsheet_utils import sanitize_excel_row


class RegimeComparisonFolderExporter:
    """Exporte une comparaison de régimes en plusieurs classeurs dans un dossier structuré."""

    EXPORTS = [
        ("01_tous_les_resultats.xlsx", "Tous les résultats", ""),
        ("02_payes_dans_les_deux.xlsx", "Payés dans les deux", "DOUBLE_PAIEMENT"),
        ("03_uniquement_regime_a.xlsx", "Uniquement régime A", "UNIQUEMENT_REGIME_A"),
        ("04_uniquement_regime_b.xlsx", "Uniquement régime B", "UNIQUEMENT_REGIME_B"),
        ("05_ecarts_financiers.xlsx", "Écarts financiers", "ECART_FINANCIER"),
        ("06_ecarts_financiers_et_admin.xlsx", "Écarts financiers et administratifs", "ECART_FINANCIER_ET_ADMIN"),
        ("07_ecarts_administratifs.xlsx", "Écarts administratifs", "ECART_ADMINISTRATIF"),
        ("08_paiements_multiples.xlsx", "Paiements multiples", "PAIEMENT_MULTIPLE"),
        ("09_identites_incoherentes.xlsx", "Identités incohérentes", "IDENTITE_INCOHERENTE"),
        ("10_communs_identiques.xlsx", "Communs identiques", "COMMUN_IDENTIQUE"),
    ]

    HEADERS = [
        "Statut", "Clé", "Matricule", "Nom", "Occurrences A", "Occurrences B",
        "Brut A", "Brut B", "Écart brut", "Net A", "Net B", "Écart net", "Écart %",
        "Grade A", "Grade B", "Catégorie A", "Catégorie B", "Affectation A", "Affectation B", "Diagnostic",
    ]

    def __init__(self, service):
        self.service = service

    @staticmethod
    def _safe_name(value: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value or "").strip())
        return cleaned.strip("_") or "comparaison"

    def export_all(self, comparison_id: str, parent_folder: str | Path, progress=None) -> str:
        summary = self.service.get_summary(comparison_id)
        folder_name = self._safe_name(
            f"comparaison_{summary['regime_a']}_vs_{summary['regime_b']}_{summary['quarter']}_{summary['year']}"
        )
        target = Path(parent_folder) / folder_name
        target.mkdir(parents=True, exist_ok=True)

        progress and progress(5, "Création du dossier d'export de la comparaison")
        self._write_summary(summary, target / "00_synthese_generale.xlsx")

        total = len(self.EXPORTS)
        for index, (filename, title, status) in enumerate(self.EXPORTS, 1):
            progress and progress(10 + int(85 * index / total), f"Export : {title}")
            rows = self.service.list_results(comparison_id, status, 10000)
            self._write_results(target / filename, title, summary, rows)

        with self.service.db.connect() as con:
            con.execute(
                "UPDATE comparaisons_regimes SET fichier_export=? WHERE comparaison_id=?",
                [str(target), comparison_id],
            )
        progress and progress(100, "Toutes les analyses ont été exportées")
        return str(target)

    def _write_summary(self, summary: dict, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Synthèse"
        ws.append(["Indicateur", "Valeur"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        rows = [
            ("Comparaison", f"{summary['regime_a']} vs {summary['regime_b']}"),
            ("Période", f"{summary['quarter']} {summary['year']}"),
            ("Lignes régime A", summary["rows_a"]),
            ("Lignes régime B", summary["rows_b"]),
            ("Agents communs / payés dans les deux", summary["common"]),
            ("Uniquement régime A", summary["only_a"]),
            ("Uniquement régime B", summary["only_b"]),
            ("Écarts financiers", summary["financial"]),
            ("Écarts administratifs", summary["administrative"]),
            ("Masse régime A", summary["mass_a"]),
            ("Masse régime B", summary["mass_b"]),
            ("Écart de masse", (summary["mass_a"] or 0) - (summary["mass_b"] or 0)),
            ("Seuil financier", summary["threshold_amount"]),
            ("Seuil pourcentage", summary["threshold_percent"]),
        ]
        for row in rows:
            ws.append(list(sanitize_excel_row(row)))
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 30
        wb.save(path)

    def _write_results(self, path: Path, title: str, summary: dict, rows: list[tuple]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats"
        ws.append(self.HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append(list(sanitize_excel_row(row)))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        info = wb.create_sheet("Informations")
        info.append(["Analyse", title])
        info.append(["Comparaison", f"{summary['regime_a']} vs {summary['regime_b']}"])
        info.append(["Période", f"{summary['quarter']} {summary['year']}"])
        info.append(["Nombre de lignes", len(rows)])
        info.append(["Identifiant comparaison", summary["id"]])
        wb.save(path)
