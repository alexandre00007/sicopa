from __future__ import annotations

import re
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .spreadsheet_utils import sanitize_excel_row


class RawFusionMultiRegimeService:
    """Fusionne plusieurs raw_* et analyse les agents présents dans plusieurs régimes."""

    def __init__(self, db):
        self.db = db
        self.ensure_schema()

    def ensure_schema(self):
        with self.db.connect() as con:
            con.execute("""CREATE TABLE IF NOT EXISTS fusions_raw (
                fusion_id VARCHAR PRIMARY KEY,
                table_fusion VARCHAR NOT NULL,
                trimestre VARCHAR NOT NULL,
                annee INTEGER NOT NULL,
                statut VARCHAR NOT NULL,
                lignes_fusion BIGINT DEFAULT 0,
                nombre_sources BIGINT DEFAULT 0,
                nombre_regimes BIGINT DEFAULT 0,
                cree_le TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                termine_le TIMESTAMP,
                dossier_export VARCHAR
            )""")
            con.execute("""CREATE TABLE IF NOT EXISTS sources_fusion_raw (
                fusion_id VARCHAR,
                table_source VARCHAR,
                execution_id VARCHAR,
                institution_id VARCHAR,
                regime VARCHAR,
                lignes BIGINT DEFAULT 0
            )""")
            con.execute("""CREATE TABLE IF NOT EXISTS resultats_fusion_multi (
                fusion_id VARCHAR,
                person_key VARCHAR,
                matricule_normalise VARCHAR,
                nom_normalise VARCHAR,
                nom VARCHAR,
                prenom VARCHAR,
                regimes VARCHAR,
                institutions VARCHAR,
                nb_regimes BIGINT DEFAULT 0,
                nb_institutions BIGINT DEFAULT 0,
                occurrences BIGINT DEFAULT 0,
                masse_brute DECIMAL(38,2) DEFAULT 0,
                masse_net DECIMAL(38,2) DEFAULT 0,
                remuneration_base DECIMAL(38,2) DEFAULT 0,
                transport DECIMAL(38,2) DEFAULT 0,
                prime DECIMAL(38,2) DEFAULT 0,
                logement DECIMAL(38,2) DEFAULT 0,
                pension_rente DECIMAL(38,2) DEFAULT 0,
                autres_remunerations DECIMAL(38,2) DEFAULT 0,
                retenues DECIMAL(38,2) DEFAULT 0,
                sections VARCHAR,
                categories VARCHAR,
                grades VARCHAR,
                unites_affectation VARCHAR,
                provinces VARCHAR,
                paiement_multi_regime BOOLEAN DEFAULT FALSE,
                paiement_multiple_meme_regime BOOLEAN DEFAULT FALSE,
                identite_incoherente BOOLEAN DEFAULT FALSE,
                statut VARCHAR,
                diagnostic VARCHAR
            )""")
            con.execute("CREATE INDEX IF NOT EXISTS idx_fusion_multi ON resultats_fusion_multi(fusion_id,statut)")

    @staticmethod
    def _quote(name: str) -> str:
        return '"' + str(name).replace('"', '""') + '"'

    @staticmethod
    def _safe_table_name(quarter: str, year: int, suffix: str = "") -> str:
        base = f"raw_multi_regimes_{quarter}_{int(year)}"
        suffix = re.sub(r"[^A-Za-z0-9_]+", "_", suffix or "").strip("_")
        return f"{base}_{suffix}" if suffix else base

    def list_raw_tables(self) -> list[tuple]:
        with self.db.connect() as con:
            names = [r[0] for r in con.execute("""SELECT table_name FROM information_schema.tables
                WHERE table_schema='main' AND table_name LIKE 'raw_%'
                  AND table_name NOT LIKE 'raw_multi_regimes_%'
                ORDER BY table_name""").fetchall()]
            out = []
            for name in names:
                count = int(con.execute(f"SELECT COUNT(*) FROM {self._quote(name)}").fetchone()[0])
                meta = con.execute("""SELECT regime,institution_id,trimestre,annee,
                        COUNT(DISTINCT execution_id)
                    FROM journal_executions
                    WHERE table_destination=? AND type_operation='IMPORT_ACCESS'
                    GROUP BY regime,institution_id,trimestre,annee
                    ORDER BY MAX(date_debut) DESC LIMIT 1""", [name]).fetchone()
                if meta:
                    regime, institution, quarter, year, executions = meta
                else:
                    regime = institution = quarter = year = None
                    executions = 0
                out.append((name, count, regime or "", institution or "", quarter or "", year or "", int(executions or 0)))
        return out

    def table_columns(self, table_name: str) -> list[tuple]:
        with self.db.connect() as con:
            exists = con.execute("SELECT COUNT(*) FROM information_schema.tables WHERE table_name=?", [table_name]).fetchone()[0]
            if not exists:
                raise ValueError("Table RAW introuvable.")
            return con.execute(f"DESCRIBE {self._quote(table_name)}").fetchall()

    def preview_schema(self, table_names: list[str]) -> dict:
        if not table_names:
            raise ValueError("Sélectionnez au moins une table RAW.")
        schemas = {name: [r[0] for r in self.table_columns(name)] for name in table_names}
        all_cols = sorted(set().union(*(set(v) for v in schemas.values())))
        common = sorted(set.intersection(*(set(v) for v in schemas.values()))) if schemas else []
        specifics = {name: sorted(set(cols) - set(common)) for name, cols in schemas.items()}
        return {"all": all_cols, "common": common, "specifics": specifics}

    def _source_executions(self, table_names: list[str], quarter: str, year: int) -> list[tuple]:
        placeholders = ",".join("?" for _ in table_names)
        with self.db.connect() as con:
            return con.execute(f"""SELECT DISTINCT table_destination,execution_id,institution_id,regime
                FROM journal_executions
                WHERE table_destination IN ({placeholders})
                  AND type_operation='IMPORT_ACCESS'
                  AND trimestre=? AND annee=? AND statut='TERMINE'
                  AND execution_id IS NOT NULL""", table_names + [quarter, int(year)]).fetchall()

    def create_fusion(self, table_names: list[str], quarter: str, year: int, suffix: str = "", progress=None) -> dict:
        if len(table_names) < 2:
            raise ValueError("Sélectionnez au moins deux tables RAW à fusionner.")
        quarter = str(quarter).strip().upper()
        if quarter not in {"T1", "T2", "T3", "T4"}:
            raise ValueError("Trimestre invalide.")
        year = int(year)
        available = {row[0] for row in self.list_raw_tables()}
        missing = [name for name in table_names if name not in available]
        if missing:
            raise ValueError("Tables introuvables : " + ", ".join(missing))
        executions = self._source_executions(table_names, quarter, year)
        if not executions:
            raise ValueError("Aucune exécution standardisée correspondante n'existe pour cette période.")

        fusion_id = str(uuid.uuid4())
        table_fusion = self._safe_table_name(quarter, year, suffix)
        progress and progress(5, "Préparation de la fusion RAW")
        with self.db.connect() as con:
            con.execute("BEGIN")
            try:
                con.execute(f"DROP TABLE IF EXISTS {self._quote(table_fusion)}")
                selects = []
                for name in table_names:
                    literal = name.replace("'", "''")
                    selects.append(f"SELECT *, '{literal}' AS _fusion_source_table FROM {self._quote(name)}")
                union_sql = " UNION ALL BY NAME ".join(selects)
                con.execute(f"CREATE TABLE {self._quote(table_fusion)} AS {union_sql}")
                rows = int(con.execute(f"SELECT COUNT(*) FROM {self._quote(table_fusion)}").fetchone()[0])
                regimes = len({r[3] for r in executions if r[3]})
                con.execute("""INSERT INTO fusions_raw
                    (fusion_id,table_fusion,trimestre,annee,statut,lignes_fusion,nombre_sources,nombre_regimes)
                    VALUES (?,?,?,?, 'EN_COURS',?,?,?)""",
                    [fusion_id, table_fusion, quarter, year, rows, len(table_names), regimes])
                for table_source, execution_id, institution_id, regime in executions:
                    source_rows = con.execute("SELECT COUNT(*) FROM paie_standardisee WHERE execution_id=?", [execution_id]).fetchone()[0]
                    con.execute("INSERT INTO sources_fusion_raw VALUES (?,?,?,?,?,?)",
                                [fusion_id, table_source, execution_id, institution_id, regime, source_rows])
                progress and progress(35, "Analyse des agents et des régimes")
                execution_ids = [r[1] for r in executions]
                ph = ",".join("?" for _ in execution_ids)
                con.execute(f"""INSERT INTO resultats_fusion_multi
                    WITH base AS (
                        SELECT p.*,
                          CASE WHEN COALESCE(matricule_normalise,'') NOT IN ('','NU')
                               THEN 'M:'||matricule_normalise
                               WHEN COALESCE(nom_normalise,'')<>'' THEN 'N:'||nom_normalise
                               ELSE 'L:'||ligne_paie_id END person_key
                        FROM paie_standardisee p
                        WHERE execution_id IN ({ph}) AND trimestre=? AND annee=?
                    ), per_regime AS (
                        SELECT person_key,regime,COUNT(*) n
                        FROM base GROUP BY person_key,regime
                    ), stats AS (
                        SELECT b.person_key,
                          MIN(NULLIF(b.matricule_normalise,'')) matricule_normalise,
                          MIN(NULLIF(b.nom_normalise,'')) nom_normalise,
                          MIN(NULLIF(b.nom,'')) nom,
                          MIN(NULLIF(b.prenom,'')) prenom,
                          STRING_AGG(DISTINCT COALESCE(b.regime,''), ', ' ORDER BY COALESCE(b.regime,'')) regimes,
                          STRING_AGG(DISTINCT COALESCE(b.institution_id,''), ', ' ORDER BY COALESCE(b.institution_id,'')) institutions,
                          COUNT(DISTINCT b.regime) nb_regimes,
                          COUNT(DISTINCT b.institution_id) nb_institutions,
                          COUNT(*) occurrences,
                          SUM(COALESCE(b.remuneration_brute_calculee,0)) masse_brute,
                          SUM(COALESCE(b.montant_net,0)) masse_net,
                          SUM(COALESCE(b.remuneration_base,0)) remuneration_base,
                          SUM(COALESCE(b.transport,0)) transport,
                          SUM(COALESCE(b.prime,0)) prime,
                          SUM(COALESCE(b.logement,0)) logement,
                          SUM(COALESCE(b.pension_rente,0)) pension_rente,
                          SUM(COALESCE(b.autres_remunerations,0)) autres_remunerations,
                          SUM(COALESCE(b.retenues,0)) retenues,
                          STRING_AGG(DISTINCT NULLIF(b.section,''), ', ' ORDER BY NULLIF(b.section,'')) sections,
                          STRING_AGG(DISTINCT NULLIF(b.categorie,''), ', ' ORDER BY NULLIF(b.categorie,'')) categories,
                          STRING_AGG(DISTINCT NULLIF(b.grade,''), ', ' ORDER BY NULLIF(b.grade,'')) grades,
                          STRING_AGG(DISTINCT NULLIF(b.unite_affectation,''), ', ' ORDER BY NULLIF(b.unite_affectation,'')) unites,
                          STRING_AGG(DISTINCT NULLIF(b.province,''), ', ' ORDER BY NULLIF(b.province,'')) provinces,
                          COUNT(DISTINCT COALESCE(b.nom_normalise,'')) noms_distincts,
                          MAX(pr.n) max_occ_regime
                        FROM base b
                        JOIN per_regime pr ON pr.person_key=b.person_key AND pr.regime=b.regime
                        GROUP BY b.person_key
                    )
                    SELECT ?,person_key,matricule_normalise,nom_normalise,nom,prenom,regimes,institutions,
                      nb_regimes,nb_institutions,occurrences,masse_brute,masse_net,remuneration_base,transport,prime,
                      logement,pension_rente,autres_remunerations,retenues,sections,categories,grades,unites,provinces,
                      nb_regimes>1,max_occ_regime>1,noms_distincts>1,
                      CASE
                        WHEN noms_distincts>1 THEN 'IDENTITE_INCOHERENTE'
                        WHEN nb_regimes>=3 THEN 'TROIS_REGIMES_OU_PLUS'
                        WHEN nb_regimes=2 THEN 'DEUX_REGIMES'
                        WHEN max_occ_regime>1 THEN 'PAIEMENT_MULTIPLE_MEME_REGIME'
                        WHEN nb_institutions>1 THEN 'PLUSIEURS_INSTITUTIONS'
                        ELSE 'UN_SEUL_REGIME' END,
                      TRIM(CONCAT_WS(' ; ',
                        CASE WHEN nb_regimes>1 THEN CAST(nb_regimes AS VARCHAR)||' régimes' END,
                        CASE WHEN max_occ_regime>1 THEN 'Plusieurs paiements dans un même régime' END,
                        CASE WHEN nb_institutions>1 THEN CAST(nb_institutions AS VARCHAR)||' institutions' END,
                        CASE WHEN noms_distincts>1 THEN 'Même clé associée à plusieurs identités' END
                      ))
                    FROM stats""", execution_ids + [quarter, year, fusion_id])
                con.execute("UPDATE fusions_raw SET statut='TERMINE',termine_le=CURRENT_TIMESTAMP WHERE fusion_id=?", [fusion_id])
                con.execute("COMMIT")
            except Exception:
                con.execute("ROLLBACK")
                raise
        progress and progress(100, "Fusion et analyse multi-régimes terminées")
        return self.get_fusion(fusion_id)

    def get_fusion(self, fusion_id: str) -> dict:
        with self.db.connect() as con:
            row = con.execute("""SELECT fusion_id,table_fusion,trimestre,annee,statut,lignes_fusion,
                nombre_sources,nombre_regimes,cree_le,termine_le,dossier_export
                FROM fusions_raw WHERE fusion_id=?""", [fusion_id]).fetchone()
        if not row:
            raise ValueError("Fusion introuvable.")
        keys = ["id","table","quarter","year","status","rows","sources","regimes","created","finished","export"]
        return dict(zip(keys, row))

    def list_history(self, limit: int = 100) -> list[tuple]:
        with self.db.connect() as con:
            return con.execute("""SELECT fusion_id,table_fusion,trimestre,annee,statut,lignes_fusion,
                nombre_sources,nombre_regimes,cree_le,dossier_export
                FROM fusions_raw ORDER BY cree_le DESC LIMIT ?""", [max(1,min(int(limit),500))]).fetchall()

    def sample_fusion(self, fusion_id: str, limit: int = 100) -> tuple[list[str], list[tuple]]:
        info = self.get_fusion(fusion_id)
        with self.db.connect() as con:
            cur = con.execute(f"SELECT * FROM {self._quote(info['table'])} LIMIT ?", [max(1,min(int(limit),500))])
            return [d[0] for d in cur.description], cur.fetchall()

    def list_results(self, fusion_id: str, status: str = "", limit: int = 3000) -> list[tuple]:
        condition = "fusion_id=?"
        params = [fusion_id]
        if status:
            condition += " AND statut=?"
            params.append(status)
        params.append(max(1,min(int(limit),10000)))
        with self.db.connect() as con:
            return con.execute(f"""SELECT statut,matricule_normalise,nom,prenom,regimes,nb_regimes,nb_institutions,
                occurrences,masse_brute,masse_net,sections,categories,grades,unites_affectation,provinces,
                paiement_multi_regime,paiement_multiple_meme_regime,identite_incoherente,diagnostic
                FROM resultats_fusion_multi WHERE {condition}
                ORDER BY nb_regimes DESC,occurrences DESC,masse_brute DESC LIMIT ?""", params).fetchall()

    def summary(self, fusion_id: str) -> list[tuple]:
        with self.db.connect() as con:
            return con.execute("""SELECT statut,COUNT(*),SUM(occurrences),SUM(masse_brute),SUM(masse_net)
                FROM resultats_fusion_multi WHERE fusion_id=? GROUP BY statut ORDER BY statut""", [fusion_id]).fetchall()

    def regime_matrix(self, fusion_id: str) -> tuple[list[str], list[list]]:
        with self.db.connect() as con:
            regimes = [r[0] for r in con.execute("""SELECT DISTINCT regime FROM sources_fusion_raw
                WHERE fusion_id=? AND COALESCE(regime,'')<>'' ORDER BY regime""", [fusion_id]).fetchall()]
            rows = []
            for a in regimes:
                values = [a]
                for b in regimes:
                    count = con.execute("""WITH persons AS (
                        SELECT CASE WHEN COALESCE(matricule_normalise,'') NOT IN ('','NU') THEN 'M:'||matricule_normalise
                                    ELSE 'N:'||COALESCE(nom_normalise,'') END person_key,regime
                        FROM paie_standardisee
                        WHERE execution_id IN (SELECT execution_id FROM sources_fusion_raw WHERE fusion_id=?)
                    )
                    SELECT COUNT(*) FROM (
                        SELECT person_key FROM persons WHERE regime IN (?,?)
                        GROUP BY person_key HAVING COUNT(DISTINCT regime)=?
                    ) x""", [fusion_id,a,b,1 if a==b else 2]).fetchone()[0]
                    values.append(int(count or 0))
                rows.append(values)
        return regimes, rows

    def export_all(self, fusion_id: str, parent_folder: str | Path, progress=None) -> str:
        info = self.get_fusion(fusion_id)
        folder = Path(parent_folder) / f"fusion_multi_regimes_{info['quarter']}_{info['year']}_{datetime.now():%Y%m%d_%H%M%S}"
        folder.mkdir(parents=True, exist_ok=True)
        progress and progress(5, "Création des exports")

        wb = Workbook(); ws = wb.active; ws.title = "Synthèse"
        ws.append(["Indicateur","Valeur"])
        for c in ws[1]: c.font = Font(bold=True)
        for row in [("Table fusionnée",info["table"]),("Période",f"{info['quarter']} {info['year']}"),
                    ("Lignes RAW",info["rows"]),("Sources",info["sources"]),("Régimes",info["regimes"])]:
            ws.append(list(sanitize_excel_row(row)))
        ws.append([]); ws.append(["Statut","Agents","Occurrences","Masse brute","Masse nette"])
        for c in ws[ws.max_row]: c.font = Font(bold=True)
        for row in self.summary(fusion_id): ws.append(list(sanitize_excel_row(row)))
        wb.save(folder / "00_synthese.xlsx")

        headers = ["Statut","Matricule","Nom","Prénom","Régimes","Nb régimes","Nb institutions","Occurrences",
                   "Masse brute","Masse nette","Sections","Catégories","Grades","Unités d'affectation","Provinces",
                   "Multi-régimes","Paiement multiple même régime","Identité incohérente","Diagnostic"]
        exports = [("01_tous_les_agents.xlsx", ""),("02_agents_deux_regimes.xlsx","DEUX_REGIMES"),
                   ("03_agents_trois_regimes_plus.xlsx","TROIS_REGIMES_OU_PLUS"),
                   ("04_paiements_multiples.xlsx","PAIEMENT_MULTIPLE_MEME_REGIME"),
                   ("05_identites_incoherentes.xlsx","IDENTITE_INCOHERENTE"),
                   ("06_plusieurs_institutions.xlsx","PLUSIEURS_INSTITUTIONS")]
        for i,(filename,status) in enumerate(exports,1):
            progress and progress(10+int(60*i/len(exports)), f"Export {filename}")
            book=Workbook(); sh=book.active; sh.title="Résultats"; sh.append(headers)
            for c in sh[1]: c.font=Font(bold=True)
            for row in self.list_results(fusion_id,status,10000): sh.append(list(sanitize_excel_row(row)))
            sh.freeze_panes="A2"; sh.auto_filter.ref=sh.dimensions; book.save(folder/filename)

        regimes, matrix = self.regime_matrix(fusion_id)
        book=Workbook(); sh=book.active; sh.title="Matrice"; sh.append(["Régime"]+regimes)
        for c in sh[1]: c.font=Font(bold=True)
        for row in matrix: sh.append(row)
        book.save(folder/"07_matrice_regimes.xlsx")

        cols, sample = self.sample_fusion(fusion_id, 5000)
        book=Workbook(); sh=book.active; sh.title="Listing fusionné"; sh.append(cols)
        for c in sh[1]: c.font=Font(bold=True)
        for row in sample: sh.append(list(sanitize_excel_row(row)))
        book.save(folder/"08_listing_fusionne_apercu.xlsx")
        with self.db.connect() as con:
            con.execute("UPDATE fusions_raw SET dossier_export=? WHERE fusion_id=?", [str(folder),fusion_id])
        progress and progress(100, "Export terminé")
        return str(folder)

    def delete_fusion(self, fusion_id: str):
        info = self.get_fusion(fusion_id)
        with self.db.connect() as con:
            con.execute("BEGIN")
            try:
                con.execute(f"DROP TABLE IF EXISTS {self._quote(info['table'])}")
                con.execute("DELETE FROM resultats_fusion_multi WHERE fusion_id=?", [fusion_id])
                con.execute("DELETE FROM sources_fusion_raw WHERE fusion_id=?", [fusion_id])
                con.execute("DELETE FROM fusions_raw WHERE fusion_id=?", [fusion_id])
                con.execute("COMMIT")
            except Exception:
                con.execute("ROLLBACK"); raise
