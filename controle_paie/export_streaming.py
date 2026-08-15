from __future__ import annotations

import csv
import os
from pathlib import Path

from openpyxl import Workbook

from .spreadsheet_utils import sanitize_excel_row


EXCEL_MAX_DATA_ROWS = 1_048_575  # une ligne réservée à l'en-tête


def _temporary_target(path: str | Path) -> tuple[Path, Path]:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f".{target.name}.part")
    return target, temporary


def atomic_save_workbook(book: Workbook, path: str | Path) -> Path:
    """Sauvegarde un classeur sans laisser de fichier final partiel en cas d'échec."""
    target, temporary = _temporary_target(path)
    try:
        book.save(temporary)
        os.replace(temporary, target)
    finally:
        if temporary.exists():
            temporary.unlink()
    return target


def append_query_sheets(book: Workbook, con, query: str, params=None, headers=None,
                        sheet_name: str = "Résultats", chunk_size: int = 5000) -> int:
    """Ajoute tout un résultat DuckDB à un classeur write-only, en plusieurs feuilles si nécessaire."""
    cursor = con.execute(query, params or [])
    if headers is None:
        headers = [column[0] for column in (cursor.description or [])]
    headers = list(headers)

    sheet_index = 1
    sheet = book.create_sheet(sheet_name[:31])
    sheet.append(list(sanitize_excel_row(headers)))
    rows_in_sheet = 0
    total = 0

    while True:
        chunk = cursor.fetchmany(max(100, int(chunk_size)))
        if not chunk:
            break
        for row in chunk:
            if rows_in_sheet >= EXCEL_MAX_DATA_ROWS:
                sheet_index += 1
                suffix = f"_{sheet_index}"
                sheet = book.create_sheet((sheet_name[:31-len(suffix)] + suffix)[:31])
                sheet.append(list(sanitize_excel_row(headers)))
                rows_in_sheet = 0
            sheet.append(list(sanitize_excel_row(row)))
            rows_in_sheet += 1
            total += 1
    return total


def write_query_xlsx(con, path: str | Path, query: str, params=None, headers=None,
                     sheet_name: str = "Résultats", chunk_size: int = 5000) -> int:
    """Écrit tout le résultat DuckDB en streaming et de manière atomique."""
    book = Workbook(write_only=True)
    total = append_query_sheets(book, con, query, params, headers, sheet_name, chunk_size)
    atomic_save_workbook(book, path)
    return total


def write_query_csv(con, path: str | Path, query: str, params=None, headers=None,
                    chunk_size: int = 10000) -> int:
    """Écrit tout le résultat DuckDB en CSV UTF-8 BOM sans chargement complet en mémoire."""
    target, temporary = _temporary_target(path)
    cursor = con.execute(query, params or [])
    if headers is None:
        headers = [column[0] for column in (cursor.description or [])]
    total = 0
    try:
        with temporary.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow(list(headers))
            while True:
                chunk = cursor.fetchmany(max(100, int(chunk_size)))
                if not chunk:
                    break
                writer.writerows(chunk)
                total += len(chunk)
        os.replace(temporary, target)
    finally:
        if temporary.exists():
            temporary.unlink()
    return total
