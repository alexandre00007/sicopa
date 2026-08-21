from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

from .export_streaming import atomic_save_workbook
from .raw_period_occurrence_exports import OccurrenceExportRawPeriodComparisonService
from .spreadsheet_utils import sanitize_excel_row


GLOBAL_HEADERS = [
    "Cle de Matching", "Cles A analysees", "Trouvees dans B", "Non trouvees dans B",
    "Occurrences A trouvees", "Occurrences B correspondantes", "Brut A", "Brut B",
    "Ecart Brut A-B", "Net A", "Net B", "Ecart Net A-B", "Montant Brut potentiellement chevauche",
    "Lien Direct",
]

SUMMARY_HEADERS = [
    "Cle / Valeur A", "Occurrences A", "Occurrences B", "Total Occurrences", "Ecart A-B",
    "Regime(s) A", "Regime(s) B", "Statut A -> B", "Niveau de Confiance",
    "Brut A", "Brut B", "Ecart Brut", "Net A", "Net B", "Ecart Net",
    "Brut potentiellement chevauche", "Net potentiellement chevauche", "Diagnostic Financier", "Action",
]

DETAIL_HEADERS = [
    "Cle de Matching",
    "ID_A", "Regime_A", "Institution_A", "Matricule_A", "Nom_A", "Prenom_A", "Ligne_Source_A", "Brut_A", "Net_A",
    "ID_B", "Regime_B", "Institution_B", "Matricule_B", "Nom_B", "Prenom_B", "Ligne_Source_B", "Brut_B", "Net_B",
    "Occurrences_A", "Occurrences_B", "Statut_A_vers_B", "Niveau_Confiance",
    "Brut_Total_A", "Brut_Total_B", "Ecart_Brut", "Net_Total_A", "Net_Total_B", "Ecart_Net",
    "Brut_Potentiellement_Chevauche", "Diagnostic_Financier",
]


class AToBRawPeriodExporter:
    """Annexe 20 : analyse directionnelle A -> B inspiree de l'annexe 12 multi-regimes.

    A est la population de reference. Les syntheses comptent aussi les cles non trouvees dans B,
    mais le drill-down detaille uniquement les cles ayant une correspondance dans B. Cela evite
    de recopier des centaines de milliers de lignes exclusives deja disponibles dans les exports
    RAW standards.
    """

    SHEETS = {
        "global": "Synthese Globale A vers B",
        "nom_summary": "Synthese_Nom_A_vers_B",
        "mat_summary": "Synthese_Matricule_A_vers_B",
        "combo_summary": "Synthese_MatNom_A_vers_B",
        "nom_detail": "Detail_Nom_A_vers_B",
        "mat_detail": "Detail_Matricule_A_vers_B",
        "combo_detail": "Detail_MatNom_A_vers_B",
    }

    def __init__(self, db):
        self.db = db

    @staticmethod
    def _name_sql(alias: str = "o") -> str:
        return f"""array_to_string(list_sort(regexp_split_to_array(
            trim(regexp_replace(upper(strip_accents(coalesce({alias}.nom,'')) || ' ' ||
                 strip_accents(coalesce({alias}.prenom,''))), '[^A-Z0-9]+', ' ', 'g')),
            '\\s+')), '')"""

    @staticmethod
    def _mat_sql(alias: str = "o") -> str:
        return f"""CASE
            WHEN regexp_replace(upper(coalesce({alias}.matricule_normalise,'')), '[^A-Z0-9]', '', 'g')='' THEN ''
            ELSE coalesce(nullif(ltrim(regexp_replace(upper(coalesce({alias}.matricule_normalise,'')),
                 '[^A-Z0-9]', '', 'g'),'0'),''),'0') END"""

    @staticmethod
    def _link(ws, target_sheet: str, row: int, label: str, bold: bool = False):
        cell = WriteOnlyCell(ws, value=label)
        cell.hyperlink = f"#'{target_sheet}'!A{row}"
        cell.font = Font(color="0563C1", underline="single", bold=bold)
        return cell

    def _prepare(self, con, comparison_id: str):
        con.execute("SET threads=1")
        con.execute("SET preserve_insertion_order=false")
        for table in ("tmp_a2b_combo", "tmp_a2b_mat", "tmp_a2b_nom", "tmp_a2b_scope"):
            con.execute(f"DROP TABLE IF EXISTS {table}")

        name_key = self._name_sql("o")
        mat_key = self._mat_sql("o")
        con.execute(f"""CREATE TEMP TABLE tmp_a2b_scope AS
            SELECT o.cote,o.table_source,o.execution_id,o.ligne_paie_id,o.ligne_source,
                   o.matricule_normalise,o.nom,o.prenom,o.institution_id,o.regime,
                   o.section,o.categorie,o.grade,o.unite_affectation,o.province,
                   COALESCE(o.brut,0) brut,COALESCE(o.net,0) net,
                   {name_key} AS nom_cle,{mat_key} AS mat_cle,
                   CASE WHEN ({mat_key})<>'' AND ({name_key})<>''
                        THEN ({mat_key}) || '|' || ({name_key}) ELSE '' END AS combo_cle
            FROM occurrences_comparaison_raw o WHERE o.comparaison_id=?""", [comparison_id])

        self._build_key_table(con, "tmp_a2b_nom", "nom_cle", "NOM")
        self._build_key_table(con, "tmp_a2b_mat", "mat_cle", "MATRICULE")
        self._build_key_table(con, "tmp_a2b_combo", "combo_cle", "MATNOM")

    def _build_key_table(self, con, table: str, key_col: str, kind: str):
        con.execute(f"""CREATE TEMP TABLE {table} AS
            WITH a AS (
                SELECT {key_col} cle,COUNT(*) occurrences,
                       string_agg(DISTINCT regime,' | ' ORDER BY regime) regimes,
                       COUNT(DISTINCT nom_cle) nb_noms,COUNT(DISTINCT mat_cle) nb_mats,
                       string_agg(DISTINCT nom_cle,' | ' ORDER BY nom_cle) noms,
                       string_agg(DISTINCT mat_cle,' | ' ORDER BY mat_cle) mats,
                       COALESCE(SUM(brut),0) brut,COALESCE(SUM(net),0) net
                FROM tmp_a2b_scope WHERE cote='A' AND {key_col}<>'' GROUP BY {key_col}
            ), b AS (
                SELECT {key_col} cle,COUNT(*) occurrences,
                       string_agg(DISTINCT regime,' | ' ORDER BY regime) regimes,
                       COUNT(DISTINCT nom_cle) nb_noms,COUNT(DISTINCT mat_cle) nb_mats,
                       string_agg(DISTINCT nom_cle,' | ' ORDER BY nom_cle) noms,
                       string_agg(DISTINCT mat_cle,' | ' ORDER BY mat_cle) mats,
                       COALESCE(SUM(brut),0) brut,COALESCE(SUM(net),0) net
                FROM tmp_a2b_scope WHERE cote='B' AND {key_col}<>'' GROUP BY {key_col}
            )
            SELECT a.cle,a.occurrences occurrences_a,COALESCE(b.occurrences,0) occurrences_b,
                   a.regimes regimes_a,COALESCE(b.regimes,'') regimes_b,
                   a.noms noms_a,COALESCE(b.noms,'') noms_b,a.mats mats_a,COALESCE(b.mats,'') mats_b,
                   a.nb_noms nb_noms_a,COALESCE(b.nb_noms,0) nb_noms_b,
                   a.nb_mats nb_mats_a,COALESCE(b.nb_mats,0) nb_mats_b,
                   a.brut brut_a,COALESCE(b.brut,0) brut_b,a.net net_a,COALESCE(b.net,0) net_b,
                   CASE
                     WHEN b.cle IS NULL THEN 'NON_TROUVE_DANS_B'
                     WHEN '{kind}'='MATRICULE' AND b.nb_noms>1 THEN 'PLUSIEURS_CORRESPONDANCES_DANS_B'
                     WHEN '{kind}'='MATRICULE' AND a.noms<>b.noms THEN 'TROUVE_DANS_B_DISCORDANCE_IDENTITE'
                     WHEN '{kind}'='NOM' AND b.nb_mats>1 THEN 'PLUSIEURS_CORRESPONDANCES_DANS_B'
                     WHEN '{kind}'='MATNOM' THEN 'TROUVE_DANS_B_NOM_MATRICULE'
                     WHEN '{kind}'='MATRICULE' THEN 'TROUVE_DANS_B_PAR_MATRICULE'
                     ELSE 'TROUVE_DANS_B_PAR_NOM' END statut,
                   CASE
                     WHEN b.cle IS NULL THEN 'NON_APPLICABLE'
                     WHEN '{kind}'='MATNOM' THEN 'TRES_ELEVEE'
                     WHEN '{kind}'='MATRICULE' AND (b.nb_noms>1 OR a.noms<>b.noms) THEN 'ALERTE'
                     WHEN '{kind}'='MATRICULE' THEN 'ELEVEE'
                     WHEN '{kind}'='NOM' AND b.nb_mats>1 THEN 'A_CONTROLER'
                     ELSE 'MOYENNE' END confiance,
                   CASE WHEN b.cle IS NULL THEN 0 ELSE LEAST(GREATEST(a.brut,0),GREATEST(b.brut,0)) END brut_chevauche,
                   CASE WHEN b.cle IS NULL THEN 0 ELSE LEAST(GREATEST(a.net,0),GREATEST(b.net,0)) END net_chevauche,
                   CASE
                     WHEN b.cle IS NULL THEN 'NON_APPLICABLE'
                     WHEN a.brut>0 AND b.brut>0 THEN 'RISQUE_DOUBLE_PAIEMENT'
                     WHEN ABS(a.brut-COALESCE(b.brut,0))>0.01 OR ABS(a.net-COALESCE(b.net,0))>0.01 THEN 'A_CONTROLER'
                     ELSE 'COHERENT' END diagnostic_financier
            FROM a LEFT JOIN b ON a.cle=b.cle""")

    @staticmethod
    def _cleanup(con):
        for table in ("tmp_a2b_combo", "tmp_a2b_mat", "tmp_a2b_nom", "tmp_a2b_scope"):
            try:
                con.execute(f"DROP TABLE IF EXISTS {table}")
            except Exception:
                pass

    @staticmethod
    def _anchors(con, table: str) -> dict[str, int]:
        row = 3
        anchors = {}
        for key, oa, ob in con.execute(f"""SELECT cle,occurrences_a,occurrences_b FROM {table}
            WHERE occurrences_b>0 ORDER BY cle""").fetchall():
            anchors[str(key)] = row
            row += max(int(oa or 0), int(ob or 0), 1)
        return anchors

    def _write_summary(self, book, con, sheet: str, table: str, detail_sheet: str, anchors):
        ws = book.create_sheet(sheet)
        ws.append([self._link(ws, self.SHEETS['global'], 1, '< Retour a la Synthese Globale', True)] + ['']*(len(SUMMARY_HEADERS)-1))
        ws.append(list(sanitize_excel_row(SUMMARY_HEADERS)))
        cursor = con.execute(f"""SELECT cle,occurrences_a,occurrences_b,regimes_a,regimes_b,statut,confiance,
            brut_a,brut_b,(brut_a-brut_b),net_a,net_b,(net_a-net_b),brut_chevauche,net_chevauche,diagnostic_financier
            FROM {table}
            ORDER BY CASE WHEN occurrences_b>0 THEN 0 ELSE 1 END,occurrences_b DESC,occurrences_a DESC,cle""")
        while True:
            rows = cursor.fetchmany(2000)
            if not rows:
                break
            for key,oa,ob,ra,rb,status,conf,ba,bb,eba,na,nb,ena,bch,nch,diag in rows:
                oa=int(oa or 0); ob=int(ob or 0); total=oa+ob
                if ob>0:
                    action=self._link(ws, detail_sheet, anchors[str(key)], 'Voir le detail')
                    total_cell=self._link(ws, detail_sheet, anchors[str(key)], str(total))
                else:
                    action='Synthese uniquement - absent de B'
                    total_cell=total
                ws.append([key,oa,ob,total_cell,oa-ob,ra or '',rb or '',status,conf,
                           float(ba or 0),float(bb or 0),float(eba or 0),float(na or 0),float(nb or 0),float(ena or 0),
                           float(bch or 0),float(nch or 0),diag,action])

    def _detail_query(self, table: str, key_col: str):
        return f"""WITH a AS (
            SELECT {key_col} cle,ligne_paie_id,regime,institution_id,matricule_normalise,nom,prenom,ligne_source,brut,net,
                   row_number() OVER (PARTITION BY {key_col} ORDER BY execution_id,ligne_source,ligne_paie_id) rn
            FROM tmp_a2b_scope WHERE cote='A' AND {key_col} IN (SELECT cle FROM {table} WHERE occurrences_b>0)
        ), b AS (
            SELECT {key_col} cle,ligne_paie_id,regime,institution_id,matricule_normalise,nom,prenom,ligne_source,brut,net,
                   row_number() OVER (PARTITION BY {key_col} ORDER BY execution_id,ligne_source,ligne_paie_id) rn
            FROM tmp_a2b_scope WHERE cote='B' AND {key_col} IN (SELECT cle FROM {table} WHERE occurrences_b>0)
        )
        SELECT COALESCE(a.cle,b.cle),
               a.ligne_paie_id,a.regime,a.institution_id,a.matricule_normalise,a.nom,a.prenom,a.ligne_source,a.brut,a.net,
               b.ligne_paie_id,b.regime,b.institution_id,b.matricule_normalise,b.nom,b.prenom,b.ligne_source,b.brut,b.net,
               k.occurrences_a,k.occurrences_b,k.statut,k.confiance,k.brut_a,k.brut_b,(k.brut_a-k.brut_b),
               k.net_a,k.net_b,(k.net_a-k.net_b),k.brut_chevauche,k.diagnostic_financier
        FROM a FULL OUTER JOIN b ON a.cle=b.cle AND a.rn=b.rn
        JOIN {table} k ON k.cle=COALESCE(a.cle,b.cle)
        ORDER BY COALESCE(a.cle,b.cle),COALESCE(a.rn,b.rn)"""

    def _write_detail(self, book, con, sheet: str, summary_sheet: str, table: str, key_col: str, progress=None):
        ws = book.create_sheet(sheet)
        ws.append([self._link(ws, summary_sheet, 1, '< Retour a la Synthese', True)] + ['']*(len(DETAIL_HEADERS)-1))
        ws.append(list(sanitize_excel_row(DETAIL_HEADERS)))
        cursor = con.execute(self._detail_query(table,key_col))
        exported=0
        while True:
            rows=cursor.fetchmany(2000)
            if not rows:
                break
            for row in rows:
                ws.append(list(sanitize_excel_row(row)))
                exported += 1
            if progress and exported % 50000 < 2000:
                progress(98, f"Annexe A vers B : {sheet} - {exported:,} lignes".replace(',', ' '))
        return exported

    def _global_metrics(self, con, table: str):
        row=con.execute(f"""SELECT COUNT(*),
            SUM(CASE WHEN occurrences_b>0 THEN 1 ELSE 0 END),
            SUM(CASE WHEN occurrences_b=0 THEN 1 ELSE 0 END),
            COALESCE(SUM(CASE WHEN occurrences_b>0 THEN occurrences_a ELSE 0 END),0),
            COALESCE(SUM(occurrences_b),0),
            COALESCE(SUM(CASE WHEN occurrences_b>0 THEN brut_a ELSE 0 END),0),
            COALESCE(SUM(brut_b),0),
            COALESCE(SUM(CASE WHEN occurrences_b>0 THEN net_a ELSE 0 END),0),
            COALESCE(SUM(net_b),0),COALESCE(SUM(brut_chevauche),0)
            FROM {table}""").fetchone()
        return [int(row[0] or 0),int(row[1] or 0),int(row[2] or 0),int(row[3] or 0),int(row[4] or 0),
                float(row[5] or 0),float(row[6] or 0),float(row[7] or 0),float(row[8] or 0),float(row[9] or 0)]

    def export(self, comparison_id: str, folder: str | Path, progress=None) -> Path:
        target=Path(folder)/"20_analyse_RAW_A_vers_B.xlsx"
        with self.db.connect() as con:
            self._prepare(con,comparison_id)
            try:
                progress and progress(97,"Annexe 20 : analyse A vers B")
                nom_anchors=self._anchors(con,"tmp_a2b_nom")
                mat_anchors=self._anchors(con,"tmp_a2b_mat")
                combo_anchors=self._anchors(con,"tmp_a2b_combo")
                book=Workbook(write_only=True)
                global_ws=book.create_sheet(self.SHEETS['global'])
                global_ws.append(list(sanitize_excel_row(GLOBAL_HEADERS)))
                specs=[
                    ("Matching par Nom Normalise","tmp_a2b_nom",self.SHEETS['nom_summary']),
                    ("Matching par Matricule","tmp_a2b_mat",self.SHEETS['mat_summary']),
                    ("Matching par Nom + Matricule","tmp_a2b_combo",self.SHEETS['combo_summary']),
                ]
                for label,table,target_sheet in specs:
                    total,found,missing,oa,ob,ba,bb,na,nb,bch=self._global_metrics(con,table)
                    global_ws.append([label,total,found,missing,oa,ob,ba,bb,ba-bb,na,nb,na-nb,bch,
                                      self._link(global_ws,target_sheet,1,'Ouvrir')])

                self._write_summary(book,con,self.SHEETS['nom_summary'],'tmp_a2b_nom',self.SHEETS['nom_detail'],nom_anchors)
                self._write_summary(book,con,self.SHEETS['mat_summary'],'tmp_a2b_mat',self.SHEETS['mat_detail'],mat_anchors)
                self._write_summary(book,con,self.SHEETS['combo_summary'],'tmp_a2b_combo',self.SHEETS['combo_detail'],combo_anchors)
                self._write_detail(book,con,self.SHEETS['nom_detail'],self.SHEETS['nom_summary'],'tmp_a2b_nom','nom_cle',progress)
                self._write_detail(book,con,self.SHEETS['mat_detail'],self.SHEETS['mat_summary'],'tmp_a2b_mat','mat_cle',progress)
                self._write_detail(book,con,self.SHEETS['combo_detail'],self.SHEETS['combo_summary'],'tmp_a2b_combo','combo_cle',progress)
            finally:
                self._cleanup(con)
        atomic_save_workbook(book,target)
        progress and progress(99,f"Annexe A vers B generee : {target.name}")
        return target


class AToBExportRawPeriodComparisonService(OccurrenceExportRawPeriodComparisonService):
    """Exports RAW standards + annexe 20 directionnelle A -> B."""

    def __init__(self,db):
        super().__init__(db)
        self.a_to_b_exporter=AToBRawPeriodExporter(db)

    def export_all(self,comparison_id: str,parent_folder,progress=None):
        def previous_progress(value,text=""):
            if progress:
                progress(min(96,int(max(0,value)*0.96)),text)
        folder=Path(super().export_all(comparison_id,parent_folder,progress=previous_progress))
        self.a_to_b_exporter.export(comparison_id,folder,progress=progress)
        progress and progress(100,"Export RAW termine : analyse A vers B incluse")
        return str(folder)
