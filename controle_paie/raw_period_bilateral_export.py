from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

from .export_streaming import atomic_save_workbook
from .raw_period_occurrence_exports import OccurrenceExportRawPeriodComparisonService
from .spreadsheet_utils import sanitize_excel_row


GLOBAL_HEADERS = [
    "Cle de Matching", "Agents Communs (Match)", "Uniques Regime A", "Uniques Regime B",
    "Discordances", "Taux de Chevauchement %", "Lien Direct",
]
SUMMARY_HEADERS = [
    "Cle / Valeur", "Occurrences Regime A", "Occurrences Regime B", "Total Occurrences",
    "Ecart Occurrences A-B", "Statut Comparatif", "Niveau de Confiance",
    "Regime(s) A", "Regime(s) B", "Action",
]
DETAIL_HEADERS = [
    "ID Cle de Matching",
    "ID_A", "Matricule_A", "Nom_A", "Prenom_A", "Statut_A",
    "ID_B", "Matricule_B", "Nom_B", "Prenom_B", "Statut_B",
    "Ecart Detecte", "Niveau de Confiance",
]


class BilateralRawPeriodComparisonExporter:
    """Annexe comparative A/B avec syntheses et drill-down sur les lignes physiques."""

    SHEETS = {
        "global": "Synthese Globale Comparee",
        "nom_summary": "Synthese_Nom_Comparee",
        "mat_summary": "Synthese_Matricule_Comparee",
        "combo_summary": "Synthese_MatNom_Comparee",
        "nom_detail": "Detail_Nom_Compare",
        "mat_detail": "Detail_Matricule_Compare",
        "combo_detail": "Detail_MatNom_Compare",
    }

    def __init__(self, db):
        self.db = db

    @staticmethod
    def _name_sql(alias: str = "o") -> str:
        return f"""array_to_string(list_sort(regexp_split_to_array(
            trim(regexp_replace(upper(strip_accents(coalesce({alias}.nom,'')) || ' ' ||
                 strip_accents(coalesce({alias}.prenom,''))), '[^A-Z0-9]+', ' ', 'g')),
            '\\s+')), '')"""

    @staticmethod
    def _mat_sql(alias: str = "o") -> str:
        return f"""CASE
            WHEN regexp_replace(upper(coalesce({alias}.matricule_normalise,'')), '[^A-Z0-9]', '', 'g')='' THEN ''
            ELSE coalesce(nullif(ltrim(regexp_replace(upper(coalesce({alias}.matricule_normalise,'')),
                 '[^A-Z0-9]', '', 'g'),'0'),''),'0')
        END"""

    @staticmethod
    def _link(ws, target_sheet: str, row: int, label: str, bold: bool = False):
        cell = WriteOnlyCell(ws, value=label)
        cell.hyperlink = f"#'{target_sheet}'!A{row}"
        cell.font = Font(color="0563C1", underline="single", bold=bold)
        return cell

    def _prepare(self, con, comparison_id: str):
        con.execute("SET threads=1")
        con.execute("SET preserve_insertion_order=false")
        for table in ("tmp_rpc_combo", "tmp_rpc_mat", "tmp_rpc_nom", "tmp_rpc_scope"):
            con.execute(f"DROP TABLE IF EXISTS {table}")

        name_key = self._name_sql("o")
        mat_key = self._mat_sql("o")
        con.execute(f"""CREATE TEMP TABLE tmp_rpc_scope AS
            SELECT o.cote,o.table_source,o.execution_id,o.ligne_paie_id,o.ligne_source,
                   o.matricule_normalise,o.nom,o.prenom,o.institution_id,o.regime,
                   o.section,o.categorie,o.grade,o.unite_affectation,o.province,o.brut,o.net,
                   {name_key} AS nom_cle,
                   {mat_key} AS mat_cle,
                   CASE WHEN ({mat_key})<>'' AND ({name_key})<>''
                        THEN ({mat_key}) || '|' || ({name_key}) ELSE '' END AS combo_cle
            FROM occurrences_comparaison_raw o WHERE o.comparaison_id=?""", [comparison_id])

        self._build_compare_table(con, "tmp_rpc_nom", "nom_cle", "NOM")
        self._build_compare_table(con, "tmp_rpc_mat", "mat_cle", "MATRICULE")
        self._build_compare_table(con, "tmp_rpc_combo", "combo_cle", "MATNOM")

    def _build_compare_table(self, con, table_name: str, key_col: str, kind: str):
        con.execute(f"""CREATE TEMP TABLE {table_name} AS
            WITH a AS (
                SELECT {key_col} cle,COUNT(*) occurrences,
                       string_agg(DISTINCT regime,' | ' ORDER BY regime) regimes,
                       string_agg(DISTINCT nom_cle,' | ' ORDER BY nom_cle) noms,
                       string_agg(DISTINCT mat_cle,' | ' ORDER BY mat_cle) matricules
                FROM tmp_rpc_scope WHERE cote='A' AND {key_col}<>'' GROUP BY {key_col}
            ), b AS (
                SELECT {key_col} cle,COUNT(*) occurrences,
                       string_agg(DISTINCT regime,' | ' ORDER BY regime) regimes,
                       string_agg(DISTINCT nom_cle,' | ' ORDER BY nom_cle) noms,
                       string_agg(DISTINCT mat_cle,' | ' ORDER BY mat_cle) matricules
                FROM tmp_rpc_scope WHERE cote='B' AND {key_col}<>'' GROUP BY {key_col}
            )
            SELECT COALESCE(a.cle,b.cle) cle,
                   COALESCE(a.occurrences,0) occurrences_a,
                   COALESCE(b.occurrences,0) occurrences_b,
                   COALESCE(a.regimes,'') regimes_a,COALESCE(b.regimes,'') regimes_b,
                   COALESCE(a.noms,'') noms_a,COALESCE(b.noms,'') noms_b,
                   COALESCE(a.matricules,'') matricules_a,COALESCE(b.matricules,'') matricules_b,
                   CASE
                     WHEN a.cle IS NULL THEN 'EXCLUSIF_REGIME_B'
                     WHEN b.cle IS NULL THEN 'EXCLUSIF_REGIME_A'
                     WHEN '{kind}'='MATRICULE' AND COALESCE(a.noms,'')<>COALESCE(b.noms,'') THEN 'COMMUN_DISCORDANCE'
                     WHEN a.occurrences<>b.occurrences THEN 'COMMUN_AVEC_OCCURRENCES_DIFFERENTES'
                     ELSE 'COMMUN_MATCH_EXACT' END AS statut,
                   CASE
                     WHEN a.cle IS NULL OR b.cle IS NULL THEN 'NON_APPLICABLE'
                     WHEN '{kind}'='MATRICULE' AND COALESCE(a.noms,'')<>COALESCE(b.noms,'') THEN 'ALERTE'
                     WHEN '{kind}'='MATNOM' THEN 'TRES_ELEVE'
                     WHEN '{kind}'='MATRICULE' THEN 'ELEVE'
                     ELSE 'MOYEN' END AS confiance
            FROM a FULL OUTER JOIN b ON a.cle=b.cle""")

    @staticmethod
    def _cleanup(con):
        for table in ("tmp_rpc_combo", "tmp_rpc_mat", "tmp_rpc_nom", "tmp_rpc_scope"):
            try:
                con.execute(f"DROP TABLE IF EXISTS {table}")
            except Exception:
                pass

    @staticmethod
    def _anchors(con, table_name: str) -> dict[str, int]:
        row = 3
        starts = {}
        for key, occ_a, occ_b in con.execute(
            f"SELECT cle,occurrences_a,occurrences_b FROM {table_name} ORDER BY cle"
        ).fetchall():
            starts[str(key)] = row
            row += max(int(occ_a or 0), int(occ_b or 0), 1)
        return starts

    def _write_summary(self, book, con, sheet_name: str, table_name: str, detail_sheet: str, anchors):
        ws = book.create_sheet(sheet_name)
        ws.append([self._link(ws, self.SHEETS["global"], 1, "< Retour a la Synthese Globale", True)] + [""] * (len(SUMMARY_HEADERS)-1))
        ws.append(list(sanitize_excel_row(SUMMARY_HEADERS)))
        for key, oa, ob, ra, rb, status, confidence in con.execute(f"""SELECT
            cle,occurrences_a,occurrences_b,regimes_a,regimes_b,statut,confiance
            FROM {table_name}
            ORDER BY CASE WHEN statut LIKE 'COMMUN%' THEN 0 ELSE 1 END,
                     (occurrences_a+occurrences_b) DESC,cle""").fetchall():
            total = int(oa or 0) + int(ob or 0)
            target = anchors[str(key)]
            ws.append([
                key, int(oa or 0), int(ob or 0),
                self._link(ws, detail_sheet, target, str(total)),
                int(oa or 0) - int(ob or 0), status, confidence, ra or "", rb or "",
                self._link(ws, detail_sheet, target, "Voir le detail"),
            ])

    def _detail_cursor(self, con, table_name: str, key_col: str, key: str):
        status, confidence, noms_a, noms_b = con.execute(
            f"SELECT statut,confiance,noms_a,noms_b FROM {table_name} WHERE cle=?", [key]
        ).fetchone()
        if status == "EXCLUSIF_REGIME_A":
            ecart = "ABSENT_DANS_B"
        elif status == "EXCLUSIF_REGIME_B":
            ecart = "ABSENT_DANS_A"
        elif status == "COMMUN_DISCORDANCE":
            ecart = f"MATRICULE_COMMUN_NOMS_DIFFERENTS: A={noms_a} / B={noms_b}"
        elif status == "COMMUN_AVEC_OCCURRENCES_DIFFERENTES":
            ecart = "NOMBRE_OCCURRENCES_DIFFERENT"
        else:
            ecart = "AUCUN_ECART_SUR_LA_CLE"

        return con.execute(f"""WITH a AS (
            SELECT ligne_paie_id,matricule_normalise,nom,prenom,
                   row_number() OVER (ORDER BY execution_id,ligne_source,ligne_paie_id) rn
            FROM tmp_rpc_scope WHERE cote='A' AND {key_col}=?
        ), b AS (
            SELECT ligne_paie_id,matricule_normalise,nom,prenom,
                   row_number() OVER (ORDER BY execution_id,ligne_source,ligne_paie_id) rn
            FROM tmp_rpc_scope WHERE cote='B' AND {key_col}=?
        )
        SELECT ?,a.ligne_paie_id,a.matricule_normalise,a.nom,a.prenom,'' AS statut_a,
                 b.ligne_paie_id,b.matricule_normalise,b.nom,b.prenom,'' AS statut_b,
                 ?,?
        FROM a FULL OUTER JOIN b ON a.rn=b.rn ORDER BY COALESCE(a.rn,b.rn)""",
        [key, key, key, ecart, confidence])

    def _write_detail(self, book, con, sheet_name: str, summary_sheet: str, table_name: str, key_col: str, progress=None):
        ws = book.create_sheet(sheet_name)
        ws.append([self._link(ws, summary_sheet, 1, "< Retour a la Synthese", True)] + [""] * (len(DETAIL_HEADERS)-1))
        ws.append(list(sanitize_excel_row(DETAIL_HEADERS)))
        keys = con.execute(f"SELECT cle FROM {table_name} ORDER BY cle").fetchall()
        total_keys = max(1, len(keys))
        exported = 0
        for index, (key,) in enumerate(keys, 1):
            cursor = self._detail_cursor(con, table_name, key_col, str(key))
            while True:
                rows = cursor.fetchmany(1000)
                if not rows:
                    break
                for row in rows:
                    ws.append(list(sanitize_excel_row(row)))
                    exported += 1
            if progress and index % 250 == 0:
                progress(98, f"Annexe bilaterale : {sheet_name} {index}/{total_keys}")
        return exported

    def _global_row(self, con, table_name: str):
        common, only_a, only_b, discord, union_count = con.execute(f"""SELECT
            SUM(CASE WHEN statut LIKE 'COMMUN%' THEN 1 ELSE 0 END),
            SUM(CASE WHEN statut='EXCLUSIF_REGIME_A' THEN 1 ELSE 0 END),
            SUM(CASE WHEN statut='EXCLUSIF_REGIME_B' THEN 1 ELSE 0 END),
            SUM(CASE WHEN statut='COMMUN_DISCORDANCE' THEN 1 ELSE 0 END),
            COUNT(*) FROM {table_name}""").fetchone()
        common = int(common or 0); union_count = int(union_count or 0)
        overlap = (100.0 * common / union_count) if union_count else 0.0
        return common, int(only_a or 0), int(only_b or 0), int(discord or 0), overlap

    def export(self, comparison_id: str, folder: str | Path, progress=None) -> Path:
        target = Path(folder) / "20_comparaison_bilaterale_RAW_A_vs_RAW_B.xlsx"
        with self.db.connect() as con:
            self._prepare(con, comparison_id)
            try:
                nom_anchors = self._anchors(con, "tmp_rpc_nom")
                mat_anchors = self._anchors(con, "tmp_rpc_mat")
                combo_anchors = self._anchors(con, "tmp_rpc_combo")

                book = Workbook(write_only=True)
                global_ws = book.create_sheet(self.SHEETS["global"])
                global_ws.append(list(sanitize_excel_row(GLOBAL_HEADERS)))
                specs = [
                    ("Matching par Nom Normalise", "tmp_rpc_nom", self.SHEETS["nom_summary"]),
                    ("Matching par Matricule", "tmp_rpc_mat", self.SHEETS["mat_summary"]),
                    ("Matching par Matricule + Nom", "tmp_rpc_combo", self.SHEETS["combo_summary"]),
                ]
                for label, table, target_sheet in specs:
                    common, only_a, only_b, discord, overlap = self._global_row(con, table)
                    global_ws.append([
                        label, common, only_a, only_b, discord, round(overlap, 2),
                        self._link(global_ws, target_sheet, 1, "Ouvrir"),
                    ])

                self._write_summary(book, con, self.SHEETS["nom_summary"], "tmp_rpc_nom", self.SHEETS["nom_detail"], nom_anchors)
                self._write_summary(book, con, self.SHEETS["mat_summary"], "tmp_rpc_mat", self.SHEETS["mat_detail"], mat_anchors)
                self._write_summary(book, con, self.SHEETS["combo_summary"], "tmp_rpc_combo", self.SHEETS["combo_detail"], combo_anchors)

                self._write_detail(book, con, self.SHEETS["nom_detail"], self.SHEETS["nom_summary"], "tmp_rpc_nom", "nom_cle", progress)
                self._write_detail(book, con, self.SHEETS["mat_detail"], self.SHEETS["mat_summary"], "tmp_rpc_mat", "mat_cle", progress)
                self._write_detail(book, con, self.SHEETS["combo_detail"], self.SHEETS["combo_summary"], "tmp_rpc_combo", "combo_cle", progress)
            finally:
                self._cleanup(con)
        atomic_save_workbook(book, target)
        progress and progress(99, f"Annexe comparative generee : {target.name}")
        return target


class BilateralExportRawPeriodComparisonService(OccurrenceExportRawPeriodComparisonService):
    """Ajoute l'annexe 20 de comparaison bilaterale aux exports RAW existants."""

    def __init__(self, db):
        super().__init__(db)
        self.bilateral_exporter = BilateralRawPeriodComparisonExporter(db)

    def export_all(self, comparison_id: str, parent_folder, progress=None):
        def previous_progress(value, text=""):
            if progress:
                progress(min(96, int(max(0, value) * 0.96)), text)
        folder = Path(super().export_all(comparison_id, parent_folder, progress=previous_progress))
        self.bilateral_exporter.export(comparison_id, folder, progress=progress)
        progress and progress(100, "Export Comparaison RAW termine : annexe bilaterale incluse")
        return str(folder)
