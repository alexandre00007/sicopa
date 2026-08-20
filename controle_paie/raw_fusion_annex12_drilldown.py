from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

from .export_streaming import atomic_save_workbook
from .spreadsheet_utils import sanitize_excel_row


DETAIL_HEADERS = [
    "ID Ligne Base", "Regime Source", "Institution Source", "Matricule Source",
    "Nom Brut en Base", "Prenom Brut", "Nom Normalise", "Matricule Normalise",
    "Cle de Matching", "Statut Agent", "Date Affiliation", "Table Source",
    "Execution ID", "Ligne Source", "Section", "Categorie", "Grade",
    "Unite d'affectation", "Province", "Remuneration Brute", "Montant Net",
    "Brut Total Identite", "Net Total Identite", "Brut Moyen", "Net Moyen",
    "Brut Minimum", "Brut Maximum", "Ecart Brut Max-Min",
    "Net Minimum", "Net Maximum", "Ecart Net Max-Min",
    "Brut Potentiellement Duplique", "Net Potentiellement Duplique",
    "Part Brut Occurrence (%)", "Diagnostic Financier",
]
SUMMARY_HEADERS = [
    "Cle / Valeur", "Nombre de Regimes Impactes", "Nombre d'Occurrences",
    "Liste des Regimes", "Brut Total", "Net Total", "Brut Moyen", "Net Moyen",
    "Brut Minimum", "Brut Maximum", "Ecart Brut Max-Min",
    "Net Minimum", "Net Maximum", "Ecart Net Max-Min",
    "Brut Potentiellement Duplique", "Net Potentiellement Duplique",
    "Occurrences / Regime", "Brut / Regime", "Diagnostic Financier", "Action / Detail",
]
GLOBAL_HEADERS = [
    "Rubrique / Cle de Matching", "Nombre d'Agents Uniques",
    "Nombre Total d'Occurrences", "Regimes Impactes", "Brut Total Concerne",
    "Net Total Concerne", "Brut Potentiellement Duplique",
    "Net Potentiellement Duplique", "Agents a Risque Double Paiement", "Lien Direct",
]


def normalize_name_python(value: str | None) -> str:
    """Majuscules, sans accents/ponctuation, mots tries puis concatenes."""
    text = unicodedata.normalize("NFKD", str(value or "").upper())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return "".join(sorted(re.findall(r"[A-Z0-9]+", text)))


def normalize_matricule_python(value: str | None) -> str:
    """Alphanumerique uniquement, sans separateurs ni zeros initiaux."""
    text = re.sub(r"[^A-Z0-9]", "", str(value or "").upper())
    if not text:
        return ""
    stripped = text.lstrip("0")
    return stripped or "0"


class DrillDownAnnex12Exporter:
    """Annexe 12 interactive a 7 feuilles pour les cles presentes dans >=2 regimes."""

    SHEETS = {
        "global": "Synthese Globale",
        "nom_summary": "Synthese_Nom",
        "mat_summary": "Synthese_Matricule",
        "combo_summary": "Synthese_Matricule_Nom",
        "nom_detail": "Detail_Nom",
        "mat_detail": "Detail_Matricule",
        "combo_detail": "Detail_Matricule_Nom",
    }

    def __init__(self, db):
        self.db = db

    @staticmethod
    def _name_sql(alias: str = "p") -> str:
        return f"""array_to_string(list_sort(regexp_split_to_array(
            trim(regexp_replace(upper(strip_accents(coalesce({alias}.nom,'')) || ' ' ||
                 strip_accents(coalesce({alias}.prenom,''))), '[^A-Z0-9]+', ' ', 'g')),
            '\\s+')), '')"""

    @staticmethod
    def _mat_sql(alias: str = "p") -> str:
        return f"""CASE
            WHEN regexp_replace(upper(coalesce({alias}.matricule_source,'')), '[^A-Z0-9]', '', 'g')='' THEN ''
            ELSE coalesce(nullif(ltrim(regexp_replace(upper(coalesce({alias}.matricule_source,'')), '[^A-Z0-9]', '', 'g'),'0'),''),'0')
        END"""

    @staticmethod
    def _financial_diagnostic_sql(prefix: str = "") -> str:
        p = f"{prefix}." if prefix else ""
        return f"""CASE
            WHEN COALESCE({p}nb_regimes,0)>=2 AND COALESCE({p}positive_occurrences,0)>=2
                 AND COALESCE({p}brut_potentiellement_duplique,0)>0 THEN 'RISQUE_DOUBLE_PAIEMENT'
            WHEN COALESCE({p}ecart_brut,0)>0 OR COALESCE({p}ecart_net,0)>0 THEN 'A_CONTROLER'
            ELSE 'COHERENT'
        END"""

    def _prepare(self, con, fusion_id: str) -> tuple[str, int]:
        con.execute("SET threads=1")
        con.execute("SET preserve_insertion_order=false")
        row = con.execute("SELECT trimestre,annee FROM fusions_raw WHERE fusion_id=?", [fusion_id]).fetchone()
        if not row:
            raise ValueError(f"Fusion introuvable : {fusion_id}")
        quarter, year = row[0], int(row[1])
        for table in ("tmp_ann12_combo_keys", "tmp_ann12_mat_keys", "tmp_ann12_nom_keys", "tmp_ann12_scope"):
            con.execute(f"DROP TABLE IF EXISTS {table}")

        name_key = self._name_sql("p")
        mat_key = self._mat_sql("p")
        con.execute(f"""CREATE TEMP TABLE tmp_ann12_scope AS
            SELECT p.ligne_paie_id,p.execution_id,p.institution_id,p.regime,p.table_source,
                   p.matricule_source,p.nom,p.prenom,p.ligne_source,p.section,p.categorie,p.grade,
                   p.unite_affectation,p.province,
                   COALESCE(p.remuneration_brute_calculee,0) AS remuneration_brute_calculee,
                   COALESCE(p.montant_net,0) AS montant_net,
                   {name_key} AS nom_cle,{mat_key} AS matricule_cle,
                   CASE WHEN ({mat_key})<>'' AND ({name_key})<>''
                        THEN ({mat_key}) || '|' || ({name_key}) ELSE '' END AS combo_cle
            FROM paie_standardisee p
            WHERE p.trimestre=? AND p.annee=?
              AND p.execution_id IN (
                SELECT DISTINCT execution_id FROM sources_fusion_raw
                WHERE fusion_id=? AND execution_id IS NOT NULL
              )""", [quarter, year, fusion_id])

        for table, column in (
            ("tmp_ann12_nom_keys", "nom_cle"),
            ("tmp_ann12_mat_keys", "matricule_cle"),
            ("tmp_ann12_combo_keys", "combo_cle"),
        ):
            con.execute(f"""CREATE TEMP TABLE {table} AS
                SELECT {column} AS cle,
                       COUNT(DISTINCT regime) nb_regimes,
                       COUNT(*) occurrences,
                       COUNT(*) FILTER (WHERE remuneration_brute_calculee>0 OR montant_net>0) positive_occurrences,
                       string_agg(DISTINCT regime, ' | ' ORDER BY regime) regimes,
                       COALESCE(SUM(remuneration_brute_calculee),0) brut_total,
                       COALESCE(SUM(montant_net),0) net_total,
                       COALESCE(AVG(remuneration_brute_calculee),0) brut_moyen,
                       COALESCE(AVG(montant_net),0) net_moyen,
                       COALESCE(MIN(remuneration_brute_calculee),0) brut_min,
                       COALESCE(MAX(remuneration_brute_calculee),0) brut_max,
                       COALESCE(MAX(remuneration_brute_calculee)-MIN(remuneration_brute_calculee),0) ecart_brut,
                       COALESCE(MIN(montant_net),0) net_min,
                       COALESCE(MAX(montant_net),0) net_max,
                       COALESCE(MAX(montant_net)-MIN(montant_net),0) ecart_net,
                       GREATEST(COALESCE(SUM(remuneration_brute_calculee),0)-COALESCE(MAX(remuneration_brute_calculee),0),0) brut_potentiellement_duplique,
                       GREATEST(COALESCE(SUM(montant_net),0)-COALESCE(MAX(montant_net),0),0) net_potentiellement_duplique
                FROM tmp_ann12_scope WHERE {column}<>'' GROUP BY {column}
                HAVING COUNT(DISTINCT regime)>=2""")
        return quarter, year

    @staticmethod
    def _cleanup(con):
        for table in ("tmp_ann12_combo_keys", "tmp_ann12_mat_keys", "tmp_ann12_nom_keys", "tmp_ann12_scope"):
            try:
                con.execute(f"DROP TABLE IF EXISTS {table}")
            except Exception:
                pass

    @staticmethod
    def _link_cell(ws, target_sheet: str, row: int, label: str, bold: bool = False):
        cell = WriteOnlyCell(ws, value=label)
        cell.hyperlink = f"#'{target_sheet}'!A{row}"
        cell.font = Font(color="0563C1", underline="single", bold=bold)
        return cell

    @staticmethod
    def _anchor_map(con, key_table: str) -> dict[str, int]:
        starts: dict[str, int] = {}
        row = 3
        for key, occurrences in con.execute(
            f"SELECT cle,occurrences FROM {key_table} ORDER BY cle"
        ).fetchall():
            starts[str(key)] = row
            row += int(occurrences or 0)
        return starts

    @staticmethod
    def _regime_breakdown(con, key_column: str, key: str) -> tuple[str, str]:
        rows = con.execute(f"""SELECT regime,COUNT(*) occurrences,
                COALESCE(SUM(remuneration_brute_calculee),0) brut
            FROM tmp_ann12_scope WHERE {key_column}=?
            GROUP BY regime ORDER BY regime""", [key]).fetchall()
        occurrences = " | ".join(f"{regime}={int(count or 0)}" for regime, count, _ in rows)
        gross = " | ".join(f"{regime}={float(amount or 0):.2f}" for regime, _, amount in rows)
        return occurrences, gross

    def _detail_rows_for_key(self, con, key_table: str, key_column: str, key: str):
        diagnostic = self._financial_diagnostic_sql("k")
        return con.execute(f"""SELECT s.ligne_paie_id,s.regime,s.institution_id,s.matricule_source,s.nom,s.prenom,
            s.nom_cle,s.matricule_cle,s.{key_column},'' AS statut_agent,NULL AS date_affiliation,
            s.table_source,s.execution_id,s.ligne_source,s.section,s.categorie,s.grade,s.unite_affectation,
            s.province,s.remuneration_brute_calculee,s.montant_net,
            k.brut_total,k.net_total,k.brut_moyen,k.net_moyen,k.brut_min,k.brut_max,k.ecart_brut,
            k.net_min,k.net_max,k.ecart_net,k.brut_potentiellement_duplique,k.net_potentiellement_duplique,
            CASE WHEN k.brut_total<>0 THEN ROUND(100.0*s.remuneration_brute_calculee/k.brut_total,2) ELSE 0 END,
            {diagnostic}
            FROM tmp_ann12_scope s JOIN {key_table} k ON k.cle=s.{key_column}
            WHERE s.{key_column}=?
            ORDER BY s.regime,s.execution_id,s.ligne_source,s.ligne_paie_id""", [key])

    def _write_summary(self, book, con, sheet_name: str, key_table: str, key_column: str,
                       detail_sheet: str, starts: dict[str, int]):
        ws = book.create_sheet(sheet_name)
        ws.append([self._link_cell(ws, self.SHEETS["global"], 1, "< Retour a la Synthese Globale", True)] + [""] * (len(SUMMARY_HEADERS)-1))
        ws.append(list(sanitize_excel_row(SUMMARY_HEADERS)))
        diagnostic = self._financial_diagnostic_sql("k")
        rows = con.execute(f"""SELECT k.cle,k.nb_regimes,k.occurrences,k.regimes,
            k.brut_total,k.net_total,k.brut_moyen,k.net_moyen,k.brut_min,k.brut_max,k.ecart_brut,
            k.net_min,k.net_max,k.ecart_net,k.brut_potentiellement_duplique,k.net_potentiellement_duplique,
            {diagnostic}
            FROM {key_table} k ORDER BY k.occurrences DESC,k.cle""").fetchall()
        for row in rows:
            key = str(row[0])
            target = starts[key]
            occ_by_regime, gross_by_regime = self._regime_breakdown(con, key_column, key)
            ws.append([
                key,
                int(row[1] or 0),
                self._link_cell(ws, detail_sheet, target, str(int(row[2] or 0))),
                row[3] or "",
                *row[4:16],
                occ_by_regime,
                gross_by_regime,
                row[16],
                self._link_cell(ws, detail_sheet, target, "Voir les lignes"),
            ])

    def _write_detail(self, book, con, sheet_name: str, summary_sheet: str, key_table: str,
                      key_column: str, progress=None):
        ws = book.create_sheet(sheet_name)
        ws.append([self._link_cell(ws, summary_sheet, 1, "< Retour a la Synthese", True)] + [""] * (len(DETAIL_HEADERS)-1))
        ws.append(list(sanitize_excel_row(DETAIL_HEADERS)))
        keys = con.execute(f"SELECT cle FROM {key_table} ORDER BY cle").fetchall()
        total = max(1, len(keys))
        exported = 0
        for idx, (key,) in enumerate(keys, 1):
            cursor = self._detail_rows_for_key(con, key_table, key_column, str(key))
            while True:
                rows = cursor.fetchmany(1000)
                if not rows:
                    break
                for row in rows:
                    ws.append(list(sanitize_excel_row(row)))
                    exported += 1
            if progress and idx % 100 == 0:
                progress(98, f"Annexe 12 : {sheet_name} {idx}/{total}")
        return exported

    def _global_metrics(self, con):
        specs = [
            ("Matching par Nom Normalise", "tmp_ann12_nom_keys", "nom_cle", self.SHEETS["nom_summary"]),
            ("Matching par Matricule", "tmp_ann12_mat_keys", "matricule_cle", self.SHEETS["mat_summary"]),
            ("Matching par Matricule + Nom", "tmp_ann12_combo_keys", "combo_cle", self.SHEETS["combo_summary"]),
        ]
        result = []
        for label, table, column, target_sheet in specs:
            row = con.execute(f"""SELECT COUNT(*),COALESCE(SUM(occurrences),0),
                COALESCE(SUM(brut_total),0),COALESCE(SUM(net_total),0),
                COALESCE(SUM(brut_potentiellement_duplique),0),
                COALESCE(SUM(net_potentiellement_duplique),0),
                COUNT(*) FILTER (WHERE nb_regimes>=2 AND positive_occurrences>=2 AND brut_potentiellement_duplique>0)
                FROM {table}""").fetchone()
            regimes = con.execute(f"""SELECT COALESCE(string_agg(DISTINCT s.regime,' | ' ORDER BY s.regime),'')
                FROM tmp_ann12_scope s JOIN {table} k ON s.{column}=k.cle""").fetchone()[0]
            result.append((
                label,int(row[0] or 0),int(row[1] or 0),regimes or "",
                float(row[2] or 0),float(row[3] or 0),float(row[4] or 0),float(row[5] or 0),
                int(row[6] or 0),target_sheet,
            ))
        return result

    def export(self, fusion_id: str, folder: str | Path, progress=None) -> Path:
        target = Path(folder) / "12_matching_multi_regimes_drilldown.xlsx"
        with self.db.connect() as con:
            self._prepare(con, fusion_id)
            try:
                progress and progress(97, "Annexe 12 : calcul des correspondances et impacts financiers")
                nom_starts = self._anchor_map(con, "tmp_ann12_nom_keys")
                mat_starts = self._anchor_map(con, "tmp_ann12_mat_keys")
                combo_starts = self._anchor_map(con, "tmp_ann12_combo_keys")

                book = Workbook(write_only=True)
                global_ws = book.create_sheet(self.SHEETS["global"])
                global_ws.append(list(sanitize_excel_row(GLOBAL_HEADERS)))
                for row in self._global_metrics(con):
                    global_ws.append([
                        *row[:-1],
                        self._link_cell(global_ws, row[-1], 1, "Ouvrir"),
                    ])

                self._write_summary(book, con, self.SHEETS["nom_summary"], "tmp_ann12_nom_keys", "nom_cle", self.SHEETS["nom_detail"], nom_starts)
                self._write_summary(book, con, self.SHEETS["mat_summary"], "tmp_ann12_mat_keys", "matricule_cle", self.SHEETS["mat_detail"], mat_starts)
                self._write_summary(book, con, self.SHEETS["combo_summary"], "tmp_ann12_combo_keys", "combo_cle", self.SHEETS["combo_detail"], combo_starts)

                self._write_detail(book, con, self.SHEETS["nom_detail"], self.SHEETS["nom_summary"], "tmp_ann12_nom_keys", "nom_cle", progress)
                self._write_detail(book, con, self.SHEETS["mat_detail"], self.SHEETS["mat_summary"], "tmp_ann12_mat_keys", "matricule_cle", progress)
                self._write_detail(book, con, self.SHEETS["combo_detail"], self.SHEETS["combo_summary"], "tmp_ann12_combo_keys", "combo_cle", progress)
            finally:
                self._cleanup(con)

        atomic_save_workbook(book, target)
        progress and progress(99, f"Annexe 12 generee : {target.name}")
        return target
