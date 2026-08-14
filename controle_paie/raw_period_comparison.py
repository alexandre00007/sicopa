from __future__ import annotations

import re
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from .spreadsheet_utils import sanitize_excel_row


class RawPeriodComparisonService:
    """Compare deux sources raw_* sur une période, par matricule et par nom normalisés."""

    def __init__(self, db):
        self.db = db
        self.ensure_schema()

    def ensure_schema(self):
        with self.db.connect() as con:
            con.execute("""CREATE TABLE IF NOT EXISTS comparaisons_raw_periode (
                comparaison_id VARCHAR PRIMARY KEY,
                table_a VARCHAR NOT NULL,
                table_b VARCHAR NOT NULL,
                trimestre VARCHAR NOT NULL,
                annee INTEGER NOT NULL,
                statut VARCHAR NOT NULL,
                cree_le TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                termine_le TIMESTAMP,
                dossier_export VARCHAR
            )""")
            con.execute("""CREATE TABLE IF NOT EXISTS sources_comparaison_raw_periode (
                comparaison_id VARCHAR,
                cote VARCHAR,
                table_source VARCHAR,
                execution_id VARCHAR,
                institution_id VARCHAR,
                regime VARCHAR
            )""")
            con.execute("""CREATE TABLE IF NOT EXISTS resultats_comparaison_raw_periode (
                comparaison_id VARCHAR,
                cle_resultat VARCHAR,
                statut VARCHAR,
                commun_matricule BOOLEAN DEFAULT FALSE,
                commun_nom BOOLEAN DEFAULT FALSE,
                meme_matricule_nom_different BOOLEAN DEFAULT FALSE,
                meme_nom_matricule_different BOOLEAN DEFAULT FALSE,
                matricule_a VARCHAR,
                matricule_b VARCHAR,
                nom_norm_a VARCHAR,
                nom_norm_b VARCHAR,
                nom_a VARCHAR,
                nom_b VARCHAR,
                prenom_a VARCHAR,
                prenom_b VARCHAR,
                regime_a VARCHAR,
                regime_b VARCHAR,
                institution_a VARCHAR,
                institution_b VARCHAR,
                occurrences_a BIGINT DEFAULT 0,
                occurrences_b BIGINT DEFAULT 0,
                brut_a DECIMAL(38,2) DEFAULT 0,
                brut_b DECIMAL(38,2) DEFAULT 0,
                net_a DECIMAL(38,2) DEFAULT 0,
                net_b DECIMAL(38,2) DEFAULT 0,
                ecart_brut DECIMAL(38,2) DEFAULT 0,
                ecart_net DECIMAL(38,2) DEFAULT 0,
                section_a VARCHAR,
                section_b VARCHAR,
                categorie_a VARCHAR,
                categorie_b VARCHAR,
                grade_a VARCHAR,
                grade_b VARCHAR,
                unite_a VARCHAR,
                unite_b VARCHAR,
                province_a VARCHAR,
                province_b VARCHAR,
                doublon_matricule_a BOOLEAN DEFAULT FALSE,
                doublon_matricule_b BOOLEAN DEFAULT FALSE,
                doublon_nom_a BOOLEAN DEFAULT FALSE,
                doublon_nom_b BOOLEAN DEFAULT FALSE,
                diagnostic VARCHAR
            )""")
            con.execute("CREATE INDEX IF NOT EXISTS idx_raw_period_cmp ON resultats_comparaison_raw_periode(comparaison_id,statut)")

    @staticmethod
    def _quote(name: str) -> str:
        return '"' + str(name).replace('"', '""') + '"'

    def list_raw_tables(self):
        with self.db.connect() as con:
            names = [r[0] for r in con.execute("""SELECT table_name FROM information_schema.tables
                WHERE table_schema='main' AND table_name LIKE 'raw_%'
                ORDER BY table_name""").fetchall()]
            out = []
            for name in names:
                count = int(con.execute(f"SELECT COUNT(*) FROM {self._quote(name)}").fetchone()[0])
                meta = con.execute("""SELECT regime,institution_id,trimestre,annee
                    FROM journal_executions WHERE table_destination=? AND type_operation='IMPORT_ACCESS'
                    ORDER BY date_debut DESC LIMIT 1""", [name]).fetchone()
                out.append((name,count,*(meta or ("","","",""))))
            return out

    def _executions(self, table_name: str, quarter: str, year: int):
        with self.db.connect() as con:
            return con.execute("""SELECT DISTINCT execution_id,institution_id,regime
                FROM journal_executions
                WHERE table_destination=? AND type_operation='IMPORT_ACCESS'
                  AND trimestre=? AND annee=?
                  AND statut IN ('TERMINE','TERMINE_AVEC_AVERTISSEMENTS')
                  AND execution_id IS NOT NULL""", [table_name, quarter, int(year)]).fetchall()

    def analyze(self, table_a: str, table_b: str, quarter: str, year: int, progress=None):
        quarter = str(quarter).upper().strip(); year = int(year)
        if table_a == table_b:
            raise ValueError("Sélectionnez deux tables RAW différentes.")
        if quarter not in {"T1","T2","T3","T4"}:
            raise ValueError("Trimestre invalide.")
        progress and progress(5, "Vérification des sources et de la période")
        ex_a = self._executions(table_a, quarter, year); ex_b = self._executions(table_b, quarter, year)
        if not ex_a: raise ValueError(f"Aucune exécution exploitable pour {table_a} en {quarter} {year}.")
        if not ex_b: raise ValueError(f"Aucune exécution exploitable pour {table_b} en {quarter} {year}.")
        cmp_id = str(uuid.uuid4())
        ids_a = [r[0] for r in ex_a]; ids_b = [r[0] for r in ex_b]
        pha = ",".join("?" for _ in ids_a); phb = ",".join("?" for _ in ids_b)
        progress and progress(20, "Préparation des agents A et B")
        with self.db.connect() as con:
            con.execute("BEGIN")
            try:
                con.execute("INSERT INTO comparaisons_raw_periode VALUES (?,?,?,?,?,'EN_COURS',CURRENT_TIMESTAMP,NULL,NULL)",
                            [cmp_id,table_a,table_b,quarter,year])
                for side, table, rows in (("A",table_a,ex_a),("B",table_b,ex_b)):
                    for execution_id,institution,regime in rows:
                        con.execute("INSERT INTO sources_comparaison_raw_periode VALUES (?,?,?,?,?,?)",
                                    [cmp_id,side,table,execution_id,institution,regime])
                progress and progress(40, "Matching par matricule et par nom")
                params = ids_a + [quarter,year] + ids_b + [quarter,year] + [cmp_id]
                con.execute(f"""INSERT INTO resultats_comparaison_raw_periode
                    WITH a0 AS (
                      SELECT matricule_normalise,nom_normalise,MIN(NULLIF(nom,'')) nom,MIN(NULLIF(prenom,'')) prenom,
                        STRING_AGG(DISTINCT COALESCE(regime,''), ', ') regime,
                        STRING_AGG(DISTINCT COALESCE(institution_id,''), ', ') institution,
                        COUNT(*) occ,SUM(COALESCE(remuneration_brute_calculee,0)) brut,SUM(COALESCE(montant_net,0)) net,
                        STRING_AGG(DISTINCT NULLIF(section,''), ', ') section,
                        STRING_AGG(DISTINCT NULLIF(categorie,''), ', ') categorie,
                        STRING_AGG(DISTINCT NULLIF(grade,''), ', ') grade,
                        STRING_AGG(DISTINCT NULLIF(unite_affectation,''), ', ') unite,
                        STRING_AGG(DISTINCT NULLIF(province,''), ', ') province
                      FROM paie_standardisee WHERE execution_id IN ({pha}) AND trimestre=? AND annee=?
                      GROUP BY matricule_normalise,nom_normalise
                    ), b0 AS (
                      SELECT matricule_normalise,nom_normalise,MIN(NULLIF(nom,'')) nom,MIN(NULLIF(prenom,'')) prenom,
                        STRING_AGG(DISTINCT COALESCE(regime,''), ', ') regime,
                        STRING_AGG(DISTINCT COALESCE(institution_id,''), ', ') institution,
                        COUNT(*) occ,SUM(COALESCE(remuneration_brute_calculee,0)) brut,SUM(COALESCE(montant_net,0)) net,
                        STRING_AGG(DISTINCT NULLIF(section,''), ', ') section,
                        STRING_AGG(DISTINCT NULLIF(categorie,''), ', ') categorie,
                        STRING_AGG(DISTINCT NULLIF(grade,''), ', ') grade,
                        STRING_AGG(DISTINCT NULLIF(unite_affectation,''), ', ') unite,
                        STRING_AGG(DISTINCT NULLIF(province,''), ', ') province
                      FROM paie_standardisee WHERE execution_id IN ({phb}) AND trimestre=? AND annee=?
                      GROUP BY matricule_normalise,nom_normalise
                    ), a_enriched AS (
                      SELECT a.*,
                        EXISTS(SELECT 1 FROM b0 b WHERE COALESCE(a.matricule_normalise,'') NOT IN ('','NU') AND b.matricule_normalise=a.matricule_normalise) m_mat,
                        EXISTS(SELECT 1 FROM b0 b WHERE COALESCE(a.nom_normalise,'')<>'' AND b.nom_normalise=a.nom_normalise) m_nom
                      FROM a0 a
                    ), matched_a AS (
                      SELECT a.*,
                        (SELECT MIN(b.matricule_normalise) FROM b0 b WHERE
                           (a.m_mat AND b.matricule_normalise=a.matricule_normalise)
                           OR (NOT a.m_mat AND a.m_nom AND b.nom_normalise=a.nom_normalise)) bmat,
                        (SELECT MIN(b.nom_normalise) FROM b0 b WHERE
                           (a.m_mat AND b.matricule_normalise=a.matricule_normalise)
                           OR (NOT a.m_mat AND a.m_nom AND b.nom_normalise=a.nom_normalise)) bnom
                      FROM a_enriched a
                    ), rows_a AS (
                      SELECT
                        CASE WHEN COALESCE(a.matricule_normalise,'') NOT IN ('','NU') THEN 'A:M:'||a.matricule_normalise ELSE 'A:N:'||COALESCE(a.nom_normalise,'') END cle,
                        CASE WHEN a.m_mat AND a.m_nom THEN 'COMMUN_PAR_MATRICULE_ET_NOM'
                             WHEN a.m_mat THEN 'COMMUN_PAR_MATRICULE'
                             WHEN a.m_nom THEN 'COMMUN_PAR_NOM'
                             ELSE 'UNIQUEMENT_A' END statut,
                        a.m_mat,a.m_nom,
                        a.m_mat AND COALESCE(a.nom_normalise,'')<>'' AND EXISTS(SELECT 1 FROM b0 x WHERE x.matricule_normalise=a.matricule_normalise AND COALESCE(x.nom_normalise,'')<>COALESCE(a.nom_normalise,'')) mat_nom_diff,
                        a.m_nom AND COALESCE(a.matricule_normalise,'') NOT IN ('','NU') AND EXISTS(SELECT 1 FROM b0 x WHERE x.nom_normalise=a.nom_normalise AND COALESCE(x.matricule_normalise,'')<>COALESCE(a.matricule_normalise,'')) nom_mat_diff,
                        a.*,
                        b.matricule_normalise b_matricule,b.nom_normalise b_nom_norm,b.nom b_nom,b.prenom b_prenom,b.regime b_regime,b.institution b_institution,
                        b.occ b_occ,b.brut b_brut,b.net b_net,b.section b_section,b.categorie b_categorie,b.grade b_grade,b.unite b_unite,b.province b_province
                      FROM matched_a a LEFT JOIN b0 b ON b.matricule_normalise=a.bmat AND b.nom_normalise=a.bnom
                    ), rows_b_only AS (
                      SELECT 'B:'||COALESCE(NULLIF(b.matricule_normalise,''),'N:'||COALESCE(b.nom_normalise,'')) cle,
                        'UNIQUEMENT_B' statut,FALSE m_mat,FALSE m_nom,FALSE mat_nom_diff,FALSE nom_mat_diff,
                        NULL matricule_normalise,NULL nom_normalise,NULL nom,NULL prenom,NULL regime,NULL institution,0 occ,0 brut,0 net,
                        NULL section,NULL categorie,NULL grade,NULL unite,NULL province,
                        b.matricule_normalise b_matricule,b.nom_normalise b_nom_norm,b.nom b_nom,b.prenom b_prenom,b.regime b_regime,b.institution b_institution,
                        b.occ b_occ,b.brut b_brut,b.net b_net,b.section b_section,b.categorie b_categorie,b.grade b_grade,b.unite b_unite,b.province b_province
                      FROM b0 b WHERE NOT EXISTS (
                        SELECT 1 FROM a0 a WHERE
                          (COALESCE(a.matricule_normalise,'') NOT IN ('','NU') AND a.matricule_normalise=b.matricule_normalise)
                          OR (COALESCE(a.nom_normalise,'')<>'' AND a.nom_normalise=b.nom_normalise)
                      )
                    ), u AS (SELECT * FROM rows_a UNION ALL SELECT * FROM rows_b_only)
                    SELECT ?,cle,statut,m_mat,m_nom,mat_nom_diff,nom_mat_diff,
                      matricule_normalise,b_matricule,nom_normalise,b_nom_norm,nom,b_nom,prenom,b_prenom,regime,b_regime,institution,b_institution,
                      occ,b_occ,brut,b_brut,net,b_net,COALESCE(brut,0)-COALESCE(b_brut,0),COALESCE(net,0)-COALESCE(b_net,0),
                      section,b_section,categorie,b_categorie,grade,b_grade,unite,b_unite,province,b_province,
                      occ>1,b_occ>1,
                      (SELECT COUNT(*) FROM a0 ax WHERE ax.nom_normalise=u.nom_normalise)>1,
                      (SELECT COUNT(*) FROM b0 bx WHERE bx.nom_normalise=u.b_nom_norm)>1,
                      TRIM(CONCAT_WS(' ; ',
                        CASE WHEN m_mat THEN 'Présent dans A et B par matricule' END,
                        CASE WHEN m_nom THEN 'Présent dans A et B par nom' END,
                        CASE WHEN mat_nom_diff THEN 'Même matricule avec nom différent' END,
                        CASE WHEN nom_mat_diff THEN 'Même nom avec matricule différent' END,
                        CASE WHEN COALESCE(brut,0)<>COALESCE(b_brut,0) AND (m_mat OR m_nom) THEN 'Écart brut' END,
                        CASE WHEN COALESCE(section,'')<>COALESCE(b_section,'') AND (m_mat OR m_nom) THEN 'Section différente' END,
                        CASE WHEN COALESCE(unite,'')<>COALESCE(b_unite,'') AND (m_mat OR m_nom) THEN 'Unité d’affectation différente' END
                      ))
                    FROM u""", params)
                progress and progress(80, "Calcul des écarts et doublons")
                con.execute("UPDATE comparaisons_raw_periode SET statut='TERMINE',termine_le=CURRENT_TIMESTAMP WHERE comparaison_id=?", [cmp_id])
                con.execute("COMMIT")
            except Exception:
                con.execute("ROLLBACK"); raise
        progress and progress(100, "Comparaison RAW terminée")
        return self.get_comparison(cmp_id)

    def reanalyze(self, comparison_id: str, progress=None):
        info = self.get_comparison(comparison_id)
        progress and progress(5, "Préparation de la réanalyse")
        with self.db.connect() as con:
            con.execute("DELETE FROM resultats_comparaison_raw_periode WHERE comparaison_id=?", [comparison_id])
            con.execute("DELETE FROM sources_comparaison_raw_periode WHERE comparaison_id=?", [comparison_id])
            con.execute("DELETE FROM comparaisons_raw_periode WHERE comparaison_id=?", [comparison_id])
        return self.analyze(info['table_a'], info['table_b'], info['quarter'], info['year'], progress=progress)

    def get_comparison(self, comparison_id: str):
        with self.db.connect() as con:
            row = con.execute("SELECT comparaison_id,table_a,table_b,trimestre,annee,statut,cree_le,termine_le,dossier_export FROM comparaisons_raw_periode WHERE comparaison_id=?", [comparison_id]).fetchone()
        if not row: raise ValueError("Comparaison RAW introuvable.")
        keys = ['id','table_a','table_b','quarter','year','status','created','finished','export']
        return dict(zip(keys,row))

    def list_history(self, limit=100):
        with self.db.connect() as con:
            return con.execute("SELECT comparaison_id,table_a,table_b,trimestre,annee,statut,cree_le,dossier_export FROM comparaisons_raw_periode ORDER BY cree_le DESC LIMIT ?", [max(1,min(int(limit),500))]).fetchall()

    def list_results(self, comparison_id: str, status="", limit=3000):
        cond="comparaison_id=?"; params=[comparison_id]
        special={
            'MEME_MATRICULE_NOM_DIFFERENT':'meme_matricule_nom_different',
            'MEME_NOM_MATRICULE_DIFFERENT':'meme_nom_matricule_different',
            'DOUBLON_MATRICULE_A':'doublon_matricule_a','DOUBLON_MATRICULE_B':'doublon_matricule_b',
            'DOUBLON_NOM_A':'doublon_nom_a','DOUBLON_NOM_B':'doublon_nom_b',
        }
        if status in special: cond += f" AND {special[status]}"
        elif status: cond += " AND statut=?"; params.append(status)
        params.append(max(1,min(int(limit),10000)))
        with self.db.connect() as con:
            return con.execute(f"""SELECT statut,matricule_a,matricule_b,nom_a,nom_b,prenom_a,prenom_b,
                commun_matricule,commun_nom,regime_a,regime_b,institution_a,institution_b,occurrences_a,occurrences_b,
                brut_a,brut_b,ecart_brut,net_a,net_b,ecart_net,section_a,section_b,categorie_a,categorie_b,
                grade_a,grade_b,unite_a,unite_b,province_a,province_b,diagnostic
                FROM resultats_comparaison_raw_periode WHERE {cond}
                ORDER BY commun_matricule DESC,commun_nom DESC,ABS(ecart_brut) DESC LIMIT ?""",params).fetchall()

    def summary(self, comparison_id: str):
        with self.db.connect() as con:
            base=con.execute("SELECT statut,COUNT(*),SUM(brut_a),SUM(brut_b),SUM(net_a),SUM(net_b) FROM resultats_comparaison_raw_periode WHERE comparaison_id=? GROUP BY statut ORDER BY statut",[comparison_id]).fetchall()
            metrics=con.execute("""SELECT
                SUM(CASE WHEN commun_matricule THEN 1 ELSE 0 END),
                SUM(CASE WHEN commun_nom THEN 1 ELSE 0 END),
                SUM(CASE WHEN commun_matricule AND commun_nom THEN 1 ELSE 0 END),
                SUM(CASE WHEN meme_matricule_nom_different THEN 1 ELSE 0 END),
                SUM(CASE WHEN meme_nom_matricule_different THEN 1 ELSE 0 END)
                FROM resultats_comparaison_raw_periode WHERE comparaison_id=?""",[comparison_id]).fetchone()
        return base, metrics

    def delete(self, comparison_id: str):
        with self.db.connect() as con:
            con.execute("BEGIN")
            try:
                con.execute("DELETE FROM resultats_comparaison_raw_periode WHERE comparaison_id=?",[comparison_id])
                con.execute("DELETE FROM sources_comparaison_raw_periode WHERE comparaison_id=?",[comparison_id])
                con.execute("DELETE FROM comparaisons_raw_periode WHERE comparaison_id=?",[comparison_id])
                con.execute("COMMIT")
            except Exception:
                con.execute("ROLLBACK"); raise

    def _write_query_xlsx(self, path: Path, headers, rows):
        wb=Workbook(write_only=True); ws=wb.create_sheet("Résultats"); ws.append(list(headers))
        for row in rows: ws.append(list(sanitize_excel_row(row)))
        wb.save(path)

    def _export_raw_source(self, comparison_id: str, side: str, path: Path):
        with self.db.connect() as con:
            src=con.execute("SELECT table_source FROM sources_comparaison_raw_periode WHERE comparaison_id=? AND cote=? LIMIT 1",[comparison_id,side]).fetchone()
            ids=[r[0] for r in con.execute("SELECT execution_id FROM sources_comparaison_raw_periode WHERE comparaison_id=? AND cote=?",[comparison_id,side]).fetchall()]
            if not src: return
            table=src[0]; safe=self._quote(table); columns=[r[0] for r in con.execute(f"DESCRIBE {safe}").fetchall()]
            wb=Workbook(write_only=True); ws=wb.create_sheet(f"RAW_{side}"); ws.append(columns)
            if 'execution_id' in columns and ids:
                ph=','.join('?' for _ in ids); cur=con.execute(f"SELECT * FROM {safe} WHERE execution_id IN ({ph})",ids)
            else:
                cur=con.execute(f"SELECT * FROM {safe}")
            while True:
                chunk=cur.fetchmany(5000)
                if not chunk: break
                for row in chunk: ws.append(list(sanitize_excel_row(row)))
            wb.save(path)

    def export_all(self, comparison_id: str, parent_folder, progress=None):
        info=self.get_comparison(comparison_id)
        folder=Path(parent_folder)/f"comparaison_raw_{info['quarter']}_{info['year']}_{datetime.now():%Y%m%d_%H%M%S}"
        folder.mkdir(parents=True,exist_ok=True)
        progress and progress(5,"Création de la synthèse")
        base,metrics=self.summary(comparison_id)
        wb=Workbook();ws=wb.active;ws.title="Synthèse";ws.append(["Indicateur","Valeur"])
        for k,v in [("Table A",info['table_a']),("Table B",info['table_b']),("Période",f"{info['quarter']} {info['year']}"),
                    ("Communs par matricule",metrics[0] or 0),("Communs par nom",metrics[1] or 0),("Communs matricule + nom",metrics[2] or 0),
                    ("Même matricule / nom différent",metrics[3] or 0),("Même nom / matricule différent",metrics[4] or 0)]:ws.append([k,v])
        ws.append([]);ws.append(["Statut","Agents","Brut A","Brut B","Net A","Net B"])
        for r in base:ws.append(list(sanitize_excel_row(r)))
        wb.save(folder/'00_synthese.xlsx')
        headers=["Statut","Matricule A","Matricule B","Nom A","Nom B","Prénom A","Prénom B","Commun matricule","Commun nom","Régime A","Régime B","Institution A","Institution B","Occ A","Occ B","Brut A","Brut B","Écart brut","Net A","Net B","Écart net","Section A","Section B","Catégorie A","Catégorie B","Grade A","Grade B","Unité A","Unité B","Province A","Province B","Diagnostic"]
        exports=[('01_tous_resultats.xlsx',''),('02_communs_matricule.xlsx','COMMUN_PAR_MATRICULE'),('03_communs_nom.xlsx','COMMUN_PAR_NOM'),('04_communs_matricule_et_nom.xlsx','COMMUN_PAR_MATRICULE_ET_NOM'),('05_uniquement_A.xlsx','UNIQUEMENT_A'),('06_uniquement_B.xlsx','UNIQUEMENT_B'),('07_meme_matricule_nom_different.xlsx','MEME_MATRICULE_NOM_DIFFERENT'),('08_meme_nom_matricule_different.xlsx','MEME_NOM_MATRICULE_DIFFERENT'),('09_doublons_matricule_A.xlsx','DOUBLON_MATRICULE_A'),('10_doublons_matricule_B.xlsx','DOUBLON_MATRICULE_B'),('11_doublons_nom_A.xlsx','DOUBLON_NOM_A'),('12_doublons_nom_B.xlsx','DOUBLON_NOM_B')]
        for i,(fn,status) in enumerate(exports,1):
            progress and progress(10+int(60*i/len(exports)),f"Export {fn}")
            self._write_query_xlsx(folder/fn,headers,self.list_results(comparison_id,status,10000))
        progress and progress(75,"Annexe RAW A complète")
        self._export_raw_source(comparison_id,'A',folder/'13_annexe_RAW_A_complete.xlsx')
        progress and progress(88,"Annexe RAW B complète")
        self._export_raw_source(comparison_id,'B',folder/'14_annexe_RAW_B_complete.xlsx')
        with self.db.connect() as con: con.execute("UPDATE comparaisons_raw_periode SET dossier_export=? WHERE comparaison_id=?",[str(folder),comparison_id])
        progress and progress(100,"Export complet terminé")
        return str(folder)
