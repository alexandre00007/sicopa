from __future__ import annotations

import uuid
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .spreadsheet_utils import sanitize_excel_row


class RegimeComparisonService:
    """Compare deux périmètres de paie sans dépendre d'un déclaratif."""

    STATUSES = [
        "COMMUN_IDENTIQUE",
        "ECART_FINANCIER",
        "ECART_ADMINISTRATIF",
        "ECART_FINANCIER_ET_ADMIN",
        "PAIEMENT_MULTIPLE",
        "IDENTITE_INCOHERENTE",
        "UNIQUEMENT_REGIME_A",
        "UNIQUEMENT_REGIME_B",
    ]

    def __init__(self, db):
        self.db = db
        self.ensure_schema()

    def ensure_schema(self) -> None:
        with self.db.connect() as con:
            con.execute("""CREATE TABLE IF NOT EXISTS comparaisons_regimes (
                comparaison_id VARCHAR PRIMARY KEY,
                institution_a VARCHAR NOT NULL, regime_a VARCHAR NOT NULL,
                institution_b VARCHAR NOT NULL, regime_b VARCHAR NOT NULL,
                trimestre VARCHAR NOT NULL, annee INTEGER NOT NULL,
                seuil_montant DECIMAL(38,2) DEFAULT 0,
                seuil_pourcentage DOUBLE DEFAULT 0,
                statut VARCHAR NOT NULL,
                lignes_a BIGINT DEFAULT 0, lignes_b BIGINT DEFAULT 0,
                communs BIGINT DEFAULT 0, uniquement_a BIGINT DEFAULT 0,
                uniquement_b BIGINT DEFAULT 0, doubles BIGINT DEFAULT 0,
                ecarts_financiers BIGINT DEFAULT 0, ecarts_admin BIGINT DEFAULT 0,
                masse_a DECIMAL(38,2) DEFAULT 0, masse_b DECIMAL(38,2) DEFAULT 0,
                cree_le TIMESTAMP DEFAULT CURRENT_TIMESTAMP, termine_le TIMESTAMP,
                fichier_export VARCHAR
            )""")
            con.execute("""CREATE TABLE IF NOT EXISTS resultats_comparaison_regimes (
                resultat_id VARCHAR,
                comparaison_id VARCHAR,
                cle_type VARCHAR,
                cle_match VARCHAR,
                matricule_a VARCHAR, matricule_b VARCHAR,
                matricule_normalise_a VARCHAR, matricule_normalise_b VARCHAR,
                nom_a VARCHAR, nom_b VARCHAR,
                nom_normalise_a VARCHAR, nom_normalise_b VARCHAR,
                occurrences_a BIGINT DEFAULT 0, occurrences_b BIGINT DEFAULT 0,
                remuneration_a DECIMAL(38,2) DEFAULT 0,
                remuneration_b DECIMAL(38,2) DEFAULT 0,
                net_a DECIMAL(38,2) DEFAULT 0,
                net_b DECIMAL(38,2) DEFAULT 0,
                ecart_remuneration DECIMAL(38,2) DEFAULT 0,
                ecart_net DECIMAL(38,2) DEFAULT 0,
                ecart_pourcentage DOUBLE DEFAULT 0,
                grade_a VARCHAR, grade_b VARCHAR,
                categorie_a VARCHAR, categorie_b VARCHAR,
                affectation_a VARCHAR, affectation_b VARCHAR,
                province_a VARCHAR, province_b VARCHAR,
                double_paiement BOOLEAN DEFAULT FALSE,
                ecart_financier BOOLEAN DEFAULT FALSE,
                ecart_administratif BOOLEAN DEFAULT FALSE,
                statut VARCHAR,
                diagnostic VARCHAR
            )""")
            con.execute("CREATE INDEX IF NOT EXISTS idx_cmp_regime_resultats ON resultats_comparaison_regimes(comparaison_id,statut)")

    def available_count(self, institution_id: str, regime: str, quarter: str, year: int) -> int:
        with self.db.connect() as con:
            return int(con.execute(
                "SELECT COUNT(*) FROM paie_standardisee WHERE institution_id=? AND regime=? AND trimestre=? AND annee=?",
                [institution_id, regime, quarter, int(year)],
            ).fetchone()[0])

    def run(self, institution_a: str, regime_a: str, institution_b: str, regime_b: str,
            quarter: str, year: int, threshold_amount: float = 0,
            threshold_percent: float = 0, progress=None) -> dict:
        if not all([institution_a, regime_a, institution_b, regime_b, quarter]):
            raise ValueError("Institution, régime et période sont obligatoires pour les deux côtés.")
        year = int(year)
        threshold_amount = max(0.0, float(threshold_amount or 0))
        threshold_percent = max(0.0, float(threshold_percent or 0))
        if institution_a == institution_b and regime_a == regime_b:
            raise ValueError("Choisissez deux périmètres différents à comparer.")

        rows_a = self.available_count(institution_a, regime_a, quarter, year)
        rows_b = self.available_count(institution_b, regime_b, quarter, year)
        if not rows_a:
            raise ValueError("Aucune donnée de paie n'est disponible pour le régime A.")
        if not rows_b:
            raise ValueError("Aucune donnée de paie n'est disponible pour le régime B.")

        comparison_id = str(uuid.uuid4())
        progress and progress(5, "Préparation de la comparaison des régimes")
        with self.db.connect() as con:
            con.execute("""INSERT INTO comparaisons_regimes
                (comparaison_id,institution_a,regime_a,institution_b,regime_b,trimestre,annee,
                 seuil_montant,seuil_pourcentage,statut,lignes_a,lignes_b)
                VALUES (?,?,?,?,?,?,?,?,?,'EN_COURS',?,?)""",
                [comparison_id, institution_a, regime_a, institution_b, regime_b, quarter, year,
                 threshold_amount, threshold_percent, rows_a, rows_b])
            con.execute("BEGIN")
            try:
                progress and progress(20, "Agrégation des agents et des masses financières")
                con.execute("""CREATE OR REPLACE TEMP TABLE cmp_a AS
                    SELECT
                        CASE WHEN COALESCE(matricule_normalise,'') NOT IN ('','NU')
                             THEN 'M:' || matricule_normalise ELSE 'N:' || COALESCE(nom_normalise,'') END cle_match,
                        CASE WHEN COALESCE(matricule_normalise,'') NOT IN ('','NU') THEN 'MATRICULE' ELSE 'NOM' END cle_type,
                        MIN(matricule_source) matricule_source,
                        MIN(matricule_normalise) matricule_normalise,
                        MIN(TRIM(COALESCE(nom,'') || ' ' || COALESCE(prenom,''))) nom_complet,
                        MIN(nom_normalise) nom_normalise,
                        COUNT(*) occurrences,
                        SUM(COALESCE(remuneration_brute_calculee,0)) remuneration,
                        SUM(COALESCE(montant_net,0)) net,
                        MIN(COALESCE(grade,'')) grade,
                        MIN(COALESCE(categorie,'')) categorie,
                        MIN(COALESCE(unite_affectation,'')) affectation,
                        MIN(COALESCE(province,'')) province
                    FROM paie_standardisee
                    WHERE institution_id=? AND regime=? AND trimestre=? AND annee=?
                      AND (COALESCE(matricule_normalise,'') NOT IN ('','NU') OR COALESCE(nom_normalise,'')<>'')
                    GROUP BY 1,2""", [institution_a, regime_a, quarter, year])
                con.execute("""CREATE OR REPLACE TEMP TABLE cmp_b AS
                    SELECT
                        CASE WHEN COALESCE(matricule_normalise,'') NOT IN ('','NU')
                             THEN 'M:' || matricule_normalise ELSE 'N:' || COALESCE(nom_normalise,'') END cle_match,
                        CASE WHEN COALESCE(matricule_normalise,'') NOT IN ('','NU') THEN 'MATRICULE' ELSE 'NOM' END cle_type,
                        MIN(matricule_source) matricule_source,
                        MIN(matricule_normalise) matricule_normalise,
                        MIN(TRIM(COALESCE(nom,'') || ' ' || COALESCE(prenom,''))) nom_complet,
                        MIN(nom_normalise) nom_normalise,
                        COUNT(*) occurrences,
                        SUM(COALESCE(remuneration_brute_calculee,0)) remuneration,
                        SUM(COALESCE(montant_net,0)) net,
                        MIN(COALESCE(grade,'')) grade,
                        MIN(COALESCE(categorie,'')) categorie,
                        MIN(COALESCE(unite_affectation,'')) affectation,
                        MIN(COALESCE(province,'')) province
                    FROM paie_standardisee
                    WHERE institution_id=? AND regime=? AND trimestre=? AND annee=?
                      AND (COALESCE(matricule_normalise,'') NOT IN ('','NU') OR COALESCE(nom_normalise,'')<>'')
                    GROUP BY 1,2""", [institution_b, regime_b, quarter, year])

                progress and progress(45, "Rapprochement par matricule puis par nom normalisé")
                con.execute("""INSERT INTO resultats_comparaison_regimes
                    SELECT
                        uuid(), ?, COALESCE(a.cle_type,b.cle_type), COALESCE(a.cle_match,b.cle_match),
                        a.matricule_source,b.matricule_source,a.matricule_normalise,b.matricule_normalise,
                        a.nom_complet,b.nom_complet,a.nom_normalise,b.nom_normalise,
                        COALESCE(a.occurrences,0),COALESCE(b.occurrences,0),
                        COALESCE(a.remuneration,0),COALESCE(b.remuneration,0),
                        COALESCE(a.net,0),COALESCE(b.net,0),
                        COALESCE(a.remuneration,0)-COALESCE(b.remuneration,0),
                        COALESCE(a.net,0)-COALESCE(b.net,0),
                        CASE WHEN GREATEST(ABS(COALESCE(a.remuneration,0)),ABS(COALESCE(b.remuneration,0)),1)=0 THEN 0
                             ELSE 100.0*ABS(COALESCE(a.remuneration,0)-COALESCE(b.remuneration,0)) /
                                  GREATEST(ABS(COALESCE(a.remuneration,0)),ABS(COALESCE(b.remuneration,0)),1) END,
                        a.grade,b.grade,a.categorie,b.categorie,a.affectation,b.affectation,a.province,b.province,
                        a.cle_match IS NOT NULL AND b.cle_match IS NOT NULL,
                        FALSE,FALSE,'EN_COURS',''
                    FROM cmp_a a FULL OUTER JOIN cmp_b b ON a.cle_match=b.cle_match""", [comparison_id])

                progress and progress(65, "Classification des écarts financiers et administratifs")
                con.execute("""UPDATE resultats_comparaison_regimes SET
                    ecart_financier = CASE
                        WHEN occurrences_a=0 OR occurrences_b=0 THEN FALSE
                        ELSE ABS(ecart_remuneration)>=? OR ABS(ecart_net)>=? OR ABS(ecart_pourcentage)>=? END,
                    ecart_administratif = CASE
                        WHEN occurrences_a=0 OR occurrences_b=0 THEN FALSE
                        ELSE COALESCE(grade_a,'')<>COALESCE(grade_b,'')
                          OR COALESCE(categorie_a,'')<>COALESCE(categorie_b,'')
                          OR COALESCE(affectation_a,'')<>COALESCE(affectation_b,'')
                          OR COALESCE(province_a,'')<>COALESCE(province_b,'') END
                    WHERE comparaison_id=?""",
                    [threshold_amount, threshold_amount, threshold_percent, comparison_id])

                con.execute("""UPDATE resultats_comparaison_regimes SET
                    statut=CASE
                        WHEN occurrences_a=0 THEN 'UNIQUEMENT_REGIME_B'
                        WHEN occurrences_b=0 THEN 'UNIQUEMENT_REGIME_A'
                        WHEN cle_type='MATRICULE' AND COALESCE(nom_normalise_a,'')<>'' AND COALESCE(nom_normalise_b,'')<>''
                             AND nom_normalise_a<>nom_normalise_b THEN 'IDENTITE_INCOHERENTE'
                        WHEN occurrences_a>1 OR occurrences_b>1 THEN 'PAIEMENT_MULTIPLE'
                        WHEN ecart_financier AND ecart_administratif THEN 'ECART_FINANCIER_ET_ADMIN'
                        WHEN ecart_financier THEN 'ECART_FINANCIER'
                        WHEN ecart_administratif THEN 'ECART_ADMINISTRATIF'
                        ELSE 'COMMUN_IDENTIQUE' END,
                    diagnostic=TRIM(CONCAT_WS(' ; ',
                        CASE WHEN occurrences_a>1 THEN 'Paiements multiples dans le régime A' END,
                        CASE WHEN occurrences_b>1 THEN 'Paiements multiples dans le régime B' END,
                        CASE WHEN ecart_financier THEN 'Écart financier supérieur au seuil' END,
                        CASE WHEN ecart_administratif THEN 'Informations administratives différentes' END,
                        CASE WHEN cle_type='MATRICULE' AND COALESCE(nom_normalise_a,'')<>'' AND COALESCE(nom_normalise_b,'')<>''
                                  AND nom_normalise_a<>nom_normalise_b THEN 'Même matricule associé à des identités différentes' END,
                        CASE WHEN double_paiement THEN 'Agent payé dans les deux régimes' END
                    ))
                    WHERE comparaison_id=?""", [comparison_id])

                # Signale les noms communs portant des matricules différents dans les deux régimes.
                con.execute("""UPDATE resultats_comparaison_regimes r SET diagnostic=TRIM(CONCAT_WS(' ; ',
                        NULLIF(r.diagnostic,''),'Même nom retrouvé avec un matricule différent dans l’autre régime'))
                    WHERE r.comparaison_id=? AND r.statut IN ('UNIQUEMENT_REGIME_A','UNIQUEMENT_REGIME_B')
                      AND COALESCE(CASE WHEN r.occurrences_a>0 THEN r.nom_normalise_a ELSE r.nom_normalise_b END,'')<>''
                      AND EXISTS (
                        SELECT 1 FROM resultats_comparaison_regimes x
                        WHERE x.comparaison_id=r.comparaison_id AND x.resultat_id<>r.resultat_id
                          AND COALESCE(CASE WHEN x.occurrences_a>0 THEN x.nom_normalise_a ELSE x.nom_normalise_b END,'')=
                              COALESCE(CASE WHEN r.occurrences_a>0 THEN r.nom_normalise_a ELSE r.nom_normalise_b END,'')
                          AND ((r.occurrences_a>0 AND x.occurrences_b>0) OR (r.occurrences_b>0 AND x.occurrences_a>0))
                    )""", [comparison_id])

                summary = con.execute("""SELECT
                    SUM(CASE WHEN occurrences_a>0 AND occurrences_b>0 THEN 1 ELSE 0 END),
                    SUM(CASE WHEN occurrences_a>0 AND occurrences_b=0 THEN 1 ELSE 0 END),
                    SUM(CASE WHEN occurrences_b>0 AND occurrences_a=0 THEN 1 ELSE 0 END),
                    SUM(CASE WHEN double_paiement THEN 1 ELSE 0 END),
                    SUM(CASE WHEN ecart_financier THEN 1 ELSE 0 END),
                    SUM(CASE WHEN ecart_administratif THEN 1 ELSE 0 END),
                    SUM(remuneration_a),SUM(remuneration_b)
                    FROM resultats_comparaison_regimes WHERE comparaison_id=?""", [comparison_id]).fetchone()
                con.execute("""UPDATE comparaisons_regimes SET statut='TERMINE',communs=?,uniquement_a=?,uniquement_b=?,
                    doubles=?,ecarts_financiers=?,ecarts_admin=?,masse_a=?,masse_b=?,termine_le=CURRENT_TIMESTAMP
                    WHERE comparaison_id=?""", [*(int(value or 0) for value in summary[:6]), summary[6] or 0, summary[7] or 0, comparison_id])
                con.execute("COMMIT")
            except Exception:
                con.execute("ROLLBACK")
                con.execute("UPDATE comparaisons_regimes SET statut='ERREUR',termine_le=CURRENT_TIMESTAMP WHERE comparaison_id=?", [comparison_id])
                raise

        progress and progress(100, "Comparaison des régimes terminée")
        return self.get_summary(comparison_id)

    def get_summary(self, comparison_id: str) -> dict:
        with self.db.connect() as con:
            row = con.execute("""SELECT comparaison_id,institution_a,regime_a,institution_b,regime_b,trimestre,annee,
                    seuil_montant,seuil_pourcentage,statut,lignes_a,lignes_b,communs,uniquement_a,uniquement_b,
                    doubles,ecarts_financiers,ecarts_admin,masse_a,masse_b,cree_le,termine_le,fichier_export
                FROM comparaisons_regimes WHERE comparaison_id=?""", [comparison_id]).fetchone()
        if not row:
            raise ValueError("Comparaison introuvable.")
        keys = ["id","institution_a","regime_a","institution_b","regime_b","quarter","year","threshold_amount",
                "threshold_percent","status","rows_a","rows_b","common","only_a","only_b","double","financial",
                "administrative","mass_a","mass_b","created","finished","export"]
        return dict(zip(keys,row))

    def list_results(self, comparison_id: str, status: str = "", limit: int = 2000) -> list[tuple]:
        params = [comparison_id]
        condition = "comparaison_id=?"
        if status:
            if status == "DOUBLE_PAIEMENT":
                condition += " AND double_paiement"
            else:
                condition += " AND statut=?"
                params.append(status)
        params.append(max(1, min(int(limit), 10000)))
        with self.db.connect() as con:
            return con.execute(f"""SELECT statut,cle_type,COALESCE(matricule_a,matricule_b,''),
                    COALESCE(NULLIF(nom_a,''),nom_b,''),occurrences_a,occurrences_b,
                    remuneration_a,remuneration_b,ecart_remuneration,net_a,net_b,ecart_net,ecart_pourcentage,
                    COALESCE(grade_a,''),COALESCE(grade_b,''),COALESCE(categorie_a,''),COALESCE(categorie_b,''),
                    COALESCE(affectation_a,''),COALESCE(affectation_b,''),diagnostic
                FROM resultats_comparaison_regimes WHERE {condition}
                ORDER BY CASE WHEN statut='COMMUN_IDENTIQUE' THEN 1 ELSE 0 END,ABS(ecart_remuneration) DESC,nom_a,nom_b
                LIMIT ?""", params).fetchall()

    def list_history(self, limit: int = 100) -> list[tuple]:
        with self.db.connect() as con:
            return con.execute("""SELECT c.comparaison_id,COALESCE(ia.nom_officiel,c.institution_a),c.regime_a,
                    COALESCE(ib.nom_officiel,c.institution_b),c.regime_b,c.trimestre,c.annee,c.statut,
                    c.communs,c.uniquement_a,c.uniquement_b,c.ecarts_financiers,c.cree_le
                FROM comparaisons_regimes c
                LEFT JOIN institutions ia ON ia.institution_id=c.institution_a
                LEFT JOIN institutions ib ON ib.institution_id=c.institution_b
                ORDER BY c.cree_le DESC LIMIT ?""", [max(1,min(int(limit),500))]).fetchall()

    def export(self, comparison_id: str, path: str) -> str:
        summary = self.get_summary(comparison_id)
        target = Path(path)
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")
        target.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Synthèse"
        summary_rows = [
            ("Comparaison", f"{summary['regime_a']} vs {summary['regime_b']}"),
            ("Période", f"{summary['quarter']} {summary['year']}"),
            ("Lignes régime A", summary["rows_a"]), ("Lignes régime B", summary["rows_b"]),
            ("Agents communs / payés dans les deux", summary["common"]),
            ("Uniquement régime A", summary["only_a"]), ("Uniquement régime B", summary["only_b"]),
            ("Écarts financiers", summary["financial"]), ("Écarts administratifs", summary["administrative"]),
            ("Masse régime A", summary["mass_a"]), ("Masse régime B", summary["mass_b"]),
            ("Écart de masse", (summary["mass_a"] or 0) - (summary["mass_b"] or 0)),
        ]
        ws.append(["Indicateur", "Valeur"])
        for cell in ws[1]: cell.font = Font(bold=True)
        for row in summary_rows: ws.append(list(sanitize_excel_row(row)))

        headers = ["Statut","Clé","Matricule","Nom","Occurrences A","Occurrences B","Brut A","Brut B",
                   "Écart brut","Net A","Net B","Écart net","Écart %","Grade A","Grade B","Catégorie A",
                   "Catégorie B","Affectation A","Affectation B","Diagnostic"]
        sheets = [
            ("Tous les résultats", ""),
            ("Payés dans les deux", "DOUBLE_PAIEMENT"),
            ("Uniquement A", "UNIQUEMENT_REGIME_A"),
            ("Uniquement B", "UNIQUEMENT_REGIME_B"),
            ("Écarts financiers", "ECART_FINANCIER"),
            ("Écarts fin+admin", "ECART_FINANCIER_ET_ADMIN"),
            ("Écarts administratifs", "ECART_ADMINISTRATIF"),
            ("Paiements multiples", "PAIEMENT_MULTIPLE"),
            ("Identités incohérentes", "IDENTITE_INCOHERENTE"),
        ]
        for title, status in sheets:
            sheet = wb.create_sheet(title[:31])
            sheet.append(headers)
            for cell in sheet[1]: cell.font = Font(bold=True)
            for row in self.list_results(comparison_id, status, 10000):
                sheet.append(list(sanitize_excel_row(row)))
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

        wb.save(target)
        with self.db.connect() as con:
            con.execute("UPDATE comparaisons_regimes SET fichier_export=? WHERE comparaison_id=?", [str(target), comparison_id])
        return str(target)
