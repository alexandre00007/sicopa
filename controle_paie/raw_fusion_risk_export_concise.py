from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from .export_streaming import atomic_save_workbook
from .raw_fusion_export_partitioned import (
    _WorkbookPartitionWriter,
    _cleanup_compact_stats,
    _configure_partitioned,
    _fusion_executions,
    _fusion_scope,
    _person_key_sql,
    _prepare_compact_identity_stats,
    PartitionedRawFusionRiskExporter,
)
from .raw_fusion_risk_export import RISK_DETAIL_HEADERS


GLOBAL_HEADERS = [
    "Type anomalie", "Agents concernes", "Occurrences physiques", "Repetitions",
    "Max regimes par agent", "Masse brute concernee", "Masse nette concernee", "Lecture",
]

MATRICULE_HEADERS = [
    "Matricule", "Noms observes", "Occurrences", "Repetitions", "Nb regimes",
    "Regimes", "Nb executions", "Nb tables RAW", "Masse brute", "Masse nette",
    "Identite incoherente", "Diagnostic",
]

NOM_HEADERS = [
    "Nom normalise", "Matricules observes", "Occurrences", "Repetitions", "Nb regimes max",
    "Regimes observes", "Nb executions max", "Nb tables RAW max", "Masse brute", "Masse nette",
    "Doublon nom", "Diagnostic",
]

MULTI_HEADERS = [
    "Matricule", "Nom", "Occurrences", "Repetitions", "Nb regimes", "Regimes",
    "Institutions", "Nb executions", "Nb tables RAW", "Masse brute", "Masse nette", "Diagnostic",
]


class ConcisePartitionedRawFusionRiskExporter(PartitionedRawFusionRiskExporter):
    """Annexe 12 orientee audit : synthese concise, puis occurrences detaillees.

    Contrairement a l'annexe 11 qui est exhaustive, cette annexe ne reprend que les
    situations a investiguer. Les familles d'anomalies ne sont pas exclusives : un agent
    peut apparaitre dans plusieurs syntheses/detailles lorsqu'il cumule plusieurs risques.
    """

    @staticmethod
    def _valid_matricule() -> str:
        return """COALESCE(NULLIF(TRIM(r.matricule_normalise),''),'')<>''
            AND UPPER(TRIM(COALESCE(r.matricule_normalise,''))) NOT IN
                ('NU','NULL','N/A','NA','NEANT','NÉANT','INCONNU','NONE')"""

    @staticmethod
    def _invalid_matricule_stats() -> str:
        return """(COALESCE(s.has_matricule_vide,FALSE)
            OR COALESCE(s.has_matricule_nu,FALSE)
            OR COALESCE(s.has_matricule_non_exploitable,FALSE))"""

    def _lens_predicates(self) -> list[tuple[str, str, str]]:
        valid = self._valid_matricule()
        invalid = self._invalid_matricule_stats()
        return [
            (
                "ANOMALIE_PAR_MATRICULE",
                f"({valid}) AND (COALESCE(r.occurrences,0)>1 OR COALESCE(d.doublon_matricule,FALSE) OR COALESCE(r.identite_incoherente,FALSE))",
                "Matricule retrouve plusieurs fois, partage ou associe a plusieurs identites.",
            ),
            (
                "ANOMALIE_PAR_NOM",
                "COALESCE(NULLIF(TRIM(r.nom_normalise),''),'')<>'' AND (COALESCE(d.doublon_nom,FALSE) OR COALESCE(r.occurrences,0)>1)",
                "Nom retrouve sur plusieurs occurrences; verifier les matricules et regimes associes.",
            ),
            (
                "ANOMALIE_MATRICULE_ET_NOM",
                "COALESCE(d.doublon_matricule,FALSE) AND COALESCE(d.doublon_nom,FALSE)",
                "Matricule et nom se repetent simultanement.",
            ),
            (
                "MATRICULE_NU_NULL_VIDE",
                invalid,
                "Matricule absent ou non exploitable; l'identification repose principalement sur le nom.",
            ),
            (
                "MULTI_REGIME_REPETE",
                "COALESCE(r.nb_regimes,0)>1 AND COALESCE(r.occurrences,0)>1",
                "Agent present plusieurs fois dans au moins deux regimes.",
            ),
            (
                "IDENTITE_INCOHERENTE",
                "COALESCE(r.identite_incoherente,FALSE) OR r.statut='MATRICULE_PARTAGE_IDENTITES_DIFFERENTES'",
                "Un meme matricule renvoie vers plusieurs identites ou noms.",
            ),
            (
                "PAIEMENT_MULTIPLE_MEME_REGIME",
                "COALESCE(r.paiement_multiple_meme_regime,FALSE)",
                "Plusieurs lignes de paiement sont observees dans un meme regime.",
            ),
            (
                "PLUSIEURS_INSTITUTIONS",
                "COALESCE(r.nb_institutions,0)>1",
                "Agent retrouve dans plusieurs institutions sur la periode.",
            ),
        ]

    def _global_summary_row(self, con, fusion_id: str, label: str, predicate: str, reading: str):
        row = con.execute(f"""SELECT COUNT(*),COALESCE(SUM(r.occurrences),0),
            COALESCE(SUM(GREATEST(r.occurrences-1,0)),0),COALESCE(MAX(r.nb_regimes),0),
            COALESCE(SUM(r.masse_brute),0),COALESCE(SUM(r.masse_net),0)
            FROM resultats_fusion_multi r
            JOIN tmp_sicorpa_identity_stats s ON s.person_key=r.person_key
            LEFT JOIN resultats_fusion_doublons d
              ON d.fusion_id=r.fusion_id AND d.person_key=r.person_key
            WHERE r.fusion_id=? AND ({predicate})""", [fusion_id]).fetchone()
        return [label, int(row[0] or 0), int(row[1] or 0), int(row[2] or 0),
                int(row[3] or 0), float(row[4] or 0), float(row[5] or 0), reading]

    def _matricule_summary_query(self) -> str:
        valid = self._valid_matricule()
        return f"""SELECT r.matricule_normalise,
            STRING_AGG(DISTINCT NULLIF(r.nom_normalise,''),' | '),
            COALESCE(SUM(r.occurrences),0),COALESCE(SUM(GREATEST(r.occurrences-1,0)),0),
            COALESCE(MAX(r.nb_regimes),0),STRING_AGG(DISTINCT NULLIF(r.regimes,''),' | '),
            COALESCE(MAX(s.nb_executions),0),COALESCE(MAX(s.nb_tables),0),
            COALESCE(SUM(r.masse_brute),0),COALESCE(SUM(r.masse_net),0),
            BOOL_OR(COALESCE(r.identite_incoherente,FALSE)),STRING_AGG(DISTINCT NULLIF(r.diagnostic,''),' | ')
        FROM resultats_fusion_multi r
        JOIN tmp_sicorpa_identity_stats s ON s.person_key=r.person_key
        LEFT JOIN resultats_fusion_doublons d ON d.fusion_id=r.fusion_id AND d.person_key=r.person_key
        WHERE r.fusion_id=? AND ({valid})
          AND (COALESCE(r.occurrences,0)>1 OR COALESCE(d.doublon_matricule,FALSE) OR COALESCE(r.identite_incoherente,FALSE))
        GROUP BY r.matricule_normalise
        ORDER BY COALESCE(SUM(r.occurrences),0) DESC,r.matricule_normalise"""

    def _nom_summary_query(self) -> str:
        return """SELECT r.nom_normalise,
            STRING_AGG(DISTINCT NULLIF(r.matricule_normalise,''),' | '),
            COALESCE(SUM(r.occurrences),0),COALESCE(SUM(GREATEST(r.occurrences-1,0)),0),
            COALESCE(MAX(r.nb_regimes),0),STRING_AGG(DISTINCT NULLIF(r.regimes,''),' | '),
            COALESCE(MAX(s.nb_executions),0),COALESCE(MAX(s.nb_tables),0),
            COALESCE(SUM(r.masse_brute),0),COALESCE(SUM(r.masse_net),0),
            BOOL_OR(COALESCE(d.doublon_nom,FALSE)),STRING_AGG(DISTINCT NULLIF(r.diagnostic,''),' | ')
        FROM resultats_fusion_multi r
        JOIN tmp_sicorpa_identity_stats s ON s.person_key=r.person_key
        LEFT JOIN resultats_fusion_doublons d ON d.fusion_id=r.fusion_id AND d.person_key=r.person_key
        WHERE r.fusion_id=? AND COALESCE(NULLIF(TRIM(r.nom_normalise),''),'')<>''
          AND (COALESCE(d.doublon_nom,FALSE) OR COALESCE(r.occurrences,0)>1)
        GROUP BY r.nom_normalise
        ORDER BY COALESCE(SUM(r.occurrences),0) DESC,r.nom_normalise"""

    def _multi_summary_query(self) -> str:
        return """SELECT r.matricule_normalise,r.nom_normalise,r.occurrences,
            GREATEST(r.occurrences-1,0),r.nb_regimes,r.regimes,r.institutions,
            s.nb_executions,s.nb_tables,r.masse_brute,r.masse_net,r.diagnostic
        FROM resultats_fusion_multi r
        JOIN tmp_sicorpa_identity_stats s ON s.person_key=r.person_key
        WHERE r.fusion_id=? AND COALESCE(r.nb_regimes,0)>1 AND COALESCE(r.occurrences,0)>1
        ORDER BY r.occurrences DESC,r.nb_regimes DESC,r.matricule_normalise,r.nom_normalise"""

    def _execution_lens_detail(self, predicate: str) -> str:
        key = _person_key_sql("p")
        return f"""SELECT ?,?,p.execution_id,p.ligne_paie_id,p.ligne_source,p.regime,p.institution_id,
            p.trimestre,p.annee,p.matricule_source,p.matricule_normalise,p.nom,p.prenom,p.nom_normalise,
            p.section,p.categorie,p.grade,p.unite_affectation,p.province,
            p.remuneration_brute_calculee,p.montant_net,r.statut,r.nb_regimes,r.occurrences,
            GREATEST(r.occurrences-1,0),s.nb_executions,s.nb_tables,r.regimes,r.institutions,
            s.noms_distincts,s.matricules_distincts,COALESCE(d.doublon_matricule,FALSE),
            COALESCE(d.doublon_nom,FALSE),r.paiement_multi_regime,r.paiement_multiple_meme_regime,
            r.identite_incoherente,r.diagnostic
        FROM paie_standardisee p
        JOIN resultats_fusion_multi r ON r.fusion_id=? AND r.person_key=({key})
        JOIN tmp_sicorpa_identity_stats s ON s.person_key=r.person_key
        LEFT JOIN resultats_fusion_doublons d ON d.fusion_id=r.fusion_id AND d.person_key=r.person_key
        WHERE p.execution_id=? AND p.trimestre=? AND p.annee=? AND ({predicate})"""

    def export(self, fusion_id: str, folder: str | Path, progress=None) -> Path:
        folder = Path(folder)
        target = folder / "12_synthese_occurrences_agents_a_risque.xlsx"
        book = Workbook(write_only=True)

        lenses = self._lens_predicates()
        detail_lenses = [
            ("D01_Occ_matricule", lenses[0]),
            ("D02_Occ_nom", lenses[1]),
            ("D03_Occ_matricule_nom", lenses[2]),
            ("D04_Occ_NU_NULL", lenses[3]),
            ("D05_Occ_multi_regimes", lenses[4]),
            ("D06_Identites_incoh", lenses[5]),
            ("D07_Paiements_multiples", lenses[6]),
            ("D08_Plusieurs_instit", lenses[7]),
        ]

        with self.db.connect() as con:
            _configure_partitioned(con)
            quarter, year = _fusion_scope(con, fusion_id)
            executions = _fusion_executions(con, fusion_id)
            _prepare_compact_identity_stats(con, fusion_id, quarter, year, executions)
            try:
                # 00 - vue de direction : quelques lignes seulement.
                ws = book.create_sheet("00_Synthese_anomalies")
                ws.append(GLOBAL_HEADERS)
                global_rows = []
                for label, predicate, reading in lenses:
                    row = self._global_summary_row(con, fusion_id, label, predicate, reading)
                    ws.append(row)
                    global_rows.append(row)

                # 01/02/03 - syntheses metier prioritaires.
                progress and progress(96, "Annexe 12 : syntheses matricule, nom et multi-regimes")
                matricule_writer = _WorkbookPartitionWriter(book, MATRICULE_HEADERS, "01_Synthese_matricule")
                matricule_writer.append_cursor(con.execute(self._matricule_summary_query(), [fusion_id]), 1000)
                nom_writer = _WorkbookPartitionWriter(book, NOM_HEADERS, "02_Synthese_nom")
                nom_writer.append_cursor(con.execute(self._nom_summary_query(), [fusion_id]), 1000)
                multi_writer = _WorkbookPartitionWriter(book, MULTI_HEADERS, "03_Synthese_multi_regimes")
                multi_writer.append_cursor(con.execute(self._multi_summary_query(), [fusion_id]), 1000)

                # Details : lecture execution par execution pour conserver une memoire bornee.
                detail_counts = {}
                total_exec = max(1, len(executions))
                for lens_index, (sheet_name, (label, predicate, _reading)) in enumerate(detail_lenses, 1):
                    writer = _WorkbookPartitionWriter(book, RISK_DETAIL_HEADERS, sheet_name)
                    query = self._execution_lens_detail(predicate)
                    for exec_index, (execution_id, table_source) in enumerate(executions, 1):
                        progress and progress(
                            97 + int(exec_index / total_exec),
                            f"Annexe 12 : {label} - source {exec_index}/{total_exec}",
                        )
                        cursor = con.execute(
                            query,
                            [label, table_source or "", fusion_id, execution_id, quarter, year],
                        )
                        writer.append_cursor(cursor, 1000)
                    detail_counts[label] = writer.total

                all_agents = int(con.execute(
                    "SELECT COUNT(*) FROM resultats_fusion_multi WHERE fusion_id=?", [fusion_id]
                ).fetchone()[0] or 0)
                control = book.create_sheet("Controle")
                control.append(["Indicateur", "Valeur"])
                control.append(["Mode export", "PARTITIONNE_PAR_EXECUTION"])
                control.append(["Organisation", "SYNTHESE_CONCISE_PUIS_OCCURRENCES"])
                control.append(["Nature categories", "NON_EXCLUSIVES"])
                control.append(["Agents analyses au total", all_agents])
                control.append(["Executions traitees", len(executions)])
                control.append(["Synthese matricules", matricule_writer.total])
                control.append(["Synthese noms", nom_writer.total])
                control.append(["Synthese multi-regimes", multi_writer.total])
                for row in global_rows:
                    control.append([f"Agents {row[0]}", row[1]])
                    control.append([f"Occurrences {row[0]}", row[2]])
                    control.append([f"Lignes detail {row[0]}", detail_counts.get(row[0], 0)])
                control.append(["Controle", "OK"])
            finally:
                _cleanup_compact_stats(con)

        atomic_save_workbook(book, target)
        progress and progress(99, f"Fichier genere : {target.name}")
        return target
